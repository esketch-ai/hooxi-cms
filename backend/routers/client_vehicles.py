"""운수사 보유 차량(fleet) 관리 — 부록 M.

- 전역 엑셀 업로드(BUS_LIST_ALL): 업체명→운수사 매칭·차량번호 upsert·참여 링크 갱신
- 고객사 상세 보유 차량 조회: 감축사업 참여 구분(participation) + 대표 참여 정보
- fleet 단건 CRUD(수기)

감축사업 참여는 ProjectVehicle.client_vehicle_id가 이 마스터를 가리켜 표현한다(참여 구분).
"""

from datetime import date, datetime
from io import BytesIO
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

import schemas
from auth import get_current_user, require_permission
from models import Asset, Client, ClientVehicle, Project, ProjectVehicle, User, get_db
from routers import common
from routers.codes import validate_active_code
from services.audit_logger import AuditLogger

router = APIRouter(tags=["client_vehicles"])

_FLEET_SHEET = "BUS_LIST_ALL"

# fleet 마스터 수기 입력/수정 대상 컬럼 (region은 vehicle_no에서 파생, client_id는 path/매칭)
_FLEET_FIELDS = (
    "operator_name", "chassis_no", "model_name", "model_year", "registered_at",
    "vehicle_class", "length_mm", "width_mm", "height_mm", "gross_weight_kg",
    "seating_capacity", "fuel", "status", "asset_id", "memo",
)

# BUS_LIST_ALL 1행 헤더 라벨 → (필드, 타입). 순서 무관·라벨 기준 매핑.
_HEADER_MAP = {
    "차량번호": ("vehicle_no", "str"),
    "업체명": ("operator_name", "str"),
    "차대번호": ("chassis_no", "str"),
    "차명": ("model_name", "str"),
    "연식": ("model_year", "int"),
    "차량등록일": ("registered_at", "date"),
    "차종": ("vehicle_class", "str"),
    "길이(mm)": ("length_mm", "int"),
    "너비(mm)": ("width_mm", "int"),
    "높이(mm)": ("height_mm", "int"),
    "총중량(kg)": ("gross_weight_kg", "int"),
    "승차정원": ("seating_capacity", "int"),
    "연료": ("fuel", "str"),
}


def _coerce(value, kind: str):
    """엑셀 셀 값 → 필드 타입. 빈값은 None. int는 float("123.0") 표기도 흡수."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    if kind == "str":
        return str(value).strip()
    if kind == "int":
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None
    if kind == "date":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None
    return value


_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
_FLEET_REQUIRED = {"vehicle_no"}  # 양식 헤더에 " *" 표기할 필수 필드
# 양식 예시행 차량번호 접두 — 파스/분류에서 이 접두로 시작하면 skip(실데이터엔 없어 회귀 0)
EXAMPLE_PREFIX = "(예시) "
# 양식 예시행 1개(필드→값). 헤더 순서(_HEADER_MAP)에 맞춰 write.
_FLEET_EXAMPLE_ROW = {
    "vehicle_no": EXAMPLE_PREFIX + "서울70사1234",
    "operator_name": "서울교통공사",
    "chassis_no": "KMJHG18BPMC000001",
    "model_name": "일렉시티",
    "model_year": 2022,
    "registered_at": "2022-03-01",
    "vehicle_class": "대형 승합",
    "length_mm": 10995,
    "width_mm": 2495,
    "height_mm": 3405,
    "gross_weight_kg": 17300,
    "seating_capacity": 47,
    "fuel": "전기",
}


def _open_fleet_workbook(content: bytes):
    """업로드 bytes → (데이터 행 이터레이터, 열인덱스→(필드,타입) 매핑).

    import/preview 공유. 크기·빈파일·엑셀오류·헤더 검증은 기존 import 동작과 동일
    (상태코드·메시지 불변). 1행 헤더를 소비하고 데이터 행 이터레이터를 돌려준다.
    """
    if not content:
        raise HTTPException(status_code=422, detail="빈 파일은 업로드할 수 없습니다")
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="파일 크기가 25MB를 초과합니다")
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=422,
            detail="엑셀(.xlsx) 파일을 읽을 수 없습니다 — BUS_LIST_ALL 시트를 확인하세요",
        )
    ws = wb[_FLEET_SHEET] if _FLEET_SHEET in wb.sheetnames else wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        raise HTTPException(status_code=422, detail="빈 파일입니다 — 1행에 컬럼 헤더가 필요합니다")

    # 헤더 매칭 — 라벨 기준(공백 trim), 순서 무관
    col_field: dict = {}  # 열 인덱스 → (필드, 타입)
    for idx, cell in enumerate(header):
        name = str(cell).strip() if cell is not None else ""
        name = name.rstrip("* ").strip()  # 양식의 필수 표기 " *" 접미 흡수
        if name in _HEADER_MAP:
            col_field[idx] = _HEADER_MAP[name]
    if not any(f == "vehicle_no" for f, _ in col_field.values()):
        raise HTTPException(
            status_code=422, detail="필수 컬럼이 없습니다: 차량번호 — 헤더를 확인하세요"
        )
    return rows_iter, col_field


def _parse_fleet_row(values, col_field: dict) -> dict:
    """엑셀 행 값 → {필드: 강제변환값}. region/client 파생은 호출측 책임."""
    parsed = {}
    for idx, (field, kind) in col_field.items():
        parsed[field] = _coerce(values[idx] if idx < len(values) else None, kind)
    return parsed


def _classify_fleet_row(parsed: dict, by_chassis: dict, by_vehicle_no: dict):
    """행 분류 순수 함수 → ("skip"|"new"|"update", 기존 cv|None). 인덱스 미변형.

    식별키: 차대번호 있으면 chassis, 없으면 차량번호. 빈 차량번호/예시행이면 skip.
    """
    vehicle_no = parsed.get("vehicle_no")
    if not vehicle_no or vehicle_no.startswith(EXAMPLE_PREFIX):
        return "skip", None
    chassis_no = parsed.get("chassis_no")
    cv = by_chassis.get(chassis_no) if chassis_no else by_vehicle_no.get(vehicle_no)
    if cv is None:
        return "new", None
    return "update", cv


def _fleet_client_by_name(db: Session) -> dict:
    """업체명 → 운수사 client_id(동명이인은 None 보류). import/preview 공유."""
    client_by_name: dict = {}
    _dup_names = set()
    for cid, name in db.query(Client.client_id, Client.company_name).filter(
        Client.client_type == "TRANSPORT"
    ):
        if name in client_by_name:
            _dup_names.add(name)  # 동명이인 — 임의 매칭 대신 보류
        client_by_name[name] = cid
    for name in _dup_names:
        client_by_name[name] = None
    return client_by_name


def _build_fleet_template() -> bytes:
    """전국 버스 명부 업로드 양식(.xlsx) 생성 — 헤더는 _HEADER_MAP 단일원천.

    시트명=BUS_LIST_ALL, 1행=라벨(필수 필드는 " *" 접미), 2행=예시 1행(차량번호 접두 방지).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = _FLEET_SHEET
    headers = []
    example = []
    for label, (field, _kind) in _HEADER_MAP.items():
        headers.append(label + (" *" if field in _FLEET_REQUIRED else ""))
        example.append(_FLEET_EXAMPLE_ROW.get(field))
    ws.append(headers)
    ws.append(example)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── A. 전역 엑셀 업로드 ────────────────────────────────────────────────────
@router.post("/fleet/import", response_model=schemas.FleetImportResult)
async def import_fleet(
    file: UploadFile = File(...),
    user: User = Depends(require_permission("master.write")),
    db: Session = Depends(get_db),
):
    """운수사 보유 차량 전량 엑셀 업로드(BUS_LIST_ALL) — 차대번호 upsert + 참여 링크 갱신.

    - 업체명(operator_name) == 운수사(Client.company_name, TRANSPORT)면 client_id 매칭(미매칭 None).
    - region = 차량번호 앞 2글자, status 기본 '운행'.
    - 식별키 = 차대번호(chassis_no) 있으면 chassis 기준, 없으면 차량번호 기준 upsert.
      5,668행 규모 — 선조회 dict로 N+1 방지, 커밋 1회.
    - 업로드 후 ProjectVehicle.client_vehicle_id를 차량번호 일치로 세팅(참여 구분 신선도).
    - 도입구분 자동 판별: 참여차량 차량번호가 내연 fleet에 있으면 대체도입, 없으면 신규도입.
      introduction_type이 비어있는(None) 참여차량만 자동설정(기존 수기값 보존).
    """
    content = await file.read()
    rows_iter, col_field = _open_fleet_workbook(content)

    # 선조회: 차대번호/차량번호→기존 마스터, 업체명→운수사 client_id (N+1 방지)
    fleet = db.query(ClientVehicle).all()
    by_chassis = {cv.chassis_no: cv for cv in fleet if cv.chassis_no}
    by_vehicle_no = {cv.vehicle_no: cv for cv in fleet if cv.vehicle_no}
    client_by_name = _fleet_client_by_name(db)

    created = updated = client_matched = skipped = 0
    for values in rows_iter:
        if values is None:
            continue
        parsed = _parse_fleet_row(values, col_field)
        kind, cv = _classify_fleet_row(parsed, by_chassis, by_vehicle_no)
        if kind == "skip":
            if any(v is not None for v in values):  # 완전 빈 행은 skip 집계 제외
                skipped += 1
            continue
        vehicle_no = parsed["vehicle_no"]
        operator_name = parsed.get("operator_name")
        matched_client_id = client_by_name.get(operator_name) if operator_name else None
        if matched_client_id:
            client_matched += 1
        parsed["region"] = vehicle_no[:2]
        chassis_no = parsed.get("chassis_no")

        if kind == "new":
            cv = ClientVehicle(vehicle_no=vehicle_no)
            for k, v in parsed.items():
                setattr(cv, k, v)
            cv.client_id = matched_client_id
            cv.status = "운행"  # 신규 기본값
            db.add(cv)
            if chassis_no:
                by_chassis[chassis_no] = cv
            by_vehicle_no[vehicle_no] = cv
            fleet.append(cv)  # 도입구분 판별용 내연 집합 최신화
            created += 1
        else:
            # 재업로드 upsert: 파일에 있는 스펙 필드만 갱신. status(수기 폐차)는 파일에
            # 없으므로 보존, client_id는 매칭 성공 시에만 갱신(미매칭 시 기존 지정 보존).
            for k, v in parsed.items():
                if k == "vehicle_no":
                    continue
                setattr(cv, k, v)
            if matched_client_id:
                cv.client_id = matched_client_id
            updated += 1
    db.flush()  # vehicle_id(gen_uuid) 확보 — 참여 링크 매핑용

    # 참여 링크 갱신: 차량번호 일치 ProjectVehicle에 client_vehicle_id 세팅(신선도)
    # (참여 EV는 fleet에 없을 수 있으나 링크는 vehicle_no 기준으로 유지)
    no_to_id = {cv.vehicle_no: cv.vehicle_id for cv in fleet if cv.vehicle_no}
    linked = 0
    pvs = (
        db.query(ProjectVehicle)
        .filter(ProjectVehicle.vehicle_no.isnot(None))
        .all()
    )
    for pv in pvs:
        target = no_to_id.get(pv.vehicle_no)
        if target and pv.client_vehicle_id != target:
            pv.client_vehicle_id = target
            linked += 1

    # 도입구분 자동 판별: 참여차량 차량번호가 내연 fleet에 있으면 대체도입, 없으면 신규도입.
    # introduction_type이 비어있는(None) 참여차량만 자동설정(기존 수기값 보존).
    ice_novs = {
        cv.vehicle_no
        for cv in fleet
        if cv.vehicle_no and (not cv.fuel or "전기" not in str(cv.fuel))
    }
    introduction_derived = 0
    for pv in db.query(ProjectVehicle).filter(
        ProjectVehicle.vehicle_no.isnot(None),
        ProjectVehicle.introduction_type.is_(None),
    ):
        pv.introduction_type = "대체도입" if pv.vehicle_no in ice_novs else "신규도입"
        introduction_derived += 1

    AuditLogger.log_action(
        db,
        user.user_id,
        "FLEET_IMPORT",
        target_type="CLIENT_VEHICLE",
        new_value=(
            "보유 차량 업로드 — 생성 {0} / 갱신 {1} / 매칭 {2} / 참여링크 {3} / "
            "도입구분 {4} / 건너뜀 {5}".format(
                created, updated, client_matched, linked, introduction_derived, skipped
            )
        ),
    )
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="참조 데이터가 변경되어 반영하지 못했습니다. 다시 시도해 주세요.",
        )
    return schemas.FleetImportResult(
        created=created,
        updated=updated,
        client_matched=client_matched,
        linked_participation=linked,
        introduction_derived=introduction_derived,
        skipped=skipped,
    )


@router.get("/fleet/template")
def download_fleet_template(
    _: User = Depends(require_permission("master.write")),
):
    """전국 버스 명부 업로드 양식(.xlsx) 다운로드 — 헤더(필수 *)+예시 1행.

    파일명 한글은 RFC 5987 인코딩(imports 다운로드와 동일 관용구)."""
    content = _build_fleet_template()
    return Response(
        content=content,
        media_type=_XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": "attachment; filename*=UTF-8''{0}".format(
                quote("전국버스명부_양식.xlsx")
            )
        },
    )


@router.post("/fleet/preview", response_model=schemas.FleetPreviewResult)
async def preview_fleet(
    file: UploadFile = File(...),
    _: User = Depends(require_permission("master.write")),
    db: Session = Depends(get_db),
):
    """전역 fleet 엑셀 dry-run — 실반영 없이 신규/갱신/건너뜀 예측(부록 M).

    DB read-only(flush/commit/add 금지). import와 동일 규칙으로 분류·집계하되
    파일 내 중복 파리티를 위해 신규 예측 행을 센티넬로 인덱스에 등록해 뒤 중복행이
    갱신으로 잡히게 한다. 응답 rows는 '건너뜀' 행 상세만(대량 페이로드 방지).
    """
    content = await file.read()
    rows_iter, col_field = _open_fleet_workbook(content)

    fleet = db.query(ClientVehicle).all()
    by_chassis = {cv.chassis_no: cv for cv in fleet if cv.chassis_no}
    by_vehicle_no = {cv.vehicle_no: cv for cv in fleet if cv.vehicle_no}
    client_by_name = _fleet_client_by_name(db)

    _sentinel = object()  # 파일 내 신규 예측 등록용(뒤 중복행 → 갱신 파리티)
    total_rows = created = updated = skipped = client_matched = 0
    rows: list = []
    for rownum, values in enumerate(rows_iter, start=2):  # 1행 헤더, 데이터는 2행부터
        if values is None:
            continue
        parsed = _parse_fleet_row(values, col_field)
        vehicle_no = parsed.get("vehicle_no")
        kind, _cv = _classify_fleet_row(parsed, by_chassis, by_vehicle_no)
        if kind == "skip":
            if not any(v is not None for v in values):
                continue  # 완전 빈 행은 집계 제외(import와 동일)
            total_rows += 1
            skipped += 1
            reason = "예시행" if (vehicle_no and vehicle_no.startswith(EXAMPLE_PREFIX)) else "차량번호 없음"
            rows.append(
                schemas.FleetPreviewRow(
                    row=rownum,
                    vehicle_no=vehicle_no,
                    chassis_no=parsed.get("chassis_no"),
                    classification="건너뜀",
                    reason=reason,
                )
            )
            continue
        total_rows += 1
        operator_name = parsed.get("operator_name")
        if operator_name and client_by_name.get(operator_name):
            client_matched += 1
        chassis_no = parsed.get("chassis_no")
        if kind == "new":
            created += 1
            # 파일 내 중복 파리티: 신규 예측을 센티넬로 등록(뒤 동일키 행 → 갱신)
            if chassis_no:
                by_chassis[chassis_no] = _sentinel
            by_vehicle_no[vehicle_no] = _sentinel
        else:
            updated += 1

    return schemas.FleetPreviewResult(
        total_rows=total_rows,
        created=created,
        updated=updated,
        skipped=skipped,
        client_matched=client_matched,
        rows=rows,
    )


# ── B. 고객사 상세 보유 차량(참여 구분) ─────────────────────────────────────
def _cv_out(
    cv: ClientVehicle, client_name: Optional[str], part: Optional[dict]
) -> schemas.ClientVehicleListItem:
    item = schemas.ClientVehicleListItem.model_validate(cv, from_attributes=True)
    update = {"client_name": client_name, "participation": part is not None}
    if part is not None:
        update.update(part)
    return item.model_copy(update=update)


@router.get(
    "/clients/{client_id}/vehicles", response_model=schemas.ClientVehicleListResponse
)
def list_client_vehicles(
    client_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    q: Optional[str] = Query(None, description="차량번호 검색"),
    participation: str = Query("all", pattern="^(all|participating|unassigned)$"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """고객사 보유 차량 목록 — 참여 구분(participating/unassigned) 필터·검색·페이지네이션."""
    client = common.get_or_404(db, Client, client_id, "고객사")

    # 참여 여부 exists 서브쿼리 — ProjectVehicle이 이 마스터를 가리키는지
    part_sub = db.query(ProjectVehicle.vehicle_id).filter(
        ProjectVehicle.client_vehicle_id == ClientVehicle.vehicle_id
    )

    base = db.query(ClientVehicle).filter(ClientVehicle.client_id == client_id)
    if q:
        keyword = "%{0}%".format(common.escape_like(q.strip()))
        base = base.filter(ClientVehicle.vehicle_no.ilike(keyword, escape="\\"))

    # 참여 집계는 필터 무관(고객사 전체 기준)
    total_all = db.query(ClientVehicle).filter(ClientVehicle.client_id == client_id).count()
    participating_count = (
        db.query(ClientVehicle)
        .filter(ClientVehicle.client_id == client_id, part_sub.exists())
        .count()
    )
    unassigned_count = total_all - participating_count

    if participation == "participating":
        base = base.filter(part_sub.exists())
    elif participation == "unassigned":
        base = base.filter(~part_sub.exists())

    total = base.count()
    rows = (
        base.order_by(ClientVehicle.vehicle_no.asc(), ClientVehicle.vehicle_id.asc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    # 대표 참여 정보 선조회(페이지 cv 한정) — cv_id → 대표 1건
    cv_ids = [cv.vehicle_id for cv in rows]
    part_map: dict = {}
    if cv_ids:
        pvs = (
            db.query(ProjectVehicle)
            .filter(ProjectVehicle.client_vehicle_id.in_(cv_ids))
            .order_by(ProjectVehicle.created_at.asc(), ProjectVehicle.vehicle_id.asc())
            .all()
        )
        proj_ids = {pv.project_id for pv in pvs}
        proj_names = {
            pid: name
            for pid, name in db.query(Project.project_id, Project.project_name).filter(
                Project.project_id.in_(proj_ids)
            )
        } if proj_ids else {}
        for pv in pvs:
            if pv.client_vehicle_id in part_map:
                continue  # 첫 건만 대표
            part_map[pv.client_vehicle_id] = {
                "project_id": pv.project_id,
                "project_name": proj_names.get(pv.project_id),
                "introduction_type": pv.introduction_type,
                "effective_reduction": (
                    float(pv.effective_reduction)
                    if pv.effective_reduction is not None
                    else None
                ),
                "expected_payout": (
                    float(pv.expected_payout) if pv.expected_payout is not None else None
                ),
            }

    items = [_cv_out(cv, client.company_name, part_map.get(cv.vehicle_id)) for cv in rows]
    return schemas.ClientVehicleListResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        participating_count=participating_count,
        unassigned_count=unassigned_count,
    )


# ── C. fleet 단건 CRUD(수기) ───────────────────────────────────────────────
@router.post(
    "/clients/{client_id}/vehicles",
    response_model=schemas.ClientVehicleOut,
    status_code=201,
)
def create_client_vehicle(
    client_id: str,
    payload: schemas.ClientVehicleIn,
    user: User = Depends(require_permission("master.write")),
    db: Session = Depends(get_db),
):
    """보유 차량 수기 등록 — region 앞2 파생, status 기본 '운행', 차대번호 중복 시 409.

    식별키는 차대번호(chassis_no). 차량번호(vehicle_no) 중복은 허용(내연/전기 공존).
    """
    common.get_or_404(db, Client, client_id, "고객사")
    if payload.asset_id:
        common.get_or_404(db, Asset, payload.asset_id, "자산")
    status = payload.status or "운행"
    validate_active_code(db, "VEHICLE_STATUS", status)
    if payload.chassis_no and db.query(ClientVehicle).filter(
        ClientVehicle.chassis_no == payload.chassis_no
    ).first():
        raise HTTPException(
            status_code=409, detail="이미 등록된 차대번호입니다: {0}".format(payload.chassis_no)
        )
    cv = ClientVehicle(
        client_id=client_id,
        vehicle_no=payload.vehicle_no,
        region=(payload.vehicle_no or "")[:2] or None,
        status=status,
        **{f: getattr(payload, f) for f in _FLEET_FIELDS if f != "status"},
    )
    db.add(cv)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409, detail="이미 등록된 차대번호입니다: {0}".format(payload.chassis_no)
        )
    # 참여 링크 신선도 — 같은 차량번호의 기존 미링크 ProjectVehicle에 역링크(수기 후생성 대비)
    if cv.vehicle_no:
        db.query(ProjectVehicle).filter(
            ProjectVehicle.vehicle_no == cv.vehicle_no,
            ProjectVehicle.client_vehicle_id.is_(None),
        ).update({"client_vehicle_id": cv.vehicle_id}, synchronize_session=False)
    AuditLogger.log_action(
        db, user.user_id, "CLIENT_VEHICLE_CREATE",
        target_type="CLIENT_VEHICLE", target_id=cv.vehicle_id,
    )
    db.commit()
    db.refresh(cv)
    return schemas.ClientVehicleOut.model_validate(cv, from_attributes=True)


@router.put(
    "/clients/{client_id}/vehicles/{vehicle_id}", response_model=schemas.ClientVehicleOut
)
def update_client_vehicle(
    client_id: str,
    vehicle_id: str,
    payload: schemas.ClientVehicleUpdate,
    user: User = Depends(require_permission("master.write")),
    db: Session = Depends(get_db),
):
    """보유 차량 부분 수정 — 전달된 필드만 반영(스코프 client_id+vehicle_id)."""
    cv = (
        db.query(ClientVehicle)
        .filter(
            ClientVehicle.client_id == client_id,
            ClientVehicle.vehicle_id == vehicle_id,
        )
        .first()
    )
    if cv is None:
        raise HTTPException(status_code=404, detail="차량을 찾을 수 없습니다")
    data = payload.model_dump(exclude_unset=True)
    if data.get("status"):
        validate_active_code(db, "VEHICLE_STATUS", data["status"])
    if data.get("asset_id"):
        common.get_or_404(db, Asset, data["asset_id"], "자산")
    # 식별키 = 차대번호: chassis_no 변경 시에만 중복 409. 차량번호 중복은 허용(내연/전기 공존).
    if "chassis_no" in data and data["chassis_no"] and data["chassis_no"] != cv.chassis_no:
        dup = (
            db.query(ClientVehicle)
            .filter(ClientVehicle.chassis_no == data["chassis_no"])
            .first()
        )
        if dup is not None:
            raise HTTPException(
                status_code=409, detail="이미 등록된 차대번호입니다: {0}".format(data["chassis_no"])
            )
    if "vehicle_no" in data and data["vehicle_no"] and data["vehicle_no"] != cv.vehicle_no:
        cv.region = data["vehicle_no"][:2] or None
    for field in ("vehicle_no", "client_id", *_FLEET_FIELDS):
        if field in data:
            setattr(cv, field, data[field])
    AuditLogger.log_action(
        db, user.user_id, "CLIENT_VEHICLE_UPDATE",
        target_type="CLIENT_VEHICLE", target_id=cv.vehicle_id,
    )
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="차대번호가 중복되어 저장하지 못했습니다")
    db.refresh(cv)
    return schemas.ClientVehicleOut.model_validate(cv, from_attributes=True)


@router.delete(
    "/clients/{client_id}/vehicles/{vehicle_id}", response_model=schemas.MessageResponse
)
def delete_client_vehicle(
    client_id: str,
    vehicle_id: str,
    user: User = Depends(require_permission("master.write")),
    db: Session = Depends(get_db),
):
    """보유 차량 삭제 — 참여 링크(ProjectVehicle.client_vehicle_id)는 끊고 프로젝트 차량은 보존."""
    cv = (
        db.query(ClientVehicle)
        .filter(
            ClientVehicle.client_id == client_id,
            ClientVehicle.vehicle_id == vehicle_id,
        )
        .first()
    )
    if cv is None:
        raise HTTPException(status_code=404, detail="차량을 찾을 수 없습니다")
    # fleet 마스터 삭제 — ProjectVehicle 링크만 끊고 참여 차량 자체는 유지
    db.query(ProjectVehicle).filter(
        ProjectVehicle.client_vehicle_id == vehicle_id
    ).update({"client_vehicle_id": None}, synchronize_session=False)
    db.delete(cv)
    AuditLogger.log_action(
        db, user.user_id, "CLIENT_VEHICLE_DELETE",
        target_type="CLIENT_VEHICLE", target_id=vehicle_id,
    )
    db.commit()
    return schemas.MessageResponse(message="차량이 삭제되었습니다")

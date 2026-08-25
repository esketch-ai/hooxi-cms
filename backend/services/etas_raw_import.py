"""운수사별 원본(eTAS .xls / BMS취합 .xlsx) 직접 파서 → 차량 월별 로그(D6, P3).

담당자가 eTAS·BMS에서 내려받은 원본을 그대로 올리면 로그로 적재(수작업 취합 제거).
- eTAS .xls: 회사명·운행기간 메타 + [자동차등록번호·운행일수·총 운행거리] 표(다중헤더). source=ETAS.
- BMS취합 .xlsx(LONG): 운수사·차량번호·월·운행거리합계·운행횟수합계. source=BMS.
파일 포맷은 구조로 자동 판별. 월은 eTAS=파일명, BMS=행의 '월' 컬럼.
"""

import re
from io import BytesIO
from typing import Dict, List, Optional

import xlrd
from openpyxl import load_workbook

# 파일명/기간에서 'YYYY년 MM월' 또는 'YYYY-MM'
_YM_KR = re.compile(r"(\d{4})\s*년\s*(\d{1,2})\s*월")
_YM_ISO = re.compile(r"(\d{4})-(\d{1,2})")


def _num(v) -> Optional[float]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _month_from_name(name: str) -> Optional[str]:
    m = _YM_KR.search(name or "") or _YM_ISO.search(name or "")
    if not m:
        return None
    return "{0}-{1:02d}".format(m.group(1), int(m.group(2)))


def parse_etas_raw(content: bytes, filename: str = "") -> List[dict]:
    """eTAS 운수사별 원본 .xls → 차량 월별 로그(운행일수·운행거리). 월=파일명."""
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    grid = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]

    def _s(v) -> str:
        return str(v).replace("\n", "").replace(" ", "").strip() if v is not None else ""

    # 1) 헤더 행 탐지: '자동차등록번호' 셀 위치
    vehicle_col = header_row = None
    for r, row in enumerate(grid):
        for c, cell in enumerate(row):
            if _s(cell) == "자동차등록번호":
                vehicle_col, header_row = c, r
                break
        if header_row is not None:
            break
    if header_row is None:
        return []

    # 2) 헤더 행과 다음 행에서 운행일수·총운행거리 컬럼 탐지
    days_col = dist_col = None
    for r in (header_row, header_row + 1):
        if r >= len(grid):
            continue
        for c, cell in enumerate(row_at := grid[r]):
            s = _s(cell)
            if days_col is None and s == "운행일수":
                days_col = c
            if dist_col is None and ("총운행거리" in s or s == "운행거리"):
                dist_col = c

    # 3) 회사명(운수사) 메타 — 헤더 앞쪽에서 '회사명' 오른쪽 값
    operator = None
    for r in range(0, header_row):
        for c, cell in enumerate(grid[r]):
            if _s(cell) == "회사명":
                for cc in range(c + 1, len(grid[r])):
                    if _s(grid[r][cc]):
                        operator = str(grid[r][cc]).strip()
                        break
            if operator:
                break
        if operator:
            break

    ym = _month_from_name(filename)
    if not ym:
        # 운행기간 메타에서 폴백
        for r in range(0, header_row):
            for cell in grid[r]:
                mm = _YM_ISO.search(str(cell or ""))
                if mm and "운행" in "".join(_s(x) for x in grid[r]):
                    ym = "{0}-{1:02d}".format(mm.group(1), int(mm.group(2)))
                    break
            if ym:
                break
    if not ym:
        return []

    out = []
    for r in range(header_row + 2, len(grid)):
        row = grid[r]
        vno = row[vehicle_col] if vehicle_col < len(row) else None
        if not vno or not str(vno).strip():
            continue
        rec = {"vehicle_no": str(vno).strip(), "year_month": ym, "source": "ETAS",
               "operator_name": operator}
        if days_col is not None and days_col < len(row):
            d = _num(row[days_col])
            if d is not None:
                rec["operating_days"] = d
        if dist_col is not None and dist_col < len(row):
            dv = _num(row[dist_col])
            if dv is not None:
                rec["distance_km"] = dv
        # 지표가 전무하면 스킵(빈 행)
        if "operating_days" in rec or "distance_km" in rec:
            out.append(rec)
    return out


def parse_bms_long(content: bytes) -> List[dict]:
    """BMS취합 LONG .xlsx → 차량 월별 로그(운행거리·운행횟수). 월=행 컬럼."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(h).replace(" ", "").strip() if h is not None else "" for h in rows[0]]
    idx: Dict[str, int] = {}
    for i, h in enumerate(header):
        if h in ("차량번호", "자동차등록번호"):
            idx["vno"] = i
        elif h == "월":
            idx["ym"] = i
        elif h in ("운수사", "운수사명", "업체명"):
            idx["op"] = i
        elif "운행거리" in h:
            idx["dist"] = i
        elif "운행횟수" in h:
            idx["trip"] = i
    if "vno" not in idx or "ym" not in idx:
        return []

    out = []
    for values in rows[1:]:
        if not values:
            continue
        vno = values[idx["vno"]] if idx["vno"] < len(values) else None
        ym_raw = values[idx["ym"]] if idx["ym"] < len(values) else None
        if not vno or not ym_raw:
            continue
        mm = _YM_ISO.search(str(ym_raw)) or _YM_KR.search(str(ym_raw))
        if not mm:
            continue
        ym = "{0}-{1:02d}".format(mm.group(1), int(mm.group(2)))
        rec = {"vehicle_no": str(vno).strip(), "year_month": ym, "source": "BMS",
               "operator_name": None}
        if "op" in idx and idx["op"] < len(values) and values[idx["op"]]:
            rec["operator_name"] = str(values[idx["op"]]).strip()
        if "dist" in idx and idx["dist"] < len(values):
            dv = _num(values[idx["dist"]])
            if dv is not None:
                rec["distance_km"] = dv
        if "trip" in idx and idx["trip"] < len(values):
            tv = _num(values[idx["trip"]])
            if tv is not None:
                rec["trip_count"] = tv
        if "distance_km" in rec or "trip_count" in rec:
            out.append(rec)
    return out


def parse_raw_file(content: bytes, filename: str = "") -> List[dict]:
    """확장자·구조로 eTAS(.xls) / BMS취합(.xlsx LONG) 자동 판별 후 파싱."""
    lower = (filename or "").lower()
    if lower.endswith(".xls"):
        return parse_etas_raw(content, filename)
    # .xlsx — BMS취합 LONG 우선 시도
    return parse_bms_long(content)

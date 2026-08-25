"""차량별 산정 입력(eTAS·BMS 크롤링 정규화) 업로드 — 중복 체크 후 upsert(CRUD, D5).

표준 템플릿(라벨 매핑): 차량번호·업체명·연료·연평균주행(베이스라인)·연평균연료·
연평균주행(사업)·연평균충전·전기차등록연도. 차량번호로 중복 판정해 갱신/생성.
크롤러 산출 포맷이 다르면 라벨 매핑만 확장하면 된다.
"""

import re
from typing import Dict, List, Optional

from openpyxl import load_workbook
from io import BytesIO

from models import Client, ReductionRegistry, VehicleCalcInput
from services.excel_import import _tf_company_clean
from services.region_norm import normalize_region

# 라벨(공백무시) → 필드. 베이스라인/사업 주행을 구분하기 위해 접미 라벨 사용.
_LABEL_FIELD = {
    "차량번호": "vehicle_no",
    "업체명": "operator_name",
    "운수사": "operator_name",
    "권역": "region",
    "지역": "region",
    "연료": "fuel",
    "베이스라인연료": "fuel",
    "연평균주행거리(베이스라인)": "baseline_distance",
    "베이스라인연평균주행거리": "baseline_distance",
    "연평균주행거리_베이스라인": "baseline_distance",
    "연평균연료사용량": "baseline_fuel",
    "연평균연료": "baseline_fuel",
    "연평균주유량": "baseline_fuel",
    "연평균주행거리(사업)": "project_distance",
    "사업연평균주행거리": "project_distance",
    "연평균주행거리_사업": "project_distance",
    "연평균충전량": "project_kwh",
    "전기차등록연도": "ev_reg_year",
    "전기차등록년도": "ev_reg_year",
    "민간투자비율": "private_ratio",
    "민간비율": "private_ratio",
    "사업구분": "introduction_type",
    "도입구분": "introduction_type",
    # 차대번호(VIN) — 대체도입 판정. 내연/전기 구분 라벨 우선, 단독 차대번호는 미지정.
    "베이스라인차대번호": "baseline_vin",
    "내연차대번호": "baseline_vin",
    "기존차대번호": "baseline_vin",
    "전기차대번호": "project_vin",
    "사업차대번호": "project_vin",
    "신규차대번호": "project_vin",
}
_NUM = {"baseline_distance", "baseline_fuel", "project_distance", "project_kwh", "private_ratio"}
_INT = {"ev_reg_year"}


def _norm(v) -> str:
    return re.sub(r"\s+", "", str(v or ""))


def _to_num(v) -> Optional[float]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None


def _to_int(v) -> Optional[int]:
    n = _to_num(v)
    return int(n) if n is not None else None


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


_LABEL_NORM = {_norm(k): f for k, f in _LABEL_FIELD.items()}


def parse_calc_inputs(content: bytes) -> List[dict]:
    """표준 템플릿 엑셀 → 차량별 입력 dict 목록(첫 시트, 1행 헤더)."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = rows[0]
    col_field = {}
    for idx, label in enumerate(header):
        f = _LABEL_NORM.get(_norm(label))
        if f:
            col_field[idx] = f
    if "vehicle_no" not in col_field.values():
        return []
    out = []
    for values in rows[1:]:
        if not values or all(v is None for v in values):
            continue
        rec: Dict[str, object] = {}
        for idx, field in col_field.items():
            if idx >= len(values):
                continue
            raw = values[idx]
            if field in _NUM:
                rec[field] = _to_num(raw)
            elif field in _INT:
                rec[field] = _to_int(raw)
            elif field == "region":
                rec[field] = normalize_region(raw) if raw else None
            else:
                rec[field] = _clean(raw)
        if rec.get("vehicle_no"):
            out.append(rec)
    return out


def _registry_vin_index(db) -> Dict[str, dict]:
    """차량번호 → {baseline_vin, project_vin, introduction_type} — 레지스트리 권위값."""
    idx: Dict[str, dict] = {}
    for r in db.query(ReductionRegistry.vehicle_no, ReductionRegistry.role,
                      ReductionRegistry.vin, ReductionRegistry.introduction_type).all():
        if not r.vehicle_no:
            continue
        slot = idx.setdefault(r.vehicle_no, {})
        if r.role == "BASELINE" and r.vin:
            slot.setdefault("baseline_vin", r.vin)
        elif r.role == "PROJECT":
            if r.vin:
                slot.setdefault("project_vin", r.vin)
            if r.introduction_type:
                slot.setdefault("introduction_type", r.introduction_type)
    return idx


def _resolve_vin(rec: dict, reg: Dict[str, dict]) -> None:
    """도입구분별 VIN 검증 — 신규도입은 VIN 쌍 검증 대상 아님(NEW), 대체도입만 OK/WARN."""
    r = reg.get(rec["vehicle_no"], {})
    if not rec.get("baseline_vin") and r.get("baseline_vin"):
        rec["baseline_vin"] = r["baseline_vin"]
    if not rec.get("project_vin") and r.get("project_vin"):
        rec["project_vin"] = r["project_vin"]
    if not rec.get("introduction_type") and r.get("introduction_type"):
        rec["introduction_type"] = r["introduction_type"]

    itype = rec.get("introduction_type")
    bv, pv = rec.get("baseline_vin"), rec.get("project_vin")
    if itype == "신규도입":
        rec["vin_status"] = "NEW"  # 유사 화석연료차 선정 — 대체도입 VIN 쌍 검증 대상 아님
        return
    # 대체도입(또는 미상): 같은 차량번호·내연≠전기 VIN 확인
    if bv and pv:
        rec["vin_status"] = "OK" if bv != pv else "WARN"
        if bv == pv:
            rec["memo"] = "VIN 동일 — 대체도입 아님(확인 필요)"
    elif not bv and not pv:
        rec["vin_status"] = "WARN"
        rec["memo"] = "차대번호 없음 — 레지스트리 미매칭"
    else:
        rec["vin_status"] = "WARN"
        rec["memo"] = "차대번호 한쪽만 확인됨"


def _client_index(db) -> Dict[tuple, str]:
    idx: Dict[tuple, str] = {}
    for c in db.query(Client.client_id, Client.company_name, Client.region).all():
        if not c.company_name:
            continue
        key = (normalize_region(c.region or ""), _tf_company_clean(c.company_name).replace(" ", ""))
        idx.setdefault(key, c.client_id)
    return idx


def apply_calc_inputs(db, rows: List[dict]) -> dict:
    """차량번호로 중복 체크 후 upsert(CRUD) — 차대번호(VIN) 레지스트리 교차검증 포함."""
    cindex = _client_index(db)
    reg = _registry_vin_index(db)
    created = updated = matched = vin_ok = vin_warn = vin_new = 0
    for r in rows:
        _resolve_vin(r, reg)  # 도입구분·VIN 보완 + vin_status
        st = r.get("vin_status")
        if st == "OK":
            vin_ok += 1
        elif st == "NEW":
            vin_new += 1
        else:
            vin_warn += 1
        vno = r["vehicle_no"]
        existing = db.query(VehicleCalcInput).filter(
            VehicleCalcInput.vehicle_no == vno).first()
        client_id = None
        op = r.get("operator_name")
        if op:
            key = (normalize_region(r.get("region") or ""),
                   _tf_company_clean(op).replace(" ", ""))
            client_id = cindex.get(key)
            if client_id:
                matched += 1
        if existing:
            for k, v in r.items():
                setattr(existing, k, v)
            if client_id:
                existing.client_id = client_id
            existing.source = "CRAWL_IMPORT"
            updated += 1
        else:
            db.add(VehicleCalcInput(client_id=client_id, source="CRAWL_IMPORT", **r))
            created += 1
    return {
        "created": created, "updated": updated, "client_matched": matched,
        "vin_ok": vin_ok, "vin_warn": vin_warn, "vin_new": vin_new, "total": len(rows),
    }

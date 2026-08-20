"""운수사 계약대수 현황 — 원본 엑셀 파서 + 매칭·미리보기·적용 (F2).

매월 발행되는 원본 엑셀(단일 탭, 복합 헤더 2~4행)을 고정 컬럼 위치로 읽어, 지역(조합)+
회사명(정제)으로 고객사(TRANSPORT)에 매칭하고 (고객사×월) 단위로 대수를 합산 upsert한다.

- 다중 사업장(같은 회사 여러 행)은 합산해 1행.
- 미매칭(지역+회사명으로 고객사 못 찾음)은 client_id NULL 보류(앱 레벨 dedup: region+회사명).
- 수작업 관리(tb_fleet_mgmt)는 건드리지 않는다.
"""

from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import load_workbook

from models import Client, Code, FleetStatus
from services.excel_import import _tf_company_clean

# 원본 탭 고정 컬럼(0-base) — 실측: A조합 B업종 C회사명 E월 F면허대수 G계 H경유 I CNG J HB K전기 L수소
_COL = {
    "region": 0, "industry": 1, "company": 2, "month": 4,
    "license": 5, "total": 6,
    "diesel": 7, "cng": 8, "hybrid": 9, "electric": 10, "hydrogen": 11,
}
_NUM_FIELDS = ["license", "total", "diesel", "cng", "hybrid", "electric", "hydrogen"]
# dict 키 → 모델 컬럼명(license/total만 상이)
_MODEL_COL = {"license": "license_count", "total": "total_count"}

# 업종 라벨 → FLEET_INDUSTRY 코드
_INDUSTRY_CODE = {"시내": "CITY", "농어촌": "RURAL", "시외": "INTERCITY"}


def _to_int(v) -> int:
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().replace(",", "")
    try:
        return int(float(s))
    except ValueError:
        return 0


def parse_rows(file_bytes: bytes) -> List[dict]:
    """원본 엑셀(단일 탭) → 행 dict 목록. 헤더 행은 자동 스킵(회사명+면허대수 숫자 기준)."""
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        from fastapi import HTTPException
        raise HTTPException(status_code=422, detail="엑셀(.xlsx) 파일을 읽을 수 없습니다")
    # '원본' 탭 우선(현행 정본은 현황+원본 2탭). 앞으로 단일 탭이면 그 시트 사용.
    ws = wb["원본"] if "원본" in wb.sheetnames else wb.worksheets[0]
    out: List[dict] = []
    for values in ws.iter_rows(values_only=True):
        if values is None:
            continue
        company = values[_COL["company"]] if len(values) > _COL["company"] else None
        lic = values[_COL["license"]] if len(values) > _COL["license"] else None
        # 데이터 행 판정 — 회사명(문자) + 면허대수(숫자). 헤더/빈행은 제외.
        if not (isinstance(company, str) and company.strip()):
            continue
        if not isinstance(lic, (int, float)):
            continue
        row = {
            "region": (str(values[_COL["region"]]).strip() if values[_COL["region"]] else ""),
            "industry_raw": (str(values[_COL["industry"]]).strip() if values[_COL["industry"]] else ""),
            "company_name": company.strip(),
        }
        for f in _NUM_FIELDS:
            row[f] = _to_int(values[_COL[f]] if len(values) > _COL[f] else None)
        out.append(row)
    return out


def _client_lookup(db) -> Dict[tuple, Client]:
    """(지역, 정제 회사명) → 운수사(TRANSPORT). 매칭 인덱스 1회 로드."""
    idx: Dict[tuple, Client] = {}
    for c in db.query(Client).filter(Client.client_type == "TRANSPORT").all():
        key = ((c.region or "").strip(), _tf_company_clean(c.company_name or ""))
        idx.setdefault(key, c)
    return idx


def _industry_code(db, raw: str) -> Optional[str]:
    """업종 라벨 → FLEET_INDUSTRY 코드. 기본표 우선, tb_code 라벨도 수용."""
    if not raw:
        return None
    if raw in _INDUSTRY_CODE:
        return _INDUSTRY_CODE[raw]
    row = (
        db.query(Code.code)
        .filter(Code.category == "FLEET_INDUSTRY", Code.label == raw)
        .first()
    )
    return row[0] if row else raw


def _aggregate(db, rows: List[dict]) -> List[dict]:
    """다중 사업장 합산 — (고객사 or 지역+정제회사명) 키로 대수 합산해 1행. 매칭 정보 포함."""
    lookup = _client_lookup(db)
    agg: Dict[tuple, dict] = {}
    for r in rows:
        clean = _tf_company_clean(r["company_name"])
        client = lookup.get((r["region"], clean))
        key = ("cid", client.client_id) if client else ("rc", r["region"], clean)
        cur = agg.get(key)
        if cur is None:
            cur = {
                "region": r["region"],
                "industry": _industry_code(db, r["industry_raw"]),
                "company_name": r["company_name"],  # 첫 원문 회사명 보존
                "matched_client_id": client.client_id if client else None,
                "matched_client_name": client.company_name if client else None,
            }
            for f in _NUM_FIELDS:
                cur[f] = 0
            agg[key] = cur
        for f in _NUM_FIELDS:
            cur[f] += r[f]
    return list(agg.values())


def analyze(db, file_bytes: bytes, period: str) -> dict:
    """미리보기(DB 무변경) — 합산·매칭·신규/갱신 판정."""
    rows = parse_rows(file_bytes)
    items = _aggregate(db, rows)
    for it in items:
        it["period"] = period
        existing = _find_existing(db, it, period)
        it["is_update"] = existing is not None
        it["matched"] = it["matched_client_id"] is not None
    matched = sum(1 for it in items if it["matched"])
    return {
        "period": period,
        "total_rows": len(rows),
        "aggregated": len(items),
        "matched": matched,
        "unmatched": len(items) - matched,
        "items": items,
    }


def _find_existing(db, item: dict, period: str) -> Optional[FleetStatus]:
    """기존 스냅샷 — 매칭 건은 (client_id, period), 미매칭은 (region, 회사명, period)."""
    q = db.query(FleetStatus).filter(FleetStatus.period == period)
    if item.get("matched_client_id"):
        return q.filter(FleetStatus.client_id == item["matched_client_id"]).first()
    return (
        q.filter(
            FleetStatus.client_id.is_(None),
            FleetStatus.region == item["region"],
            FleetStatus.company_name == item["company_name"],
        ).first()
    )


def commit(db, file_bytes: bytes, period: str, actor_id: Optional[str] = None) -> dict:
    """적용 — (고객사×월) upsert. 기존 있으면 대수 갱신, 없으면 생성. 미매칭은 보류로 적재."""
    rows = parse_rows(file_bytes)
    items = _aggregate(db, rows)
    created = updated = 0
    for it in items:
        existing = _find_existing(db, it, period)
        if existing is not None:
            existing.region = it["region"]
            existing.industry = it["industry"]
            existing.company_name = it["company_name"]
            for f in _NUM_FIELDS:
                setattr(existing, _MODEL_COL.get(f, f), it[f])
            updated += 1
        else:
            fs = FleetStatus(
                client_id=it["matched_client_id"],
                region=it["region"],
                industry=it["industry"],
                company_name=it["company_name"],
                period=period,
                source="EXCEL",
                created_by=actor_id,
                **{_MODEL_COL.get(f, f): it[f] for f in _NUM_FIELDS},
            )
            db.add(fs)
            created += 1
    db.commit()
    matched = sum(1 for it in items if it["matched_client_id"])
    return {
        "period": period,
        "total_rows": len(rows),
        "aggregated": len(items),
        "created": created,
        "updated": updated,
        "matched": matched,
        "unmatched": len(items) - matched,
    }

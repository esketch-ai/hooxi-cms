"""KISA 블록체인 데이터(500대) → 감축 참여 레지스트리 적재(M3).

프로그램 전체 차량을 role별로 파싱한다:
  BASELINE  = 시트 '1. 베이스라인 제원(대체 전 화석연료버스)'
  PROJECT   = 시트 '2. 사업대상 제원(대체도입 전기버스)'
  CANDIDATE = 시트 '3. 대체 예정 화석연료버스'
운수사 매칭은 (지역+정제 업체명)으로 best-effort(services.region_norm·excel_import 재사용).
방법론 결과가 아니라 차량 현황 원장 — 감축량은 별도(project_vehicles ingest)다.
"""

from datetime import date, datetime
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import load_workbook

from models import Client, ReductionRegistry
from services.excel_import import _tf_company_clean
from services.region_norm import normalize_region

# (시트명, role) — 시트명은 원본 정확 매칭, 없으면 건너뜀(부분 파일 허용)
_SHEETS = [
    ("1. 베이스라인 제원(대체 전 화석연료버스)", "BASELINE"),
    ("2. 사업대상 제원(대체도입 전기버스)", "PROJECT"),
    ("3. 대체 예정 화석연료버스", "CANDIDATE"),
]

# 헤더 라벨 → 표준 필드(시트마다 헤더가 조금씩 달라 라벨 매핑으로 흡수)
_LABEL_FIELD = {
    "차량번호": "vehicle_no",
    "자동차등록번호": "vehicle_no",
    "업체명": "operator_name",
    "순번": "seq",
    "사업구분": "introduction_type",
    "차명": "model_name",
    "차대번호": "vin",
    "연식": "model_year",
    "전기차량 등록일": "registered_at",
    "등록일": "registered_at",
    "차종": "vehicle_class",
    "용도": "purpose",
    "길이(mm)": "length_mm",
    "너비(mm)": "width_mm",
    "높이(mm)": "height_mm",
    "총중량(kg)": "gross_weight_kg",
    "승차정원": "seating_capacity",
    "연료": "fuel",
    "배터리종류": "battery_type",
    "사업명": "program_name",
    "권역": "region",
}
_INT_FIELDS = {"seq", "model_year", "length_mm", "width_mm", "height_mm",
               "gross_weight_kg", "seating_capacity"}


def _to_int(v) -> Optional[int]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except (TypeError, ValueError):
        return None


def _to_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse_registry(content: bytes) -> List[dict]:
    """엑셀 bytes → 레지스트리 행 dict 목록(적재 전, DB 무관)."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    out: List[dict] = []
    for sheet_name, role in _SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [(_clean(h) or "") for h in rows[0]]
        col_field = {}
        for idx, label in enumerate(header):
            field = _LABEL_FIELD.get(label)
            if field:
                col_field[idx] = field
        for values in rows[1:]:
            if not values or all(v is None for v in values):
                continue
            rec: Dict[str, object] = {"role": role, "source": "KISA_IMPORT"}
            for idx, field in col_field.items():
                if idx >= len(values):
                    continue
                raw = values[idx]
                if field in _INT_FIELDS:
                    rec[field] = _to_int(raw)
                elif field == "registered_at":
                    rec[field] = _to_date(raw)
                elif field == "region":
                    rec[field] = normalize_region(raw) if raw else None
                elif field == "operator_name":
                    rec[field] = _clean(raw)
                else:
                    rec[field] = _clean(raw)
            if not rec.get("vehicle_no"):
                continue  # 차량번호 없는 행(빈/합계) 제외
            out.append(rec)
    wb.close()
    return out


def _client_match_index(db) -> Dict[tuple, str]:
    """(정규화지역, 정제업체명) → client_id — 운수사 best-effort 매칭."""
    idx: Dict[tuple, str] = {}
    for c in db.query(Client.client_id, Client.company_name, Client.region).all():
        if not c.company_name:
            continue
        key = (normalize_region(c.region or ""), _tf_company_clean(c.company_name).replace(" ", ""))
        idx.setdefault(key, c.client_id)
    return idx


def apply_registry(db, rows: List[dict], replace: bool = True) -> dict:
    """레지스트리 적재 — replace=True면 KISA_IMPORT 출처 전량 교체(멱등 재적재).

    운수사 매칭은 (지역+정제 업체명). 커밋은 호출부 책임. 요약 dict 반환.
    """
    if replace:
        db.query(ReductionRegistry).filter(
            ReductionRegistry.source == "KISA_IMPORT"
        ).delete(synchronize_session=False)

    cindex = _client_match_index(db)
    created = matched = 0
    by_role = {"BASELINE": 0, "PROJECT": 0, "CANDIDATE": 0}
    for r in rows:
        client_id = None
        op = r.get("operator_name")
        region = r.get("region")
        if op:
            key = (normalize_region(region or ""), _tf_company_clean(op).replace(" ", ""))
            client_id = cindex.get(key)
            if client_id:
                matched += 1
        entry = ReductionRegistry(client_id=client_id, **r)
        db.add(entry)
        created += 1
        by_role[r["role"]] = by_role.get(r["role"], 0) + 1
    return {"created": created, "client_matched": matched, "by_role": by_role}

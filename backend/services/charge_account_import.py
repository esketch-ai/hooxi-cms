"""충전 관제 계정 목록 → 자산·연동(Asset) 일괄 등록(외부기관 계정 관리).

충전량(kWh)은 표준 단일 소스가 없고 운수사마다 서로 다른 충전 관제 벤더 포털
(eBAB 펌프킨·조아스차저·KEViT·ELVIS·PODO-TMS 등)에 계정별 로그인해 받는다. 그 수집 경로를
CMS 자산·연동(Asset)에 등록해 '어느 운수사 충전량을 어디서 받는지'를 관제한다.

보안: 비밀번호는 crypto(AES-256-GCM)로 암호화 저장하며, 키 미설정 시 메타만 등록(비번 제외).
평문 비밀번호는 파싱 중 메모리에서만 다루고 로그·감사·응답에 절대 남기지 않는다(R2-E6).
표준 표: [구분 · 회사명 · 시내/마을 · 홈페이지 명 · 홈페이지 주소 · 아이디 · 비밀번호 · 비고]
빈 구분/회사명은 직전 값 이어받음(운수사당 복수 사이트·계정).
"""

import re
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import load_workbook

from models import Asset, Client
from services import crypto
from services.excel_import import _tf_company_clean
from services.region_norm import normalize_region

_LABEL_FIELD = {
    "구분": "region", "구분지역": "region", "지역": "region",
    "회사명": "company", "운수사": "company", "업체명": "company",
    "시내마을": "mode", "시내/마을": "mode", "구분2": "mode",
    "홈페이지명": "system_name", "시스템명": "system_name", "관제시스템": "system_name",
    "홈페이지주소": "site_url", "주소": "site_url", "url": "site_url",
    "아이디": "login_id", "id": "login_id",
    "비밀번호": "password", "비번": "password", "pw": "password", "password": "password",
    "비고": "note", "메모": "note",
}

# 상태 판정 키워드(비고·회사명·시스템명)
_ERR_KW = ("로그인 안됨", "로그인안됨", "접속 불가", "접속불가", "접속 불가능", "접속불가능")
_INACTIVE_KW = ("전기차 없음", "전기차없음", "요청 중", "요청중", "확인 불가", "확인불가",
                "확인 필요", "확인필요", "구축 중", "구축중")


def _norm(v) -> str:
    return re.sub(r"\s+", "", str(v or "")).lower()


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v)).strip()
    return s or None


_LABEL_NORM = {_norm(k): f for k, f in _LABEL_FIELD.items()}


def parse_charge_accounts(content: bytes) -> List[dict]:
    """엑셀 표 → 계정 레코드 목록. 빈 구분/회사명은 직전 값 이어받음."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = rows[0]
    col_field: Dict[int, str] = {}
    for idx, label in enumerate(header):
        f = _LABEL_NORM.get(_norm(label))
        if f and idx not in col_field:
            col_field[idx] = f
    if "company" not in col_field.values():
        return []

    out: List[dict] = []
    last_region = last_company = None
    for values in rows[1:]:
        if not values or all(v is None or str(v).strip() == "" for v in values):
            continue
        rec: Dict[str, Optional[str]] = {}
        for idx, field in col_field.items():
            if idx < len(values):
                rec[field] = _clean(values[idx])
        # 이어받기(운수사당 복수 계정 행)
        if rec.get("region"):
            last_region = rec["region"]
        else:
            rec["region"] = last_region
        if rec.get("company"):
            last_company = rec["company"]
        else:
            rec["company"] = last_company
        if not rec.get("company"):
            continue
        # 의미 없는 완전 빈 계정행(사이트·아이디·비고 전무) 스킵
        if not any(rec.get(k) for k in ("system_name", "site_url", "login_id", "note")):
            continue
        out.append(rec)
    return out


def _derive_status(rec: dict) -> str:
    blob = " ".join(str(rec.get(k) or "") for k in ("note", "company", "system_name"))
    if any(k in blob for k in _ERR_KW):
        return "ERROR"
    if rec.get("login_id") and rec.get("site_url"):
        return "ACTIVE"
    if any(k in blob for k in _INACTIVE_KW):
        return "INACTIVE"
    return "INACTIVE"


def _client_index(db) -> Dict[tuple, str]:
    idx: Dict[tuple, str] = {}
    for c in db.query(Client.client_id, Client.company_name, Client.region).all():
        if not c.company_name:
            continue
        key = (normalize_region(c.region or ""),
               _tf_company_clean(c.company_name).replace(" ", ""))
        idx.setdefault(key, c.client_id)
    return idx


def _match_client(cindex: Dict[tuple, str], region: Optional[str],
                  company: Optional[str]) -> Optional[str]:
    if not company:
        return None
    rg = normalize_region(region or "")
    # '성우교통, 자안운수'처럼 복수 표기 → 첫 매칭
    for token in re.split(r"[,/·]", company):
        nm = _tf_company_clean(token).replace(" ", "")
        if not nm:
            continue
        cid = cindex.get((rg, nm))
        if cid:
            return cid
    return None


def apply_charge_accounts(db, rows: List[dict]) -> dict:
    """계정 레코드 → Asset(자산·연동) upsert. 비번은 키 있을 때만 암호화 저장."""
    cindex = _client_index(db)
    can_encrypt = crypto.encryption_available()
    created = updated = matched = encrypted = pw_skipped = unmatched = 0
    unmatched_names: List[str] = []
    for rec in rows:
        client_id = _match_client(cindex, rec.get("region"), rec.get("company"))
        if not client_id:
            # Asset.client_id는 NOT NULL — 고객사 미매칭이면 등록 불가(마스터 보완 안내)
            unmatched += 1
            nm = rec.get("company")
            if nm and nm not in unmatched_names:
                unmatched_names.append(nm)
            continue
        matched += 1
        status = _derive_status(rec)
        login_id = rec.get("login_id")
        password = rec.get("password")
        auth_type = "ID_PW" if login_id else "NONE"
        agency = rec.get("system_name") or rec.get("company")
        note_bits = [b for b in (rec.get("mode"), rec.get("note")) if b]

        # 멱등 키: (고객사, 사이트, 아이디, 시스템명) — 재업로드 시 갱신
        existing = db.query(Asset).filter(
            Asset.usage_purpose == "충전량 수집",
            Asset.client_id == client_id,
            Asset.site_url == rec.get("site_url"),
            Asset.login_id == login_id,
            Asset.agency_name == agency,
        ).first()

        asset = existing or Asset(client_id=client_id)
        asset.asset_group = "MOBILITY"
        asset.asset_type = "EV"
        asset.status = status
        asset.telemetry_yn = "Y"
        asset.agency_name = (agency or "")[:100]
        asset.site_url = (rec.get("site_url") or None)
        asset.auth_type = auth_type
        asset.login_id = login_id
        asset.usage_purpose = "충전량 수집"
        asset.main_spec = (rec.get("mode") or None)
        asset.location_info = (" / ".join(note_bits) or None)
        # 비밀번호 — 키 있을 때만 암호화 저장, 없으면 메타만(응답·로그에 평문 미노출)
        if password and auth_type == "ID_PW":
            if can_encrypt:
                asset.login_password = crypto.encrypt(password)
                encrypted += 1
            else:
                pw_skipped += 1
        if existing:
            updated += 1
        else:
            db.add(asset)
            created += 1
    return {
        "created": created, "updated": updated, "client_matched": matched,
        "unmatched": unmatched, "unmatched_names": unmatched_names[:50],
        "encrypted": encrypted, "password_skipped": pw_skipped,
        "encryption_available": can_encrypt, "total": len(rows),
    }

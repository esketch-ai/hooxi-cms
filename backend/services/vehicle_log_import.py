"""eTAS·BMS 원본/취합본 → 차량 월별 로그 적재(D6, P1).

지원 포맷:
  - INTEGRATED(WIDE): 담당자 취합본 — [운수사명·자동차등록번호 + YYYY년MM월_운행일수/운행거리/충전량]
    반복. 월별로 분해해 로그 1행씩.
중복키 (차량번호, 월, 출처)로 upsert. 프로그램 차량 필터는 조회단에서.
"""

import re
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import load_workbook

from models import Client, VehicleMonthlyLog
from services.excel_import import _tf_company_clean
from services.region_norm import normalize_region

# 'YYYY년MM월_지표' 헤더 파싱
_HDR_RE = re.compile(r"(\d{4})년\s*(\d{1,2})월[_\s]*(운행일수|운행거리|충전량|운행횟수)")
_METRIC_FIELD = {"운행일수": "operating_days", "운행거리": "distance_km",
                 "충전량": "charge_kwh", "운행횟수": "trip_count"}


def _num(v) -> Optional[float]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None


def parse_integrated_wide(content: bytes) -> List[dict]:
    """취합본 WIDE → 로그 행(차량×월). 값이 하나도 없는 (차량,월)은 생략."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    # 컬럼 → (year_month, field)
    month_cols: Dict[int, tuple] = {}
    op_col = vno_col = None
    for idx, h in enumerate(header):
        if h in ("운수사명", "운수사"):
            op_col = idx
        elif h in ("자동차등록번호", "차량번호"):
            vno_col = idx
        else:
            m = _HDR_RE.search(h)
            if m:
                ym = "{0}-{1:02d}".format(m.group(1), int(m.group(2)))
                month_cols[idx] = (ym, _METRIC_FIELD[m.group(3)])
    if vno_col is None or not month_cols:
        return []

    out = []
    for values in rows[1:]:
        if not values:
            continue
        vno = values[vno_col] if vno_col < len(values) else None
        if not vno or not str(vno).strip():
            continue
        op = values[op_col] if (op_col is not None and op_col < len(values)) else None
        op = str(op).strip() if op and str(op).strip() != "-" else None
        # (year_month) → {field: value}
        by_month: Dict[str, dict] = {}
        for idx, (ym, field) in month_cols.items():
            if idx >= len(values):
                continue
            val = _num(values[idx])
            if val is None:
                continue
            by_month.setdefault(ym, {})[field] = val
        for ym, fields in by_month.items():
            if not fields:
                continue
            out.append({
                "vehicle_no": str(vno).strip(), "year_month": ym,
                "source": "INTEGRATED", "operator_name": op, **fields,
            })
    return out


def _client_index(db) -> Dict[tuple, str]:
    idx: Dict[tuple, str] = {}
    for c in db.query(Client.client_id, Client.company_name, Client.region).all():
        if not c.company_name:
            continue
        key = (normalize_region(c.region or ""), _tf_company_clean(c.company_name).replace(" ", ""))
        idx.setdefault(key, c.client_id)
    return idx


def apply_logs(db, rows: List[dict], batch: Optional[str] = None) -> dict:
    """(차량번호, 월, 출처) upsert. 운수사 매칭 best-effort."""
    cindex = _client_index(db)
    created = updated = matched = 0
    veh = set()
    months = set()
    for r in rows:
        veh.add(r["vehicle_no"])
        months.add(r["year_month"])
        client_id = None
        op = r.get("operator_name")
        if op:
            key = ("", _tf_company_clean(op).replace(" ", ""))
            # 지역 미상 — 회사명만으로 1차 매칭(지역 포함 매칭은 레지스트리/마스터 조인에서 보강)
            client_id = next((cid for (rg, nm), cid in cindex.items()
                              if nm == key[1]), None)
            if client_id:
                matched += 1
        existing = db.query(VehicleMonthlyLog).filter(
            VehicleMonthlyLog.vehicle_no == r["vehicle_no"],
            VehicleMonthlyLog.year_month == r["year_month"],
            VehicleMonthlyLog.source == r["source"],
        ).first()
        if existing:
            for k in ("operating_days", "distance_km", "charge_kwh", "trip_count",
                      "operator_name"):
                if k in r:
                    setattr(existing, k, r[k])
            if client_id:
                existing.client_id = client_id
            if batch:
                existing.batch = batch
            updated += 1
        else:
            db.add(VehicleMonthlyLog(client_id=client_id, batch=batch, **r))
            created += 1
    return {
        "created": created, "updated": updated, "client_matched": matched,
        "vehicles": len(veh), "months": len(months), "total": len(rows),
    }

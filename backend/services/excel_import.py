"""엑셀 일괄 등록 — 양식 생성·파싱·행 검증 (SCR-03 고객사 / SCR-04 자산).

컬럼 규격은 services/import_spec.py **단일 원천**에서만 파생한다 — 이 파일은
스펙을 해석할 뿐, 라벨·필드명을 하드코딩하지 않는다(라벨 변경 시 수정 불필요).

- 셀 검증은 대상 Pydantic 스키마(ClientCreate/AssetCreate)에 위임하고,
  ValidationError를 스펙 라벨 기준 한국어 사유로 번역한다.
- 코드 컬럼은 tb_code 활성 코드의 라벨/코드 양방향 수용(validate_active_code 재사용).
- preview·commit이 같은 parse_and_validate를 공유 — 결과 일관성 보장(무상태).
"""

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Dict, List, Optional, Union, get_args, get_origin

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ValidationError
from sqlalchemy.orm import Session

import schemas
from models import Client, Code, User
from routers import common
from routers.codes import validate_active_code
from services import import_spec
from services.import_spec import MAX_IMPORT_ROWS, ImportColumn, ImportSpec

# Y/N 정규화 — 실무 표기(예/아니오·O/X 등) 수용
_YN_TRUE = {"Y", "YES", "예", "O", "TRUE", "1"}
_YN_FALSE = {"N", "NO", "아니오", "아니요", "X", "FALSE", "0"}

# 날짜 문자열 허용 포맷 — 엑셀 문자 셀 입력 관용 표기
_DATE_FORMATS = ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S")

# Pydantic v2 오류 타입 → 한국어 사유 (미등재 타입은 원문 유지)
_PYDANTIC_MESSAGES = {
    "missing": "필수 값입니다",
    "string_too_long": "값이 너무 깁니다",
    "string_too_short": "필수 값입니다",
    "int_parsing": "숫자여야 합니다",
    "int_type": "숫자여야 합니다",
    "int_from_float": "정수여야 합니다",
    "greater_than_equal": "0 이상이어야 합니다",
    "string_pattern_mismatch": "허용되지 않는 값입니다",
    "datetime_parsing": "날짜 형식이 올바르지 않습니다 (예: 2026-01-15)",
    "datetime_type": "날짜 형식이 올바르지 않습니다 (예: 2026-01-15)",
    "datetime_from_date_parsing": "날짜 형식이 올바르지 않습니다 (예: 2026-01-15)",
}


def get_spec(entity: str) -> ImportSpec:
    """엔티티 규격 조회 — 매 호출 시 IMPORT_SPECS를 참조(스펙 교체·테스트 용이)."""
    spec = import_spec.IMPORT_SPECS.get(entity)
    if spec is None:
        raise HTTPException(status_code=404, detail="지원하지 않는 일괄 등록 대상입니다: {0}".format(entity))
    return spec


# ---------------------------------------------------------------------------
# 양식(.xlsx) 생성 — 헤더(필수 * 표시) + 예시 행 1개
# 예시 행은 "(예시)" 접두로 표기하고 파서가 자동 스킵한다 — 지우지 않고 올려도
# 실데이터로 등록되지 않게(실무 함정 방지)
EXAMPLE_PREFIX = "(예시) "
# ---------------------------------------------------------------------------
def _example_display(db: Session, col: ImportColumn) -> Optional[str]:
    """예시 값 표시 — 코드 컬럼은 불변 코드값을 현재 라벨로 해석(라벨 변경 추종).

    관리자가 tb_code 라벨을 바꿔도 양식 예시가 항상 업로드 가능해야 하므로
    (라운드트립 보장) 스펙에는 코드값을 적고 여기서 라벨로 바꿔 보여준다.
    """
    if not col.example:
        return None
    if col.code_category:
        row = (
            db.query(Code)
            .filter(Code.category == col.code_category, Code.code == col.example)
            .first()
        )
        if row is not None:
            return row.label
    return col.example


def build_template(db: Session, entity: str) -> bytes:
    spec = get_spec(entity)
    wb = Workbook()
    ws = wb.active
    ws.title = spec.label
    for idx, col in enumerate(spec.columns, start=1):
        header = col.label + (" *" if col.required else "")
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)
        # 한글 헤더 폭 보정 — 대략 2배 폭
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(header) * 2 + 4)
        example = _example_display(db, col)
        if idx == 1 and example:
            example = EXAMPLE_PREFIX + str(example)  # 파서가 이 접두로 예시 행을 스킵
        ws.cell(row=2, column=idx, value=example)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def spec_out(db: Session, entity: str) -> schemas.ImportSpecOut:
    """GET /imports/{entity}/spec 응답 — 스펙에서 그대로 파생(예시는 현재 라벨)."""
    spec = get_spec(entity)
    return schemas.ImportSpecOut(
        entity=spec.entity,
        label=spec.label,
        max_rows=MAX_IMPORT_ROWS,
        filename=spec.filename,
        columns=[
            schemas.ImportColumnOut(
                field=c.field,
                label=c.label,
                required=c.required,
                code_category=c.code_category,
                resolver=c.resolver,
                yn=c.yn,
                allowed_values=(list(c.fixed_values.keys()) if c.fixed_values else None),
                example=_example_display(db, c),
            )
            for c in spec.columns
        ],
    )


# ---------------------------------------------------------------------------
# 파싱·검증
# ---------------------------------------------------------------------------
@dataclass
class ParsedRow:
    """데이터 행 1개의 검증 결과 — payload는 오류 없을 때만 채워진다."""

    row: int  # 엑셀 실제 행 번호 (헤더=1)
    display: Dict[str, Optional[str]] = field(default_factory=dict)  # 라벨 → 표시값
    payload: Optional[BaseModel] = None
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


@dataclass
class ParseResult:
    """parse_and_validate 결과 — preview 응답과 commit 대상(유효 행)을 함께 제공."""

    spec: ImportSpec
    rows: List[ParsedRow] = field(default_factory=list)
    unknown_columns: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)  # 예시 행 스킵 등 파일 수준 안내

    @property
    def valid_rows(self) -> List[ParsedRow]:
        return [r for r in self.rows if r.payload is not None and not r.errors]

    def to_preview(self) -> schemas.ImportPreviewOut:
        valid = len(self.valid_rows)
        return schemas.ImportPreviewOut(
            entity=self.spec.entity,
            total_rows=len(self.rows),
            valid_rows=valid,
            error_rows=len(self.rows) - valid,
            unknown_columns=self.unknown_columns,
            warnings=self.warnings,
            rows=[row_result(r) for r in self.rows],
        )


def row_result(r: ParsedRow) -> schemas.ImportRowResult:
    return schemas.ImportRowResult(
        row=r.row,
        status="ERROR" if r.errors else "OK",
        data=r.display,
        errors=r.errors,
        warnings=r.warnings,
    )


def _normalize_header(value) -> str:
    """헤더 셀 정규화 — 공백 trim + 양식의 필수 표시(*) 제거."""
    s = str(value or "").strip()
    while s.endswith("*"):
        s = s[:-1].rstrip()
    return s


def _unwrap_optional(annotation):
    if get_origin(annotation) is Union:
        args = [a for a in get_args(annotation) if a is not type(None)]
        if len(args) == 1:
            return args[0]
    return annotation


def _field_kind(spec: ImportSpec, field_name: str) -> str:
    """대상 스키마 필드 타입에서 셀 정규화 방식 도출 — 스펙 변경 시 자동 추종."""
    info = spec.schema_cls.model_fields.get(field_name)
    ann = _unwrap_optional(info.annotation) if info else str
    if ann is int:
        return "int"
    if ann is datetime or ann is date:
        return "datetime"
    return "str"


def _cell_to_str(value) -> Optional[str]:
    """문자 대상 필드의 셀 정규화 — 숫자 셀 3.0 → "3" (사업자번호·전화 등).

    주의: 엑셀 '숫자 셀'은 앞자리 0이 이미 소실된 상태(엑셀 자체 동작)라 복원
    불가 — 양식 안내는 문자 서식 입력 권장. 여기서는 소수점 꼬리(.0)만 제거한다.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return "Y" if value else "N"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    return s or None


def _cell_to_datetime(value):
    """날짜 대상 필드 — 날짜 셀은 그대로, 문자 셀은 관용 표기 파싱 후 스키마 위임."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return s  # 스키마 파싱에 위임 (ISO 등) — 실패 시 번역된 오류로 안내


def _code_label_map(db: Session, category: str) -> Dict[str, str]:
    """활성 코드의 라벨 → 코드값 매핑 (라벨/코드 양방향 수용의 라벨 쪽)."""
    rows = (
        db.query(Code)
        .filter(Code.category == category, Code.active == "Y")
        .all()
    )
    return {r.label.strip(): r.code for r in rows}


class _Resolvers:
    """이름 → ID 해석기 — 파일 1건당 1회만 DB 조회(이름별 후보 목록)."""

    def __init__(self, db: Session):
        self._db = db
        self._maps: Dict[str, Dict[str, List[str]]] = {}

    def _map(self, kind: str) -> Dict[str, List[str]]:
        if kind not in self._maps:
            built: Dict[str, List[str]] = {}
            if kind == "user_by_name":
                rows = self._db.query(User.name, User.user_id).filter(
                    User.status == "ACTIVE"
                ).all()
            elif kind == "client_by_name":
                rows = self._db.query(Client.company_name, Client.client_id).all()
            else:
                raise HTTPException(status_code=500, detail="알 수 없는 resolver: {0}".format(kind))
            for name, pk in rows:
                if name:
                    built.setdefault(name.strip(), []).append(pk)
            self._maps[kind] = built
        return self._maps[kind]

    def resolve(self, kind: str, label: str, raw: str) -> str:
        """이름 → ID. 미존재/동명이인은 행 오류(ValueError 메시지)."""
        target = "사용자" if kind == "user_by_name" else "고객사"
        candidates = self._map(kind).get(raw, [])
        if not candidates:
            raise ValueError("{0}: 존재하지 않는 {1}입니다: '{2}'".format(label, target, raw))
        if len(candidates) > 1:
            raise ValueError(
                "{0}: 같은 이름의 {1}이(가) {2}건 있어 지정할 수 없습니다: '{3}'".format(
                    label, target, len(candidates), raw
                )
            )
        return candidates[0]


def _normalize_cell(
    db: Session, spec: ImportSpec, col: ImportColumn, value, resolvers: _Resolvers
):
    """컬럼 규격에 따른 셀 정규화 — 반환 None이면 payload에서 생략(스키마 기본값)."""
    kind = _field_kind(spec, col.field)
    if col.yn:
        raw = _cell_to_str(value)
        if raw is None:
            return None
        upper = raw.strip().upper()
        if upper in _YN_TRUE:
            return "Y"
        if upper in _YN_FALSE:
            return "N"
        raise ValueError("{0}: Y/N 값이어야 합니다: '{1}'".format(col.label, raw))
    if col.fixed_values:
        raw = _cell_to_str(value)
        if raw is None:
            return None
        mapped = col.fixed_values.get(raw) or col.fixed_values.get(raw.upper())
        if mapped is None:
            raise ValueError(
                "{0}: 허용되지 않는 값입니다: '{1}' (허용: {2})".format(
                    col.label, raw, ", ".join(col.fixed_values.keys())
                )
            )
        return mapped
    if col.code_category:
        raw = _cell_to_str(value)
        if raw is None:
            return None
        # 라벨 → 코드 우선, 아니면 코드값 그대로 검증(validate_active_code 재사용 —
        # 비활성·미존재 코드 메시지를 단건 등록과 동일하게 유지)
        mapped = _code_label_map(db, col.code_category).get(raw, raw)
        try:
            validate_active_code(db, col.code_category, mapped)
        except HTTPException as exc:
            raise ValueError(str(exc.detail))
        return mapped
    if col.resolver:
        raw = _cell_to_str(value)
        if raw is None:
            return None
        return resolvers.resolve(col.resolver, col.label, raw)
    if kind == "int":
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        return value.strip() if isinstance(value, str) else value
    if kind == "datetime":
        return _cell_to_datetime(value)
    return _cell_to_str(value)


def _translate_validation_error(
    exc: ValidationError, label_by_field: Dict[str, str]
) -> List[str]:
    """Pydantic ValidationError → 스펙 라벨 기준 한국어 사유."""
    messages = []
    for err in exc.errors():
        loc = str(err["loc"][0]) if err["loc"] else ""
        label = label_by_field.get(loc, loc)
        msg = _PYDANTIC_MESSAGES.get(err["type"])
        if msg is None:
            raw = err.get("msg", "")
            # 커스텀 validator(ValueError)는 이미 한국어 — "Value error, " 접두만 제거
            msg = raw[len("Value error, "):] if raw.startswith("Value error, ") else raw
        messages.append("{0}: {1}".format(label, msg))
    return messages


def _check_duplicates(db: Session, spec: ImportSpec, parsed: ParsedRow, seen: Dict):
    """파일 내·DB 중복 검사 — 유효 payload가 만들어진 행에만 적용."""
    p = parsed.payload
    if spec.entity == "clients":
        # 파일 내 사업자번호 중복 → 뒤 행 오류 (둘 다 반영되면 단건 409 규칙 붕괴)
        norm = common.normalize_biz_no(getattr(p, "biz_reg_no", None))
        if norm:
            first = seen.setdefault(("biz", norm), parsed.row)
            if first != parsed.row:
                parsed.errors.append(
                    "사업자번호: 파일 내 중복입니다 (행 {0}과 동일)".format(first)
                )
            else:
                # DB 중복 — 단건 등록과 동일 규칙·메시지 (common으로 승격된 검사 재사용)
                try:
                    common.check_biz_reg_no_duplicate(db, getattr(p, "biz_reg_no", None))
                except HTTPException as exc:
                    parsed.errors.append(str(exc.detail))
        # 파일 내 회사명 중복 → 경고 (동명 회사 등록 자체는 단건에서도 허용)
        name_key = ("name", (p.company_name or "").strip())
        first = seen.setdefault(name_key, parsed.row)
        if first != parsed.row:
            parsed.warnings.append("회사명: 파일 내 중복 의심 (행 {0}과 동일)".format(first))
    elif spec.entity == "assets":
        key = (
            "asset",
            getattr(p, "client_id", None),
            getattr(p, "asset_group", None),
            (getattr(p, "main_spec", None) or "").strip(),
        )
        first = seen.setdefault(key, parsed.row)
        if first != parsed.row:
            parsed.warnings.append(
                "파일 내 중복 의심 — 고객사·대분류·주요 제원이 행 {0}과 동일".format(first)
            )


def parse_and_validate(db: Session, entity: str, file_bytes: bytes) -> ParseResult:
    """엑셀 파일 → 행별 검증 결과 (DB 무변경). preview·commit 공용 경로."""
    spec = get_spec(entity)
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=422,
            detail="엑셀(.xlsx) 파일을 읽을 수 없습니다 — 다운로드한 양식 파일을 사용하세요",
        )
    try:
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise HTTPException(status_code=422, detail="빈 파일입니다 — 1행에 컬럼 헤더가 필요합니다")

        # 헤더 매칭 — 라벨 기준(공백 trim·필수표시 제거), 순서 무관
        col_by_label = {c.label: c for c in spec.columns}
        matched: Dict[int, ImportColumn] = {}
        unknown: List[str] = []
        for idx, cell in enumerate(header):
            name = _normalize_header(cell)
            if not name:
                continue
            col = col_by_label.get(name)
            if col is None:
                unknown.append(name)
                continue
            if any(c.field == col.field for c in matched.values()):
                raise HTTPException(
                    status_code=422, detail="중복된 컬럼 헤더입니다: '{0}'".format(name)
                )
            matched[idx] = col
        missing = [
            c.label for c in spec.columns if c.required
            and c.field not in {m.field for m in matched.values()}
        ]
        if missing:
            raise HTTPException(
                status_code=422,
                detail="필수 컬럼이 없습니다: {0} — 양식 파일의 1행 헤더를 유지하세요".format(
                    ", ".join(missing)
                ),
            )

        result = ParseResult(spec=spec, unknown_columns=unknown)
        resolvers = _Resolvers(db)
        seen: Dict = {}
        label_by_field = {c.field: c.label for c in spec.columns}

        for row_no, values in enumerate(rows_iter, start=2):
            # 빈 행 스킵 (전 셀 공백)
            if values is None or all(
                v is None or (isinstance(v, str) and not v.strip()) for v in values
            ):
                continue
            # 양식의 예시 행 — 지우지 않고 올려도 실데이터로 등록되지 않게 스킵
            if any(
                isinstance(v, str) and v.strip().startswith(EXAMPLE_PREFIX.strip())
                for v in values
            ):
                msg = "{0}행: 예시 행은 등록에서 제외됩니다".format(row_no)
                if msg not in result.warnings:
                    result.warnings.append(msg)
                continue
            if len(result.rows) >= MAX_IMPORT_ROWS:
                raise HTTPException(
                    status_code=422,
                    detail="한 번에 최대 {0}행까지 처리할 수 있습니다 — 파일을 나눠 업로드하세요".format(
                        MAX_IMPORT_ROWS
                    ),
                )
            parsed = ParsedRow(row=row_no)
            payload: Dict = {}
            for idx, col in matched.items():
                raw = values[idx] if idx < len(values) else None
                try:
                    normalized = _normalize_cell(db, spec, col, raw, resolvers)
                except ValueError as exc:
                    parsed.errors.append(str(exc))
                    parsed.display[col.label] = _cell_to_str(raw)
                    continue
                if normalized is not None:
                    payload[col.field] = normalized
                display = normalized if normalized is not None else raw
                parsed.display[col.label] = (
                    display.strftime("%Y-%m-%d")
                    if isinstance(display, (datetime, date))
                    else (_cell_to_str(display))
                )
            if not parsed.errors:
                try:
                    parsed.payload = spec.schema_cls(**payload)
                except ValidationError as exc:
                    parsed.errors.extend(
                        _translate_validation_error(exc, label_by_field)
                    )
            if parsed.payload is not None:
                _check_duplicates(db, spec, parsed, seen)
                if parsed.errors:
                    parsed.payload = None
            result.rows.append(parsed)
        return result
    finally:
        wb.close()

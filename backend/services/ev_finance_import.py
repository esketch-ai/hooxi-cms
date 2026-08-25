"""차량별 차량가액·보조금·자부담금 → 전기버스 도입 재무(민간투자비율 근거) 적재(D2).

증빙 03의 지역별 시트(강원도/제주도/경상북도…)를 라벨 매핑으로 흡수(시트마다 컬럼 편차:
제주는 농특세 없음·취득가액 추가). 헤더는 2행, 데이터는 3행부터. 엑셀 산정값을 그대로 적재.
민간비율 = 자부담금/차량가액(엑셀 결과 사용, 없으면 파생 계산).
"""

import re
from datetime import date, datetime
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import load_workbook

from models import Client, EvFinance
from services.excel_import import _tf_company_clean
from services.region_norm import normalize_region

# 헤더 라벨(공백·개행 정규화 후) → 필드
_LABEL_FIELD = {
    "순번": "seq",
    "시/군": "sido",
    "운수사": "operator_name",
    "차량번호": "vehicle_no",
    "차대번호": "vin",
    "연도": "year",
    "차량등록일": "registered_at",
    "차종": "vehicle_class",
    "연식": "model_year",
    "자동차 출고가격(부가세 제외)": "release_price",
    "취득세": "acquisition_tax",
    "농어촌특별세": "rural_tax",
    "차량가액": "vehicle_value",
    "취득가액": "acquisition_value",  # 제주 전용(미사용 — 흡수만)
    "저상버스보조금": "low_floor_subsidy",
    "전기차보조금": "ev_subsidy",
    "자부담금": "self_payment",
    "보조금검증(70%이하)": "subsidy_check",
    "민간비율": "private_ratio",
    "공공비율": "public_ratio",
    "비고": "note",
}
_NUM_FIELDS = {"release_price", "acquisition_tax", "rural_tax", "vehicle_value",
               "acquisition_value", "low_floor_subsidy", "ev_subsidy", "self_payment",
               "subsidy_check", "private_ratio", "public_ratio"}
_INT_FIELDS = {"seq", "year", "model_year"}
# 재무 지역 시트만 대상(참고·정의 시트 제외)
_SKIP_SHEETS = {"참고", "정의", "sheet1"}


def _norm_label(v) -> str:
    return re.sub(r"\s+", "", str(v or ""))


def _to_num(v) -> Optional[float]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None


def _to_int(v) -> Optional[int]:
    n = _to_num(v)
    return int(n) if n is not None else None


def _to_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# 라벨 매핑을 공백무시로 — 헤더에 개행 포함(자동차 출고가격\n(부가세 제외))
_LABEL_NORM = {_norm_label(k): f for k, f in _LABEL_FIELD.items()}


def parse_ev_finance(content: bytes) -> List[dict]:
    """지역별 시트 → 재무 행 dict 목록. 시트명은 지역(권역)으로 기록."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    out: List[dict] = []
    for ws in wb.worksheets:
        if _norm_label(ws.title).lower() in _SKIP_SHEETS:
            continue
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        header = rows[1]  # 2행이 세부 헤더
        col_field = {}
        for idx, label in enumerate(header):
            f = _LABEL_NORM.get(_norm_label(label))
            if f:
                col_field[idx] = f
        if "vehicle_no" not in col_field.values():
            continue  # 재무 시트 아님
        region = normalize_region(ws.title) if ws.title else None
        for values in rows[2:]:
            if not values or all(v is None for v in values):
                continue
            rec: Dict[str, object] = {"source": "EVIDENCE_IMPORT", "region": region}
            for idx, field in col_field.items():
                if idx >= len(values) or field in ("acquisition_value", "year", "seq"):
                    continue
                raw = values[idx]
                if field in _INT_FIELDS:
                    rec[field] = _to_int(raw)
                elif field in _NUM_FIELDS:
                    rec[field] = _to_num(raw)
                elif field == "registered_at":
                    rec[field] = _to_date(raw)
                else:
                    rec[field] = _clean(raw)
            if not rec.get("vehicle_no"):
                continue
            # 파생 보정 — 엑셀값 없으면 계산(멱등·안전)
            _fill_derived(rec)
            out.append(rec)
    wb.close()
    return out


def _fill_derived(rec: dict) -> None:
    rp = rec.get("release_price")
    at = rec.get("acquisition_tax") or 0
    rt = rec.get("rural_tax") or 0
    lf = rec.get("low_floor_subsidy") or 0
    ev = rec.get("ev_subsidy") or 0
    if rec.get("vehicle_value") is None and rp is not None:
        rec["vehicle_value"] = rp + at + rt
    if rec.get("self_payment") is None and rp is not None:
        rec["self_payment"] = rp - lf - ev
    vv = rec.get("vehicle_value")
    sp = rec.get("self_payment")
    if rec.get("private_ratio") is None and vv and sp is not None and vv != 0:
        rec["private_ratio"] = round(sp / vv, 6)
    if rec.get("public_ratio") is None and rec.get("private_ratio") is not None:
        rec["public_ratio"] = round(1 - rec["private_ratio"], 6)
    if rec.get("subsidy_check") is None and rp:
        rec["subsidy_check"] = round((lf + ev) / rp, 6)


def _client_index(db) -> Dict[tuple, str]:
    idx: Dict[tuple, str] = {}
    for c in db.query(Client.client_id, Client.company_name, Client.region).all():
        if not c.company_name:
            continue
        key = (normalize_region(c.region or ""), _tf_company_clean(c.company_name).replace(" ", ""))
        idx.setdefault(key, c.client_id)
    return idx


def apply_ev_finance(db, rows: List[dict], replace: bool = True) -> dict:
    """적재 — replace면 EVIDENCE_IMPORT 출처 전량 교체(멱등). 운수사 매칭 best-effort."""
    if replace:
        db.query(EvFinance).filter(EvFinance.source == "EVIDENCE_IMPORT").delete(
            synchronize_session=False
        )
    cindex = _client_index(db)
    created = matched = 0
    for r in rows:
        client_id = None
        op = r.get("operator_name")
        if op:
            key = (normalize_region(r.get("region") or ""),
                   _tf_company_clean(op).replace(" ", ""))
            client_id = cindex.get(key)
            if client_id:
                matched += 1
        db.add(EvFinance(client_id=client_id, **r))
        created += 1
    return {"created": created, "client_matched": matched}

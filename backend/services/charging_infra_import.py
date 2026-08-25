"""충전소별 충전기 제원 → 충전 인프라(차고지·충전기·AC전력량계) 적재(D3, 증빙 02).

'충전기 제원 및 AC전력량계' 시트는 지역 무관 동일 위치 레이아웃:
  A 연번 · B 주소 · C 운수사 · D 충전기순번 · E 충전기제조번호 · F 충전기제조년월
  · G 계순번 · H 계제조번호 · I 계제조년월
차고지(주소·운수사)는 첫 행에만 채워지고(병합) 이후 행은 같은 차고지 → forward-fill.
권역은 주소 첫 토큰으로 파생(normalize_region). MRV 인프라 증빙.
"""

import re
import unicodedata
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import load_workbook

from models import Client
from services.excel_import import _tf_company_clean
from services.region_norm import normalize_region


def _nfc(v) -> str:
    return unicodedata.normalize("NFC", str(v)) if v is not None else ""


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = _nfc(v).strip()
    return s or None


def _region_from_address(addr: Optional[str]) -> Optional[str]:
    if not addr:
        return None
    first = addr.strip().split()[0] if addr.strip() else ""
    return normalize_region(first)


def parse_charging_infra(content: bytes) -> List[dict]:
    """엑셀 → 차고지 목록(각 chargers[]·meters[] 중첩). DB 무관."""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = None
    for ws in wb.worksheets:
        if "충전기" in _nfc(ws.title):
            sheet = ws
            break
    if sheet is None:
        wb.close()
        return []
    rows = list(sheet.iter_rows(values_only=True))
    facilities: List[dict] = []
    index: Dict[tuple, dict] = {}
    cur_addr = cur_op = None
    cur_seq = None
    for values in rows[2:]:  # 1~2행 헤더
        if not values or all(v is None for v in values):
            continue

        def col(i):
            return values[i] if i < len(values) else None

        addr = _clean(col(1))
        op = _clean(col(2))
        seq = col(0)
        if addr:
            cur_addr = addr
        if op:
            cur_op = op
        if isinstance(seq, (int, float)):
            cur_seq = int(seq)
        if not cur_addr:
            continue  # 아직 차고지 미확정(비정상 선두행)

        key = (cur_op or "", cur_addr)
        fac = index.get(key)
        if fac is None:
            fac = {
                "operator_name": cur_op,
                "address": cur_addr,
                "region": _region_from_address(cur_addr),
                "seq": cur_seq,
                "chargers": [],
                "meters": [],
            }
            index[key] = fac
            facilities.append(fac)

        chg_serial = _clean(col(4))
        chg_ym = _clean(col(5))
        if chg_serial or chg_ym:
            fac["chargers"].append({"serial": chg_serial, "ym": chg_ym})
        mtr_serial = _clean(col(7))
        mtr_ym = _clean(col(8))
        if mtr_serial or mtr_ym:
            fac["meters"].append({"serial": mtr_serial, "ym": mtr_ym})
    wb.close()
    return facilities


def _client_index(db) -> Dict[tuple, str]:
    idx: Dict[tuple, str] = {}
    for c in db.query(Client.client_id, Client.company_name, Client.region).all():
        if not c.company_name:
            continue
        key = (normalize_region(c.region or ""), _tf_company_clean(c.company_name).replace(" ", ""))
        idx.setdefault(key, c.client_id)
    return idx


def apply_charging_infra(db, facilities: List[dict], replace: bool = True) -> dict:
    """적재 — replace면 EVIDENCE_IMPORT 출처 차고지 전량 교체(CASCADE로 충전기·계도 삭제)."""
    from models import AcPowerMeter, Charger, ChargingFacility

    if replace:
        # 업로드 파일에 담긴 권역만 교체(다지역 파일을 따로 올려도 서로 안 지움)
        regions = {f.get("region") for f in facilities if f.get("region")}
        q = db.query(ChargingFacility.facility_id).filter(
            ChargingFacility.source == "EVIDENCE_IMPORT")
        if regions:
            q = q.filter(ChargingFacility.region.in_(regions))
        fac_ids = [row.facility_id for row in q.all()]
        if fac_ids:
            db.query(Charger).filter(Charger.facility_id.in_(fac_ids)).delete(synchronize_session=False)
            db.query(AcPowerMeter).filter(AcPowerMeter.facility_id.in_(fac_ids)).delete(synchronize_session=False)
            db.query(ChargingFacility).filter(ChargingFacility.facility_id.in_(fac_ids)).delete(synchronize_session=False)

    cindex = _client_index(db)
    fac_n = chg_n = mtr_n = matched = 0
    for f in facilities:
        client_id = None
        if f.get("operator_name"):
            key = (normalize_region(f.get("region") or ""),
                   _tf_company_clean(f["operator_name"]).replace(" ", ""))
            client_id = cindex.get(key)
            if client_id:
                matched += 1
        fac = ChargingFacility(
            operator_name=f.get("operator_name"), client_id=client_id,
            region=f.get("region"), address=f.get("address"), seq=f.get("seq"),
        )
        db.add(fac)
        db.flush()
        fac_n += 1
        for c in f.get("chargers", []):
            db.add(Charger(facility_id=fac.facility_id, serial_number=c.get("serial"),
                           manufacturing_ym=c.get("ym")))
            chg_n += 1
        for m in f.get("meters", []):
            db.add(AcPowerMeter(facility_id=fac.facility_id, serial_number=m.get("serial"),
                                manufacturing_ym=m.get("ym")))
            mtr_n += 1
    return {"facilities": fac_n, "chargers": chg_n, "meters": mtr_n, "client_matched": matched}

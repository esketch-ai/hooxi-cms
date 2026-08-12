"""엑셀 내보내기 — 조건(필터) 리스트를 .xlsx로 내보내는 공용 헬퍼 (EX-1).

내보내기 관련 단일 진실원(single source of truth):
- 워크북 생성(build_workbook): openpyxl write_only 모드로 대량 행에 대비하고,
  데이터 상단에 **워터마크 블록**(내보낸 사람·시각·경고 문구)을 박아 유출을
  추적·억제한다(후시 내부자료 — 무단 외부유출 금지).
- 다운로드 응답(xlsx_response): routers/imports.py·routers/projects.py에 중복
  존재하던 `_XLSX_MEDIA + Content-Disposition(filename*=UTF-8'')` 관용구를
  이 헬퍼로 수렴한다(RFC 5987 한글 파일명 인코딩).
- 파일명(export_filename): `{리소스}_{YYYYMMDD}.xlsx` (KST today 기본).

셀 타입(ColumnSpec.kind)별 number_format을 지정해 엑셀에서 서식이 유지되게 한다.
"""

from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import List, Optional
from urllib.parse import quote

from fastapi import HTTPException, Response
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models import AuditLog, utcnow

# 다운로드 미디어 타입 — routers의 _XLSX_MEDIA(_TYPE)와 동일 상수를 여기로 수렴
_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# KST = UTC+9 (common.now_kst 규약과 동일 오프셋)
_KST_OFFSET = timedelta(hours=9)

# 내보내기 균형 보안(EX-2~5 공용) — finance·asset·audit export가 중복 없이 공유한다.
# 무음 잘라내기 금지(상한 초과 400) + 일일 반출 횟수 제한(초과 429)의 단일 진실원.
MAX_EXPORT_ROWS = 10000  # 필터 결과 행 상한(초과 시 필터를 좁혀 재시도)
DAILY_EXPORT_LIMIT = 30  # actor 1인 하루(KST) DATA_EXPORT 허용 횟수(감사로그를 카운터로 재사용)

# kind → 엑셀 number_format (text/date는 별도 처리)
_NUMBER_FORMATS = {
    "money": "#,##0",
    "number": "#,##0",
    "percent": "0.0%",
}
_DATE_FORMAT = "yyyy-mm-dd"

# 워터마크 셀 강조 — write_only에서도 Cell 객체에 스타일 부여는 허용됨
_WATERMARK_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")  # 옅은 노랑
_WATERMARK_FONT = Font(bold=True, color="9C5700")


@dataclass
class ColumnSpec:
    """내보내기 컬럼 1개 규격 — key(행 dict 키)·label(헤더)·kind(셀 타입)."""

    key: str
    label: str
    kind: str = "text"  # 'text'|'date'|'money'|'number'|'percent'


def _watermark_text(watermark: dict) -> str:
    """워터마크 문구 — 내보낸 사람(이메일)·시각 + 유출 억제 경고."""
    name = watermark.get("name") or ""
    email = watermark.get("email") or ""
    issued_at = watermark.get("issued_at") or ""
    return (
        "내보낸 사람: {0}({1}) · {2} · "
        "후시 내부자료 — 무단 외부유출 금지".format(name, email, issued_at)
    )


def _data_cell(ws, col: ColumnSpec, value):
    """데이터 셀 1개 — kind별 값·number_format 지정. None은 빈 셀."""
    if value is None:
        return WriteOnlyCell(ws, value=None)
    kind = col.kind
    if kind == "date":
        # date/datetime 값이면 그대로(엑셀 날짜 셀), 그 외는 문자 그대로
        cell = WriteOnlyCell(ws, value=value)
        if isinstance(value, (date, datetime)):
            cell.number_format = _DATE_FORMAT
        return cell
    if kind in ("money", "number", "percent"):
        cell = WriteOnlyCell(ws, value=value)
        cell.number_format = _NUMBER_FORMATS[kind]
        return cell
    # text 등 그 외 — 값 그대로
    return WriteOnlyCell(ws, value=value)


def build_workbook(
    columns: List[ColumnSpec],
    rows: List[dict],
    sheet_title: str,
    watermark: dict,
    total_row: Optional[dict] = None,
) -> bytes:
    """조건 리스트 → .xlsx bytes (write_only 모드, 워터마크 상단 블록 내장).

    구성: [워터마크행] · [빈 행] · [헤더행(bold)] · [데이터행…] · [합계행(옵션)].
    total_row가 주어지면 첫 컬럼에 "합계" 라벨, 나머지는 dict 값을 kind 서식으로.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_title)

    # 컬럼 폭 보정 — 한글 헤더 대략 2배 폭 (write_only는 append 전에 지정)
    for idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(
            14, len(col.label) * 2 + 4
        )

    # 1) 워터마크 상단 블록 — 데이터 위에 눈에 띄게(굵게·음영)
    wm_cell = WriteOnlyCell(ws, value=_watermark_text(watermark))
    wm_cell.font = _WATERMARK_FONT
    wm_cell.fill = _WATERMARK_FILL
    wm_cell.alignment = Alignment(vertical="center")
    ws.append([wm_cell])
    ws.append([])  # 워터마크와 헤더 사이 여백 1행

    # 2) 헤더행 — ColumnSpec.label, bold
    header_cells = []
    for col in columns:
        cell = WriteOnlyCell(ws, value=col.label)
        cell.font = Font(bold=True)
        header_cells.append(cell)
    ws.append(header_cells)

    # 3) 데이터행 — kind별 값·number_format
    for row in rows:
        ws.append([_data_cell(ws, col, row.get(col.key)) for col in columns])

    # 4) 합계행(옵션) — 첫 컬럼 "합계", 나머지는 값이 있으면 kind 서식으로 굵게
    if total_row is not None:
        cells = []
        for idx, col in enumerate(columns):
            if idx == 0:
                cell = WriteOnlyCell(ws, value="합계")
            else:
                cell = _data_cell(ws, col, total_row.get(col.key))
            cell.font = Font(bold=True)
            cells.append(cell)
        ws.append(cells)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_response(content: bytes, filename: str) -> Response:
    """.xlsx 다운로드 응답 — 미디어 타입 + 한글 파일명(RFC 5987) 공용 관용구."""
    return Response(
        content=content,
        media_type=_XLSX_MEDIA,
        headers={
            "Content-Disposition": "attachment; filename*=UTF-8''{0}".format(
                quote(filename)
            )
        },
    )


def export_filename(resource_label: str, on_date: Optional[date] = None) -> str:
    """내보내기 파일명 — `{리소스}_{YYYYMMDD}.xlsx` (기본 KST today)."""
    day = on_date or (datetime.utcnow() + _KST_OFFSET).date()
    return "{0}_{1}.xlsx".format(resource_label, day.strftime("%Y%m%d"))


def build_watermark(user) -> dict:
    """워터마크 dict 구성 — 내보낸 사람 이름·이메일 + 발급시각(KST 벽시계).

    finance·asset·audit export가 공유하는 단일 구성부(중복 금지). build_workbook의
    watermark 인자로 그대로 전달한다.
    """
    issued_at = (utcnow() + _KST_OFFSET).strftime("%Y-%m-%d %H:%M")
    return {"name": user.name, "email": user.email, "issued_at": issued_at}


def check_export_quota(
    db: Session, user, *, daily_limit: int = DAILY_EXPORT_LIMIT
) -> None:
    """일일 반출 횟수 가드(공용) — 오늘(KST) actor의 DATA_EXPORT 감사 건수가 한도 이상이면 429.

    감사로그를 카운터로 재사용(신규 테이블 없음). 호출부는 endpoint별 한도 상수를
    daily_limit로 넘겨(테스트 monkeypatch가 endpoint 모듈 상수를 그대로 반영) 재사용한다.
    """
    today = (utcnow() + _KST_OFFSET).date()
    day_start_utc = datetime.combine(today, time.min) - _KST_OFFSET
    day_end_utc = datetime.combine(today, time.max) - _KST_OFFSET
    used_today = (
        db.query(AuditLog)
        .filter(
            AuditLog.actor_id == user.user_id,
            AuditLog.action == "DATA_EXPORT",
            AuditLog.created_at >= day_start_utc,
            AuditLog.created_at <= day_end_utc,
        )
        .count()
    )
    if used_today >= daily_limit:
        raise HTTPException(
            status_code=429, detail="오늘 내보내기 횟수 한도를 초과했습니다"
        )


def enforce_row_limit(total: int, *, max_rows: int = MAX_EXPORT_ROWS) -> None:
    """행 상한 가드(공용) — 무음 잘라내기 금지, 초과 시 400(필터를 좁혀 재시도)."""
    if total > max_rows:
        raise HTTPException(
            status_code=400, detail="행이 많습니다 — 필터를 좁혀 주세요"
        )

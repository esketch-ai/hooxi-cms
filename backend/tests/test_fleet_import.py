"""운수사 계약대수 현황 — 원본 엑셀 파서·매칭·합산·upsert(F2)."""

from io import BytesIO

from openpyxl import Workbook

import models
from services import fleet_import


def _make_excel(rows, sheet_name="원본"):
    """원본 탭 형태(헤더 1~4행 복합, 데이터 5행부터)의 최소 xlsx 바이트 생성."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # 1행 빈 줄, 2~4행 헤더 흉내(값 무관 — 파서는 위치+숫자로 데이터행 판정)
    ws.append([])
    ws.append(["조합", "업종", "회사명", None, "월", "면허대수", "계"])
    ws.append([])
    ws.append([None, None, None, None, None, None, None, "경유", "CNG", "HB", "전기", "수소"])
    for r in rows:
        # A조합 B업종 C회사명 D키 E월 F면허 G계 H경유 I CNG J HB K전기 L수소
        ws.append([r["region"], r["industry"], r["company"], f"{r['region']}{r['company']}",
                   "6월", r["lic"], r["total"], r["diesel"], r["cng"], 0, r["ev"], 0])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cleanup(db):
    db.query(models.FleetStatus).filter(
        models.FleetStatus.company_name.like("TESTF%")).delete(synchronize_session=False)
    db.query(models.Client).filter(
        models.Client.company_name.like("TESTF%")).delete(synchronize_session=False)
    db.commit()


def test_parse_skips_headers_and_reads_positions():
    xls = _make_excel([
        {"region": "서울", "industry": "시내", "company": "TESTF경성",
         "lic": 83, "total": 83, "diesel": 0, "cng": 51, "ev": 24},
    ])
    rows = fleet_import.parse_rows(xls)
    assert len(rows) == 1
    r = rows[0]
    assert r["region"] == "서울" and r["company_name"] == "TESTF경성"
    assert r["license"] == 83 and r["cng"] == 51 and r["electric"] == 24


def test_analyze_match_and_aggregate(client):
    db = models.SessionLocal()
    try:
        _cleanup(db)
        c = models.Client(client_type="TRANSPORT", company_name="TESTF금강고속",
                          region="경기", biz_reg_no="880-88-88880")
        db.add(c)
        db.commit()
        cid = c.client_id
        # 다중 사업장(경기 2행) 합산 + 미매칭 1행
        xls = _make_excel([
            {"region": "경기", "industry": "시외", "company": "TESTF금강고속",
             "lic": 47, "total": 47, "diesel": 10, "cng": 0, "ev": 5},
            {"region": "경기", "industry": "시외", "company": "TESTF금강고속",
             "lic": 17, "total": 17, "diesel": 3, "cng": 0, "ev": 2},
            {"region": "강원", "industry": "농어촌", "company": "TESTF없는회사",
             "lic": 9, "total": 9, "diesel": 9, "cng": 0, "ev": 0},
        ])
        res = fleet_import.analyze(db, xls, "2026-06")
        assert res["total_rows"] == 3 and res["aggregated"] == 2
        assert res["matched"] == 1 and res["unmatched"] == 1
        matched = [it for it in res["items"] if it["matched"]][0]
        assert matched["matched_client_id"] == cid
        assert matched["license"] == 64 and matched["diesel"] == 13 and matched["electric"] == 7
        assert matched["industry"] == "INTERCITY"  # 시외→코드
        unm = [it for it in res["items"] if not it["matched"]][0]
        assert unm["matched_client_id"] is None and unm["region"] == "강원"
    finally:
        _cleanup(db)
        db.close()


def test_commit_upsert_and_reupload(client):
    db = models.SessionLocal()
    try:
        _cleanup(db)
        c = models.Client(client_type="TRANSPORT", company_name="TESTF성진",
                          region="부산", biz_reg_no="881-88-88881")
        db.add(c)
        db.commit()
        cid = c.client_id
        xls1 = _make_excel([
            {"region": "부산", "industry": "시내", "company": "TESTF성진",
             "lic": 50, "total": 50, "diesel": 0, "cng": 20, "ev": 30},
            {"region": "충북", "industry": "시내", "company": "TESTF보류사",
             "lic": 8, "total": 8, "diesel": 8, "cng": 0, "ev": 0},
        ])
        r1 = fleet_import.commit(db, xls1, "2026-06", actor_id="tester")
        assert r1["created"] == 2 and r1["updated"] == 0
        got = db.query(models.FleetStatus).filter_by(client_id=cid, period="2026-06").first()
        assert got.electric == 30 and got.source == "EXCEL"

        # 같은 월 재업로드(대수 변경) → upsert 갱신, 중복 생성 없음
        xls2 = _make_excel([
            {"region": "부산", "industry": "시내", "company": "TESTF성진",
             "lic": 55, "total": 55, "diesel": 0, "cng": 15, "ev": 40},
            {"region": "충북", "industry": "시내", "company": "TESTF보류사",
             "lic": 8, "total": 8, "diesel": 8, "cng": 0, "ev": 0},
        ])
        r2 = fleet_import.commit(db, xls2, "2026-06", actor_id="tester")
        assert r2["created"] == 0 and r2["updated"] == 2
        cnt = db.query(models.FleetStatus).filter_by(client_id=cid, period="2026-06").count()
        assert cnt == 1
        got2 = db.query(models.FleetStatus).filter_by(client_id=cid, period="2026-06").first()
        assert got2.electric == 40 and got2.cng == 15
        # 미매칭 보류도 재업로드 시 앱-레벨 dedup(region+회사명+period) → 1행 유지
        held = db.query(models.FleetStatus).filter_by(
            client_id=None, company_name="TESTF보류사", period="2026-06").count()
        assert held == 1
    finally:
        _cleanup(db)
        db.close()


def test_fleet_status_endpoints(client, admin_headers):
    db = models.SessionLocal()
    try:
        _cleanup(db)
        c = models.Client(client_type="TRANSPORT", company_name="TESTF엔드포인트",
                          region="대전", biz_reg_no="882-88-88882")
        db.add(c)
        db.commit()
        cid = c.client_id
    finally:
        db.close()
    h = admin_headers
    xls = _make_excel([
        {"region": "대전", "industry": "시내", "company": "TESTF엔드포인트",
         "lic": 40, "total": 40, "diesel": 0, "cng": 10, "ev": 30},
    ])
    # preview
    r = client.post("/api/v1/fleet-status/preview", headers=h,
                    data={"period": "2026-06"}, files={"file": ("f.xlsx", xls)})
    assert r.status_code == 200, r.text
    assert r.json()["matched"] == 1
    # 잘못된 월 형식 → 422
    rbad = client.post("/api/v1/fleet-status/preview", headers=h,
                       data={"period": "2026/6"}, files={"file": ("f.xlsx", xls)})
    assert rbad.status_code == 422
    # commit
    r2 = client.post("/api/v1/fleet-status/commit", headers=h,
                     data={"period": "2026-06"}, files={"file": ("f.xlsx", xls)})
    assert r2.status_code == 200 and r2.json()["created"] == 1
    # 현황 조회
    r3 = client.get(f"/api/v1/fleet-status/client/{cid}", headers=h)
    assert r3.status_code == 200
    assert r3.json()["trend"][0]["electric"] == 30
    # 수작업 관리 저장(업로드 독립)
    r4 = client.put(f"/api/v1/fleet-status/client/{cid}/mgmt", headers=h,
                    json={"contract_status": "DONE", "target_type": "BIZ", "memo": "테스트"})
    assert r4.status_code == 200 and r4.json()["contract_status"] == "DONE"
    r5 = client.get(f"/api/v1/fleet-status/client/{cid}", headers=h)
    assert r5.json()["mgmt"]["contract_status"] == "DONE"
    db = models.SessionLocal()
    try:
        db.query(models.FleetMgmt).filter_by(client_id=cid).delete(synchronize_session=False)
        db.commit()
        _cleanup(db)
    finally:
        db.close()


def test_dashboard_fleet_aggregation(client, admin_headers):
    db = models.SessionLocal()
    try:
        _cleanup(db)
        c = models.Client(client_type="TRANSPORT", company_name="TESTF대시운수",
                          region="서울", biz_reg_no="883-88-88883")
        db.add(c)
        db.commit()
        cid = c.client_id
        # 2개월치 + 미매칭 1건 + 수작업 관리
        db.add_all([
            models.FleetStatus(client_id=cid, region="서울", industry="CITY",
                               company_name="TESTF대시운수", period="2026-05",
                               license_count=100, total_count=100, electric=20),
            models.FleetStatus(client_id=cid, region="서울", industry="CITY",
                               company_name="TESTF대시운수", period="2026-06",
                               license_count=100, total_count=100, electric=35),
            models.FleetStatus(client_id=None, region="경기", industry="RURAL",
                               company_name="TESTF보류대시", period="2026-06",
                               license_count=10, total_count=10, electric=1),
            models.FleetMgmt(client_id=cid, target_type="BIZ", contract_status="DONE"),
        ])
        db.commit()
    finally:
        db.close()
    r = client.get("/api/v1/dashboard/fleet", headers=admin_headers)
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["period"] == "2026-06" and d["prev_period"] == "2026-05"
    assert d["companies"] == 2 and d["matched_companies"] == 1
    assert d["total_electric"] == 36  # 35 + 1
    assert d["ev_delta"] == 16  # 36(6월) - 20(5월)
    assert d["biz_target"] == 1 and d["contracted"] == 1
    db = models.SessionLocal()
    try:
        db.query(models.FleetMgmt).filter_by(client_id=cid).delete(synchronize_session=False)
        db.commit()
        _cleanup(db)
    finally:
        db.close()


def _make_excel_with_status(orig_rows, status_rows):
    """원본 탭 + 현황 탭(분류 컬럼) 2탭 xlsx. status_rows: (지역,회사,대상,계약,조합,규제)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "원본"
    ws.append([])
    ws.append(["조합", "업종", "회사명", None, "월", "면허대수", "계"])
    ws.append([])
    ws.append([None, None, None, None, None, None, None, "경유", "CNG", "HB", "전기", "수소"])
    for r in orig_rows:
        ws.append([r["region"], r["industry"], r["company"], f"{r['region']}{r['company']}",
                   "6월", r["lic"], r["total"], r["diesel"], r["cng"], 0, r["ev"], 0])
    st = wb.create_sheet("현황")
    st.append(["조합", "코드", "업종", "회사명", "대상여부", "월", "면허대수", "계",
               "경유", "CNG", "HB", "전기", "수소", "계약여부", "조합계약", "규제여부"])
    for (region, company, target, contract, union, regulated) in status_rows:
        st.append([region, 1, "시내", company, target, "6월", 0, 0, 0, 0, 0, 0, 0,
                   contract, union, regulated])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_status_tab_ingests_mgmt(client, admin_headers):
    db = models.SessionLocal()
    try:
        _cleanup(db)
        c = models.Client(client_type="TRANSPORT", company_name="TESTF분류운수",
                          region="서울", biz_reg_no="884-88-88884")
        db.add(c)
        db.commit()
        cid = c.client_id
    finally:
        db.close()
    xls = _make_excel_with_status(
        [{"region": "서울", "industry": "시내", "company": "TESTF분류운수",
          "lic": 50, "total": 50, "diesel": 0, "cng": 10, "ev": 40}],
        [("서울", "TESTF분류운수", "사업대상", "계약완료", "MOU체결", "")],
    )
    r = client.post("/api/v1/fleet-status/commit", headers=admin_headers,
                    data={"period": "2026-06"}, files={"file": ("f.xlsx", xls)})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 1 and body["mgmt_matched"] == 1
    # tb_fleet_mgmt에 코드로 반영됐는지
    r2 = client.get(f"/api/v1/fleet-status/client/{cid}", headers=admin_headers)
    m = r2.json()["mgmt"]
    assert m["target_type"] == "BIZ" and m["contract_status"] == "DONE"
    assert m["union_contract"] == "MOU"
    db = models.SessionLocal()
    try:
        db.query(models.FleetMgmt).filter_by(client_id=cid).delete(synchronize_session=False)
        db.commit()
        _cleanup(db)
    finally:
        db.close()


def test_dashboard_fleet_tables(client, admin_headers):
    db = models.SessionLocal()
    try:
        _cleanup(db)
        c1 = models.Client(client_type="TRANSPORT", company_name="TESTF표A", region="서울",
                           biz_reg_no="885-88-88885")
        c2 = models.Client(client_type="TRANSPORT", company_name="TESTF표B", region="부산",
                           biz_reg_no="886-88-88886")
        db.add_all([c1, c2]); db.commit()
        a, b = c1.client_id, c2.client_id
        db.add_all([
            models.FleetStatus(client_id=a, region="서울", industry="CITY",
                               company_name="TESTF표A", period="2026-06",
                               license_count=100, total_count=100, electric=30, hydrogen=5),
            models.FleetStatus(client_id=b, region="부산", industry="CITY",
                               company_name="TESTF표B", period="2026-06",
                               license_count=50, total_count=50, electric=10, hydrogen=0),
            # A=외부사업 미계약, B=규제대상
            models.FleetMgmt(client_id=a, target_type="BIZ", contract_status="NONE"),
            models.FleetMgmt(client_id=b, target_type="REG", contract_status="DONE",
                             regulated_type="ALLOC"),
        ])
        db.commit()
    finally:
        db.close()
    r = client.get("/api/v1/dashboard/fleet-tables", headers=admin_headers)
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["period"] == "2026-06"
    tbl = {t["key"]: t for t in d["tables"]}
    # T1 전체현황(대수) 전국 = 면허 150 / 전기 40 / 수소 5
    assert tbl["T1"]["total"]["c1"] == 150 and tbl["T1"]["total"]["c2"] == 40
    # T2 외부사업 대상(할당/목표 제외) = A만(BIZ, 규제없음) → 면허 100
    assert tbl["T2"]["total"]["c1"] == 100
    # T3 외부사업 미계약 = A(미계약) → 100
    assert tbl["T3"]["total"]["c1"] == 100
    # T5 규제/비규제: 소계2 / 규제1(B) / 외부사업1(A)
    assert tbl["T5"]["total"]["c1"] == 2 and tbl["T5"]["total"]["c2"] == 1 and tbl["T5"]["total"]["c3"] == 1
    db = models.SessionLocal()
    try:
        db.query(models.FleetMgmt).filter(models.FleetMgmt.client_id.in_([a, b])).delete(synchronize_session=False)
        db.commit(); _cleanup(db)
    finally:
        db.close()


def test_status_tab_dedup_same_client_no_violation(client, admin_headers):
    """현황 탭에서 여러 행이 같은 고객사에 매칭돼도 tb_fleet_mgmt PK 위반 없이 1건으로 병합."""
    db = models.SessionLocal()
    try:
        _cleanup(db)
        c = models.Client(client_type="TRANSPORT", company_name="TESTF중복운수",
                          region="서울", biz_reg_no="887-88-88887")
        db.add(c)
        db.commit()
        cid = c.client_id
    finally:
        db.close()
    # 같은 (지역+회사명) 현황 행 2개 → 같은 고객사 매칭
    xls = _make_excel_with_status(
        [{"region": "서울", "industry": "시내", "company": "TESTF중복운수",
          "lic": 10, "total": 10, "diesel": 0, "cng": 0, "ev": 10}],
        [("서울", "TESTF중복운수", "사업대상", "미계약", "", ""),
         ("서울", "TESTF중복운수", "사업대상", "계약완료", "MOU체결", "")],
    )
    r = client.post("/api/v1/fleet-status/commit", headers=admin_headers,
                    data={"period": "2026-06"}, files={"file": ("dup.xlsx", xls)})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["mgmt_matched"] == 2 and body["mgmt_created"] == 1  # 2행 매칭 → 1건 병합
    cnt = db.query(models.FleetMgmt).filter_by(client_id=cid).count()
    assert cnt == 1  # PK 위반 없이 단일 행
    m = db.get(models.FleetMgmt, cid)
    assert m.contract_status == "DONE"  # 마지막 행 반영
    db2 = models.SessionLocal()
    try:
        db2.query(models.FleetMgmt).filter_by(client_id=cid).delete(synchronize_session=False)
        db2.commit(); _cleanup(db2)
    finally:
        db2.close()
    db.close()


def test_match_key_normalizes_region_and_whitespace():
    """매칭 정규화 — 광주/전남↔전남광주 지역, 회사명 줄바꿈·다중공백 흡수(저장은 원문)."""
    mk = fleet_import._match_key
    # 지역 정규화: 원본 광주/전남 → 마스터 전남광주 버킷
    assert mk("광주", "대창운수") == mk("전남광주", "대창운수")
    assert mk("전남", "나주교통") == mk("전남광주", "나주교통")
    # 회사명 공백/줄바꿈 제거
    assert mk("경북", "새천년\n미소") == mk("경북", "새천년미소")
    assert mk("전북", "무진장     여객") == mk("전북", "무진장여객")
    # 다른 지역은 그대로(오병합 방지)
    assert mk("부산", "삼성여객") != mk("서울", "삼성여객")


def test_fleet_status_match_region_alias(client):
    """원본 지역이 광주여도 마스터 전남광주 운수사에 매칭된다."""
    db = models.SessionLocal()
    try:
        _cleanup(db)
        c = models.Client(client_type="TRANSPORT", company_name="TESTF대창운수",
                          region="전남광주", biz_reg_no="889-88-88889")
        db.add(c); db.commit()
        cid = c.client_id
        xls = _make_excel([
            {"region": "광주", "industry": "시내", "company": "TESTF대창운수",
             "lic": 50, "total": 50, "diesel": 0, "cng": 0, "ev": 20},
        ])
        res = fleet_import.analyze(db, xls, "2026-06")
        assert res["matched"] == 1
        assert res["items"][0]["matched_client_id"] == cid
    finally:
        _cleanup(db)
        db.close()

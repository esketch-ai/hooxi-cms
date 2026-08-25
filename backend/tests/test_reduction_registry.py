"""감축 참여 레지스트리(KISA) — 파서·적재·목록·요약(M3)."""

import io

import openpyxl

import models
from services import reduction_registry_import as rri

API = "/api/v1/reduction-registry"


def _kisa_xlsx():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    b = wb.create_sheet("1. 베이스라인 제원(대체 전 화석연료버스)")
    b.append(["차량번호", "업체명", "순번", "사업구분", "차명", "차대번호", "연식", "차종",
              "길이(mm)", "너비(mm)", "높이(mm)", "총중량(kg)", "승차정원", "연료", "사업명", "권역"])
    b.append(["경북70자7304", "경산버스", 1, "대체도입", "뉴슈퍼", "OLDVIN001", 2012, "대형 승합",
              10955, 2490, 3180, 14260, 59, "경유", "후시005", "경북"])
    p = wb.create_sheet("2. 사업대상 제원(대체도입 전기버스)")
    p.append(["차량번호", "업체명", "순번", "사업구분", "차명", "차대번호", "연식", "전기차량 등록일",
              "차종", "길이(mm)", "너비(mm)", "높이(mm)", "총중량(kg)", "승차정원", "연료", "배터리종류"])
    p.append(["경북70자7304", "경산버스", 1, "대체도입", "일렉시티", "NEWVIN001", 2024, "2024-08-29",
              "대형 승합", 10995, 2490, 3400, 15500, 50, "전기", "리튬 폴리머"])
    c = wb.create_sheet("3. 대체 예정 화석연료버스")
    c.append(["자동차등록번호", "업체명", "순번", "차명", "연식", "등록일", "차종", "용도",
              "길이(mm)", "너비(mm)", "높이(mm)", "총중량(kg)", "승차정원", "연료", "사업명", "권역"])
    c.append(["충남70자1201", "보성여객", 200, "현대그린시티", 2015, "2015-10-14", "대형 승합", "영업용",
              9085, 2490, 3225, 10850, 21, "CNG", "NWK002", "충남"])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def test_parser_roles():
    rows = rri.parse_registry(_kisa_xlsx())
    roles = sorted(r["role"] for r in rows)
    assert roles == ["BASELINE", "CANDIDATE", "PROJECT"]
    proj = next(r for r in rows if r["role"] == "PROJECT")
    assert proj["vehicle_no"] == "경북70자7304" and proj["vin"] == "NEWVIN001"
    assert proj["battery_type"] == "리튬 폴리머"
    cand = next(r for r in rows if r["role"] == "CANDIDATE")
    assert cand["purpose"] == "영업용" and cand.get("vin") is None
    assert str(cand["registered_at"]) == "2015-10-14"


def test_import_list_summary_and_idempotent(client, staff_headers):
    db = models.SessionLocal()
    try:
        db.query(models.ReductionRegistry).delete(synchronize_session=False)
        db.commit()
    finally:
        db.close()

    files = {"file": ("kisa.xlsx", _kisa_xlsx(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post(API + "/import", headers=staff_headers, files=files)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["baseline"] == 1 and body["project"] == 1 and body["candidate"] == 1

    # 재적재 멱등(KISA_IMPORT 교체) — 여전히 3건
    files2 = {"file": ("kisa.xlsx", _kisa_xlsx(), files["file"][2])}
    client.post(API + "/import", headers=staff_headers, files=files2)
    summ = client.get(API + "/summary", headers=staff_headers).json()
    assert summ["total"] == 3 and summ["candidate"] == 1

    # 후보(미참여)만 필터
    cand = client.get(API, headers=staff_headers, params={"role": "CANDIDATE"}).json()
    assert cand["total"] == 1 and cand["items"][0]["vehicle_no"] == "충남70자1201"

    db = models.SessionLocal()
    try:
        db.query(models.ReductionRegistry).delete(synchronize_session=False)
        db.commit()
    finally:
        db.close()


def test_requires_auth(client):
    assert client.get(API).status_code == 401


def test_replacement_verification(client, staff_headers):
    db = models.SessionLocal()
    try:
        db.query(models.ReductionRegistry).delete(synchronize_session=False)
        # PASS 케이스: 같은 차량번호, VIN 상이, 경유→전기
        db.add(models.ReductionRegistry(role="BASELINE", vehicle_no="강원70자1", introduction_type="대체도입",
                                        vin="OLD1", fuel="경유", operator_name="A운수"))
        db.add(models.ReductionRegistry(role="PROJECT", vehicle_no="강원70자1", introduction_type="대체도입",
                                        vin="NEW1", fuel="전기", operator_name="A운수"))
        # FAIL 케이스: 베이스라인 없음
        db.add(models.ReductionRegistry(role="PROJECT", vehicle_no="강원70자2", introduction_type="대체도입",
                                        vin="NEW2", fuel="전기", operator_name="A운수"))
        # FAIL 케이스: VIN 동일
        db.add(models.ReductionRegistry(role="BASELINE", vehicle_no="강원70자3", introduction_type="대체도입",
                                        vin="SAME", fuel="CNG", operator_name="A운수"))
        db.add(models.ReductionRegistry(role="PROJECT", vehicle_no="강원70자3", introduction_type="대체도입",
                                        vin="SAME", fuel="전기", operator_name="A운수"))
        db.commit()
    finally:
        db.close()

    v = client.get("/api/v1/reduction-registry/verification", headers=staff_headers).json()
    assert v["total"] == 3 and v["passed"] == 1 and v["failed"] == 2
    byno = {i["vehicle_no"]: i for i in v["items"]}
    assert byno["강원70자1"]["status"] == "PASS"
    assert "베이스라인 없음" in byno["강원70자2"]["reasons"]
    assert "VIN 동일" in byno["강원70자3"]["reasons"]

    failed = client.get("/api/v1/reduction-registry/verification", headers=staff_headers,
                        params={"only_failed": True}).json()
    assert len(failed["items"]) == 2

    db = models.SessionLocal()
    try:
        db.query(models.ReductionRegistry).delete(synchronize_session=False); db.commit()
    finally:
        db.close()

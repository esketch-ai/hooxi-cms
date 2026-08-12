"""감사 로그 엑셀 내보내기(EX-5, ADMIN) — '필터=파일' 정합 + R2-E6 redact-only.

검증: 전체 필터결과 반출(목록 total == 데이터 행수)·필터 반영·ADMIN 게이트(MANAGER/STAFF 403)·
외부 403·**export 파일에 비밀/토큰/평문 없음(저장 redact값만)**·DATA_EXPORT 감사(target_type AUDIT_LOG).
목록 회귀는 기존 audit 테스트가 담당(_apply_filters 추출은 목록 결과 불변).
"""

from io import BytesIO

from openpyxl import load_workbook

import models
from routers import audit as au
from services.audit_logger import AuditLogger

API = "/api/v1"
AUDIT = API + "/audit-logs"
EXPORT = AUDIT + "/export"

_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _seed_audit(action, target_type, target_id=None, new_value=None):
    """감사 로그 1건 시드 — AuditLogger 경유(적재 시 redact 규약 그대로)."""
    db = models.SessionLocal()
    try:
        AuditLogger.log_action(
            db, "u-admin", action,
            target_type=target_type, target_id=target_id, new_value=new_value,
        )
        db.commit()
    finally:
        db.close()


def _list_total(client, headers, **params):
    r = client.get(AUDIT, headers=headers, params=params)
    assert r.status_code == 200, r.text
    return r.json()["total"]


def _load_sheet(content):
    return load_workbook(BytesIO(content)).active


def _data_rows(ws):
    """데이터 행 목록 — [워터마크][빈행][헤더] 이후 전부(감사 export는 합계행 없음)."""
    rows = [
        [c.value for c in row]
        for row in ws.iter_rows()
        if any(c.value is not None for c in row)
    ]
    # rows[0]=워터마크, rows[1]=헤더, 이후 데이터(합계 없음)
    return rows[2:]


def _login_role(client, user_id, email, role, status="ACTIVE"):
    """내부/외부 임의 역할 계정 upsert 후 dev-login 토큰."""
    db = models.SessionLocal()
    try:
        u = db.get(models.User, user_id)
        if u is None:
            u = models.User(user_id=user_id, email=email, name=email.split("@")[0])
            db.add(u)
        u.role = role
        u.status = status
        db.commit()
    finally:
        db.close()
    tok = client.post(API + "/auth/dev-login", json={"email": email})
    assert tok.status_code == 200, tok.text
    return {"Authorization": "Bearer {0}".format(tok.json()["access_token"])}


# ── 1) 200·헤더·전체행·워터마크·합계없음 ────────────────────────────────────
def test_export_ok_rows_match_list_total(client, admin_headers):
    tt = "AUDITEXP_OK"
    for i in range(3):
        _seed_audit("CONFIG_CHANGE", tt, target_id="cfg{0}".format(i), new_value="v{0}".format(i))
    r = client.get(EXPORT, headers=admin_headers, params={"target_type": tt})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(_XLSX)
    assert "filename*=UTF-8''" in r.headers["content-disposition"]

    ws = _load_sheet(r.content)
    data = _data_rows(ws)
    # 데이터 행수 == 동일 필터 목록 total(전체행, 페이지네이션 없음)
    assert len(data) == _list_total(client, admin_headers, target_type=tt) == 3

    # 워터마크에 내보낸 사람(이름) 포함 — conftest admin name('관리자')
    wm = ws.cell(row=1, column=1).value
    assert wm and "관리자" in wm

    # 합계행 없음 — 마지막 데이터 행 첫 셀이 '합계'가 아니어야
    assert data[-1][0] != "합계"


# ── 필터 반영(target_type) ──────────────────────────────────────────────────
def test_export_reflects_filters(client, admin_headers):
    a, b = "AUDITEXP_FA", "AUDITEXP_FB"
    _seed_audit("REPORT_VIEW", a, target_id="r1")
    _seed_audit("REPORT_VIEW", a, target_id="r2")
    _seed_audit("REPORT_VIEW", b, target_id="r3")

    r = client.get(EXPORT, headers=admin_headers, params={"target_type": a})
    assert r.status_code == 200
    assert len(_data_rows(_load_sheet(r.content))) == _list_total(
        client, admin_headers, target_type=a
    ) == 2


# ── 인가(목록과 동일 ADMIN 게이트) ──────────────────────────────────────────
def test_export_authz(client, manager_headers, staff_headers):
    # MANAGER 403(목록도 ADMIN 전용)
    assert client.get(EXPORT, headers=manager_headers).status_code == 403
    # STAFF 403
    assert client.get(EXPORT, headers=staff_headers).status_code == 403
    # 외부(INVESTOR) 403 — get_current_user 원천 차단
    inv = _login_role(client, "u-auexp-investor", "auexp-investor@fund.example", "INVESTOR")
    assert client.get(EXPORT, headers=inv).status_code == 403
    # 미인증 401
    assert client.get(EXPORT).status_code == 401


# ── R2-E6 redact-only: 파일에 비밀/토큰/평문 없음(저장 redact값만) ──────────
def test_export_redact_only(client, admin_headers):
    tt = "AUDITEXP_SECRET"
    # 적재 시 SENSITIVE_KEYWORDS로 redact → 저장값은 '[REDACTED]'(평문 미저장)
    _seed_audit("INTEGRATION_REVEAL", tt, target_id="int1",
                new_value="api_key=SUPERSECRET_PLAINTEXT_9911 token=abcXYZ")
    r = client.get(EXPORT, headers=admin_headers, params={"target_type": tt})
    assert r.status_code == 200

    text = "\n".join(
        str(c.value)
        for row in _load_sheet(r.content).iter_rows()
        for c in row
        if c.value is not None
    )
    # 저장된 redact값은 그대로 노출, 평문 비밀은 파일 어디에도 없어야(export가 재조회·복호화 안 함)
    assert "[REDACTED]" in text
    assert "SUPERSECRET_PLAINTEXT_9911" not in text
    assert "abcXYZ" not in text


# ── 감사(DATA_EXPORT, target_type AUDIT_LOG) ────────────────────────────────
def test_export_audit_meta(client, admin_headers):
    tt = "AUDITEXP_META"
    _seed_audit("CONFIG_CHANGE", tt, target_id="m1")
    r = client.get(EXPORT, headers=admin_headers, params={"target_type": tt})
    assert r.status_code == 200

    db = models.SessionLocal()
    try:
        logs = (
            db.query(models.AuditLog)
            .filter(models.AuditLog.action == "DATA_EXPORT")
            .filter(models.AuditLog.new_value.like("%target_type={0}%".format(tt)))
            .all()
        )
        assert len(logs) == 1
        log = logs[0]
        assert log.target_type == "AUDIT_LOG"
        assert log.new_value.startswith("rows=")
    finally:
        db.close()


# ── 행 상한(무음 잘라내기 금지) ─────────────────────────────────────────────
def test_export_row_cap(client, admin_headers, monkeypatch):
    tt = "AUDITEXP_CAP"
    _seed_audit("CONFIG_CHANGE", tt, target_id="c1")
    monkeypatch.setattr(au, "MAX_EXPORT_ROWS", 0)
    r = client.get(EXPORT, headers=admin_headers, params={"target_type": tt})
    assert r.status_code == 400, r.text


# ── 일일 반출 횟수 제한 ─────────────────────────────────────────────────────
def test_export_daily_limit(client, admin_headers, monkeypatch):
    monkeypatch.setattr(au, "DAILY_EXPORT_LIMIT", 0)
    r = client.get(EXPORT, headers=admin_headers, params={"target_type": "무관"})
    assert r.status_code == 429, r.text

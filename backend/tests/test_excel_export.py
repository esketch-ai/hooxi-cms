"""엑셀 내보내기 공용 헬퍼(EX-1) 테스트 — 워터마크·셀타입·합계·응답·파일명.

build_workbook 결과 bytes를 openpyxl로 재로드해 실제 셀 값·number_format을
검증한다(직렬화 왕복 보장).
"""

from datetime import date, datetime
from io import BytesIO
from urllib.parse import quote

from openpyxl import load_workbook

from services.excel_export import (
    ColumnSpec,
    build_workbook,
    export_filename,
    xlsx_response,
    _XLSX_MEDIA,
)


def _load(content: bytes):
    return load_workbook(BytesIO(content))


def _columns():
    return [
        ColumnSpec(key="name", label="이름", kind="text"),
        ColumnSpec(key="paid_at", label="지급일", kind="date"),
        ColumnSpec(key="amount", label="금액", kind="money"),
        ColumnSpec(key="rate", label="비율", kind="percent"),
    ]


def test_watermark_block_contains_name_and_email():
    content = build_workbook(
        _columns(),
        rows=[],
        sheet_title="원장",
        watermark={
            "name": "홍길동",
            "email": "hong@hooxi.com",
            "issued_at": "2026-08-13 10:00",
        },
    )
    ws = _load(content).active
    wm = ws.cell(row=1, column=1).value
    assert "홍길동" in wm
    assert "hong@hooxi.com" in wm
    assert "무단 외부유출 금지" in wm


def test_header_labels_match():
    cols = _columns()
    content = build_workbook(cols, rows=[], sheet_title="원장", watermark={})
    ws = _load(content).active
    # 워터마크(1행)+여백(2행) 다음 헤더행은 3행
    labels = [ws.cell(row=3, column=i + 1).value for i in range(len(cols))]
    assert labels == [c.label for c in cols]


def test_cell_types_and_formats():
    cols = _columns()
    rows = [
        {
            "name": "가맹점A",
            "paid_at": date(2026, 8, 1),
            "amount": 1234567,
            "rate": 0.125,
        }
    ]
    content = build_workbook(cols, rows=rows, sheet_title="원장", watermark={})
    ws = _load(content).active
    data_row = 4  # 헤더 3행 다음
    name_c = ws.cell(row=data_row, column=1)
    date_c = ws.cell(row=data_row, column=2)
    money_c = ws.cell(row=data_row, column=3)
    pct_c = ws.cell(row=data_row, column=4)

    assert name_c.value == "가맹점A"
    # openpyxl은 date 값을 재로드 시 datetime으로 왕복 — 날짜 부분·서식으로 검증
    assert isinstance(date_c.value, (date, datetime))
    dv = date_c.value.date() if isinstance(date_c.value, datetime) else date_c.value
    assert dv == date(2026, 8, 1)
    assert date_c.number_format == "yyyy-mm-dd"
    assert money_c.value == 1234567
    assert isinstance(money_c.value, (int, float))
    assert money_c.number_format == "#,##0"
    assert pct_c.value == 0.125
    assert pct_c.number_format == "0.0%"


def test_none_cell_is_empty():
    cols = _columns()
    rows = [{"name": None, "paid_at": None, "amount": None, "rate": None}]
    content = build_workbook(cols, rows=rows, sheet_title="원장", watermark={})
    ws = _load(content).active
    for col in range(1, len(cols) + 1):
        assert ws.cell(row=4, column=col).value is None


def test_total_row():
    cols = _columns()
    rows = [{"name": "A", "paid_at": None, "amount": 100, "rate": 0.1}]
    content = build_workbook(
        cols,
        rows=rows,
        sheet_title="원장",
        watermark={},
        total_row={"amount": 100},
    )
    ws = _load(content).active
    total_row = 5  # 워터마크1+여백2+헤더3+데이터4 → 합계 5행
    assert ws.cell(row=total_row, column=1).value == "합계"
    money_c = ws.cell(row=total_row, column=3)
    assert money_c.value == 100
    assert money_c.number_format == "#,##0"


def test_export_filename_rule():
    assert export_filename("재무원장", on_date=date(2026, 8, 13)) == "재무원장_20260813.xlsx"


def test_export_filename_default_kst():
    name = export_filename("재무원장")
    assert name.startswith("재무원장_")
    assert name.endswith(".xlsx")
    # YYYYMMDD 8자리
    stamp = name[len("재무원장_") : -len(".xlsx")]
    assert len(stamp) == 8 and stamp.isdigit()


def test_xlsx_response_media_and_disposition():
    resp = xlsx_response(b"dummy", "재무원장_20260813.xlsx")
    assert resp.media_type == _XLSX_MEDIA
    disp = resp.headers["content-disposition"]
    assert disp.startswith("attachment; filename*=UTF-8''")
    assert quote("재무원장_20260813.xlsx") in disp

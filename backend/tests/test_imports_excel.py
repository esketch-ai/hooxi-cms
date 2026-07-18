"""엑셀 일괄 등록 파서·검증 (services/excel_import.py) — 서비스 단위 테스트.

라운드트립·헤더 순서 무관·라벨/코드 양방향·숫자 셀·중복·빈 행·행 상한.
"""

from io import BytesIO

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

import models
from services import excel_import
from services.import_spec import IMPORT_SPECS, MAX_IMPORT_ROWS


@pytest.fixture()
def db(client):
    session = models.SessionLocal()
    yield session
    session.close()


def _xlsx(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_client(db, name, biz=None):
    row = models.Client(client_type="TRANSPORT", company_name=name, biz_reg_no=biz)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def _label(db, category, code):
    """현재 활성 라벨 조회 — 다른 테스트가 라벨을 바꿔도(운영에서도 가능) 견고하게."""
    return (
        db.query(models.Code)
        .filter(models.Code.category == category, models.Code.code == code)
        .one()
        .label
    )


# ---------------------------------------------------------------------------
# 라운드트립 — 양식 그대로 업로드하면 전 행 통과 (양식·파서 정합)
# ---------------------------------------------------------------------------
def _strip_example_prefix(content: bytes) -> bytes:
    """예시 행 접두를 제거해 '실무자가 예시를 실값으로 바꾼 파일'을 시뮬레이션."""
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content))
    ws = wb.worksheets[0]
    first = ws.cell(row=2, column=1).value
    ws.cell(row=2, column=1, value=str(first).replace(excel_import.EXAMPLE_PREFIX, "", 1))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_template_roundtrip_clients(db):
    content = excel_import.build_template(db, "clients")
    # 예시 행은 그대로 올리면 스킵된다 — 실데이터 오등록 함정 방지
    result = excel_import.parse_and_validate(db, "clients", content)
    assert result.rows == []
    assert any("예시 행" in w for w in result.warnings)
    # 접두를 지운(=실값으로 바꾼) 행은 정상 통과 — 양식↔파서 정합
    result = excel_import.parse_and_validate(db, "clients", _strip_example_prefix(content))
    assert len(result.rows) == 1
    row = result.rows[0]
    assert row.errors == [], row.errors
    assert row.payload is not None
    # 예시 행의 라벨 값이 코드·ID로 정규화됐는지 (운수사→TRANSPORT, 관리자→u-admin)
    assert row.payload.client_type == "TRANSPORT"
    assert row.payload.contract_status == "ACTIVE"
    assert row.payload.manager_id == "u-admin"
    assert row.payload.report_yn == "Y"


def test_template_roundtrip_assets(db):
    _add_client(db, "한빛운수")  # 양식 예시 행의 고객사명
    content = excel_import.build_template(db, "assets")
    result = excel_import.parse_and_validate(db, "assets", _strip_example_prefix(content))
    assert len(result.rows) == 1
    row = result.rows[0]
    assert row.errors == [], row.errors
    assert row.payload.asset_group == "MOBILITY"  # 모빌리티 라벨 → 코드
    assert row.payload.auth_type == "ID_PW"  # 아이디/비밀번호 → 고정값 매핑
    assert row.payload.client_id  # 이름 → client_id 해석
    # 인증 비밀값 컬럼은 양식에 존재하지 않는다 (평문 유통 차단)
    wb = load_workbook(BytesIO(content))
    headers = [c.value for c in wb.active[1]]
    assert not any(h and ("비밀번호" in h or "토큰" in h) for h in headers)


# ---------------------------------------------------------------------------
# 헤더 매칭 — 순서 무관·미지 컬럼 경고·필수 컬럼 누락 422
# ---------------------------------------------------------------------------
def test_header_shuffle_and_unknown_column(db):
    content = _xlsx(
        ["비고", "구분", "회사명 *"],
        [["메모", "운수사", "순서무관상사"]],
    )
    result = excel_import.parse_and_validate(db, "clients", content)
    assert result.unknown_columns == ["비고"]
    assert len(result.valid_rows) == 1
    assert result.valid_rows[0].payload.company_name == "순서무관상사"


def test_missing_required_column_422(db):
    content = _xlsx(["회사명"], [["구분없는상사"]])  # 구분(필수) 헤더 없음
    with pytest.raises(HTTPException) as exc:
        excel_import.parse_and_validate(db, "clients", content)
    assert exc.value.status_code == 422
    assert "구분" in exc.value.detail


def test_not_xlsx_422(db):
    with pytest.raises(HTTPException) as exc:
        excel_import.parse_and_validate(db, "clients", b"this is not xlsx")
    assert exc.value.status_code == 422


# ---------------------------------------------------------------------------
# 값 정규화 — 라벨/코드 양방향·숫자 셀·Y/N·날짜
# ---------------------------------------------------------------------------
def test_code_label_and_code_both_accepted(db):
    content = _xlsx(
        ["회사명", "구분", "계약 상태"],
        [
            # 라벨은 현재 활성 라벨을 조회해 사용 — 라벨 변경(운영 가능)에도 견고
            ["라벨입력상사", _label(db, "CLIENT_TYPE", "TRANSPORT"),
             _label(db, "CONTRACT_STATUS", "ACTIVE")],
            ["코드입력상사", "TRANSPORT", "HOLD"],
            ["오류입력상사", "우주선", None],
        ],
    )
    result = excel_import.parse_and_validate(db, "clients", content)
    ok = {r.payload.company_name: r.payload for r in result.valid_rows}
    assert ok["라벨입력상사"].client_type == "TRANSPORT"
    assert ok["코드입력상사"].client_type == "TRANSPORT"
    assert ok["코드입력상사"].contract_status == "HOLD"
    bad = [r for r in result.rows if r.errors]
    assert len(bad) == 1 and "우주선" in bad[0].errors[0]


def test_numeric_cells_normalized(db):
    c = _add_client(db, "숫자셀운수")
    content = _xlsx(
        ["회사명", "구분", "사업자번호", "대표 연락처"],
        [["숫자셀상사", "운수사", 1234509876, 1055556666.0]],
    )
    result = excel_import.parse_and_validate(db, "clients", content)
    p = result.valid_rows[0].payload
    assert p.biz_reg_no == "1234509876"  # 3.0 꼬리 제거·문자화
    assert p.ceo_contact_phone == "1055556666"

    content = _xlsx(
        ["고객사명", "대분류", "수량"],
        [[c.company_name, "모빌리티", 12.0]],
    )
    result = excel_import.parse_and_validate(db, "assets", content)
    assert result.valid_rows[0].payload.quantity == 12


def test_yn_and_date_normalization(db):
    content = _xlsx(
        ["회사명", "구분", "월간 보고서 수신", "계약 일자"],
        [
            ["와이엔상사", "운수사", "예", "2026.01.15"],
            ["엑스상사", "운수사", "X", None],
            ["와이엔오류상사", "운수사", "ㅁ", None],
        ],
    )
    result = excel_import.parse_and_validate(db, "clients", content)
    ok = {r.payload.company_name: r.payload for r in result.valid_rows}
    assert ok["와이엔상사"].report_yn == "Y"
    assert ok["와이엔상사"].contract_date.strftime("%Y-%m-%d") == "2026-01-15"
    assert ok["엑스상사"].report_yn == "N"
    bad = [r for r in result.rows if r.errors]
    assert len(bad) == 1 and "Y/N" in bad[0].errors[0]


def test_resolver_errors(db):
    content = _xlsx(
        ["회사명", "구분", "담당 PM"],
        [["피엠오류상사", "운수사", "없는사람"]],
    )
    result = excel_import.parse_and_validate(db, "clients", content)
    assert result.valid_rows == []
    assert "없는사람" in result.rows[0].errors[0]

    content = _xlsx(["고객사명", "대분류"], [["없는고객사", "모빌리티"]])
    result = excel_import.parse_and_validate(db, "assets", content)
    assert result.valid_rows == []
    assert "없는고객사" in result.rows[0].errors[0]


# ---------------------------------------------------------------------------
# 중복 — 파일 내(오류/경고)·DB(사업자번호 409 규칙 재사용)
# ---------------------------------------------------------------------------
def test_duplicate_biz_no_in_file_and_db(db):
    _add_client(db, "기존등록상사", biz="777-88-99990")
    content = _xlsx(
        ["회사명", "구분", "사업자번호"],
        [
            ["중복일호", "운수사", "555-66-77770"],
            ["중복이호", "운수사", "5556677770"],  # 표기 달라도 정규화 비교
            ["디비중복상사", "운수사", "7778899990"],  # DB 기존과 충돌
        ],
    )
    result = excel_import.parse_and_validate(db, "clients", content)
    assert len(result.valid_rows) == 1
    assert result.rows[1].errors and "파일 내 중복" in result.rows[1].errors[0]
    assert result.rows[2].errors and "이미 등록된 사업자번호" in result.rows[2].errors[0]
    assert "기존등록상사" in result.rows[2].errors[0]


def test_duplicate_warnings(db):
    c = _add_client(db, "경고중복운수")
    content = _xlsx(
        ["회사명", "구분"],
        [["같은이름상사", "운수사"], ["같은이름상사", "운수사"]],
    )
    result = excel_import.parse_and_validate(db, "clients", content)
    assert len(result.valid_rows) == 2  # 회사명 중복은 경고만 (단건 등록도 허용)
    assert result.rows[1].warnings and "중복 의심" in result.rows[1].warnings[0]

    content = _xlsx(
        ["고객사명", "대분류", "주요 제원"],
        [[c.company_name, "모빌리티", "1톤 트럭"], [c.company_name, "모빌리티", "1톤 트럭"]],
    )
    result = excel_import.parse_and_validate(db, "assets", content)
    assert len(result.valid_rows) == 2
    assert result.rows[1].warnings and "중복 의심" in result.rows[1].warnings[0]


# ---------------------------------------------------------------------------
# 빈 행 스킵·행 상한
# ---------------------------------------------------------------------------
def test_blank_rows_skipped(db):
    content = _xlsx(
        ["회사명", "구분"],
        [
            ["빈행일호", "운수사"],
            [None, None],
            ["", "  "],
            ["빈행이호", "운수사"],
        ],
    )
    result = excel_import.parse_and_validate(db, "clients", content)
    assert len(result.rows) == 2
    # 엑셀 실제 행 번호 유지 (빈 행을 건너뛰어도 원본 위치 안내)
    assert [r.row for r in result.rows] == [2, 5]


def test_row_limit_422(db):
    rows = [["상한{0}".format(i), "운수사"] for i in range(MAX_IMPORT_ROWS + 1)]
    content = _xlsx(["회사명", "구분"], rows)
    with pytest.raises(HTTPException) as exc:
        excel_import.parse_and_validate(db, "clients", content)
    assert exc.value.status_code == 422
    assert str(MAX_IMPORT_ROWS) in exc.value.detail


def test_validation_error_translated_to_korean_label(db):
    content = _xlsx(
        ["회사명", "구분", "대표 이메일"],
        [["이메일오류상사", "운수사", "notanemail"]],
    )
    result = excel_import.parse_and_validate(db, "clients", content)
    assert result.valid_rows == []
    err = result.rows[0].errors[0]
    assert err.startswith("대표 이메일:")  # 스펙 라벨 기준 번역
    assert "이메일 형식" in err

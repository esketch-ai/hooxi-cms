"""전기버스 자산 엑셀 내보내기(EX-4) — '필터=파일' 정합 + 균형 보안.

검증: 전체 필터결과 반출(목록 total == 데이터 행수)·필터 반영·MANAGER 게이트(조회보다 좁게)·
행 상한(400)·일일 반출 횟수(429)·워터마크(내보낸 사람)·DATA_EXPORT 감사(금액 원문 미기록).
목록 회귀는 test_asset_vehicles.py가 담당(_apply_filters 추출은 목록 결과 불변).
"""

from io import BytesIO

from openpyxl import load_workbook

import models
from routers import asset_vehicles as av

API = "/api/v1"
PROJECTS = API + "/projects"
ASSET_VEHICLES = API + "/asset-vehicles"
EXPORT = ASSET_VEHICLES + "/export"

_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── 시드 헬퍼(test_asset_vehicles.py와 동일 패턴) ────────────────────────────
def _mk_client(client, headers, tag):
    r = client.post(
        API + "/clients", headers=headers,
        json={"client_type": "TRANSPORT", "company_name": "반출자산운수" + tag},
    )
    assert r.status_code == 201, r.text
    return r.json()["client_id"]


def _mk_project(client, headers, name, approval_status=None):
    """프로젝트 생성(운수사 배정은 차량 client_id로) — approval_status 옵션."""
    body = {"project_name": name, "project_status": "기획"}
    if approval_status:
        body["approval_status"] = approval_status
    r = client.post(PROJECTS, headers=headers, json=body)
    assert r.status_code == 201, r.text
    return r.json()["project_id"]


def _add_vehicle(client, headers, pid, **fields):
    r = client.post(f"{PROJECTS}/{pid}/vehicles", headers=headers, json=fields)
    assert r.status_code == 201, r.text
    return r.json()["vehicle_id"]


def _list_total(client, headers, **params):
    r = client.get(ASSET_VEHICLES, headers=headers, params=params)
    assert r.status_code == 200, r.text
    return r.json()["total"]


def _load_sheet(content):
    return load_workbook(BytesIO(content)).active


def _data_rows(ws):
    """데이터 행 목록 — [워터마크][빈행][헤더] 이후, 마지막 '합계' 행 제외."""
    rows = [
        [c.value for c in row]
        for row in ws.iter_rows()
        if any(c.value is not None for c in row)
    ]
    # rows[0]=워터마크, rows[1]=헤더, rows[-1]=합계
    return rows[2:-1]


def _login_role(client, user_id, email, role, status="ACTIVE"):
    """내부/외부 임의 역할 계정 upsert 후 dev-login 토큰."""
    db = models.SessionLocal()
    try:
        u = db.get(models.User, user_id)
        if u is None:
            u = models.User(user_id=user_id, email=email, name=email.split("@")[0])
            db.add(u)
        u.role = role
        u.status = status
        db.commit()
    finally:
        db.close()
    # dev-login은 내부 전용(외부는 매직링크 전용)이므로 테스트 토큰은 직접 발급
    from auth import create_access_token

    db2 = models.SessionLocal()
    try:
        u2 = db2.query(models.User).filter(models.User.email == email).first()
        assert u2 is not None, email
        return {"Authorization": "Bearer {0}".format(create_access_token(u2))}
    finally:
        db2.close()


# ── 1) 200·헤더·전체행·워터마크·합계 ────────────────────────────────────────
def test_export_ok_rows_match_list_total(client, manager_headers):
    cid = _mk_client(client, manager_headers, "OK")
    pid = _mk_project(client, manager_headers, "AVEXP자산OK", approval_status="승인")
    for i in range(3):
        _add_vehicle(
            client, manager_headers, pid,
            vehicle_no="AVEXPOK{0}".format(i), region="제주",
            client_id=cid, registered_at="2020-01-01", reduction_y1=10,
        )
    search = "AVEXPOK"
    r = client.get(EXPORT, headers=manager_headers, params={"search": search})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(_XLSX)
    assert "filename*=UTF-8''" in r.headers["content-disposition"]

    ws = _load_sheet(r.content)
    data = _data_rows(ws)
    # 데이터 행수 == 동일 필터 목록 total(전체행, 페이지네이션 없음)
    assert len(data) == _list_total(client, manager_headers, search=search) == 3

    # 워터마크에 내보낸 사람(이름) 포함 — conftest manager name('팀장')
    wm = ws.cell(row=1, column=1).value
    assert wm and "팀장" in wm

    # 합계행 존재(첫 컬럼 '합계')
    last = [c.value for c in list(ws.iter_rows())[-1] if c.value is not None]
    assert last and last[0] == "합계"


# ── 필터 반영(승인상태) ─────────────────────────────────────────────────────
def test_export_reflects_filters(client, manager_headers):
    cid = _mk_client(client, manager_headers, "FT")
    p_ok = _mk_project(client, manager_headers, "AVEXP필터FT승인", approval_status="승인")
    p_no = _mk_project(client, manager_headers, "AVEXP필터FT미승인", approval_status="미승인")
    _add_vehicle(client, manager_headers, p_ok, vehicle_no="AVEXPFT승인1",
                 region="서울", client_id=cid, registered_at="2020-01-01", reduction_y1=10)
    _add_vehicle(client, manager_headers, p_no, vehicle_no="AVEXPFT미승인1",
                 region="서울", client_id=cid, registered_at="2021-01-01", reduction_y1=5)

    p = {"search": "AVEXPFT", "approval_status": "승인"}
    r = client.get(EXPORT, headers=manager_headers, params=p)
    assert r.status_code == 200
    assert len(_data_rows(_load_sheet(r.content))) == _list_total(
        client, manager_headers, **p
    ) == 1


# ── 인가(조회보다 좁게) ─────────────────────────────────────────────────────
def test_export_authz(client, staff_headers):
    # STAFF 403(조회는 되지만 export는 MANAGER 하한)
    assert client.get(EXPORT, headers=staff_headers).status_code == 403
    # OBSERVER(ROLE_LEVEL 미등록 내부역할) 403
    obs = _login_role(client, "u-avexp-observer", "avexp-observer@hooxipartners.com", "OBSERVER")
    assert client.get(EXPORT, headers=obs).status_code == 403
    # 외부(PARTNER) 403 — get_current_user 원천 차단
    partner = _login_role(client, "u-avexp-partner", "avexp-partner@carrier.example", "PARTNER")
    assert client.get(EXPORT, headers=partner).status_code == 403
    # 미인증 401
    assert client.get(EXPORT).status_code == 401


# ── 감사(DATA_EXPORT, 금액 원문 미기록) ─────────────────────────────────────
def test_export_audit_no_secret(client, manager_headers):
    cid = _mk_client(client, manager_headers, "AU")
    pid = _mk_project(client, manager_headers, "AVEXP감사AUDIT", approval_status="승인")
    _add_vehicle(client, manager_headers, pid, vehicle_no="AVEXP감사AUDIT1",
                 region="서울", client_id=cid, registered_at="2020-01-01",
                 reduction_y1=10, expected_payout=1234567)
    search = "AVEXP감사AUDIT"
    r = client.get(EXPORT, headers=manager_headers, params={"search": search})
    assert r.status_code == 200

    db = models.SessionLocal()
    try:
        logs = (
            db.query(models.AuditLog)
            .filter(models.AuditLog.action == "DATA_EXPORT")
            .filter(models.AuditLog.new_value.like("%search={0}%".format(search)))
            .all()
        )
        assert len(logs) == 1
        log = logs[0]
        assert log.target_type == "ASSET_VEHICLES"
        assert log.new_value.startswith("rows=")
        # 금액 원문 미기록 — 예상지급액 등 금액 문자열이 감사값에 없어야
        assert "1234567" not in log.new_value
    finally:
        db.close()


# ── 행 상한(무음 잘라내기 금지) ─────────────────────────────────────────────
def test_export_row_cap(client, manager_headers, monkeypatch):
    cid = _mk_client(client, manager_headers, "CAP")
    pid = _mk_project(client, manager_headers, "AVEXP상한CAP", approval_status="승인")
    _add_vehicle(client, manager_headers, pid, vehicle_no="AVEXP상한CAP1",
                 region="서울", client_id=cid, registered_at="2020-01-01", reduction_y1=10)
    monkeypatch.setattr(av, "MAX_EXPORT_ROWS", 0)
    r = client.get(EXPORT, headers=manager_headers, params={"search": "AVEXP상한CAP"})
    assert r.status_code == 400, r.text


# ── 일일 반출 횟수 제한 ─────────────────────────────────────────────────────
def test_export_daily_limit(client, manager_headers, monkeypatch):
    monkeypatch.setattr(av, "DAILY_EXPORT_LIMIT", 0)
    r = client.get(EXPORT, headers=manager_headers, params={"search": "무관"})
    assert r.status_code == 429, r.text

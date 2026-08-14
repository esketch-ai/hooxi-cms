"""운수사 보유 차량(fleet) — CRUD·참여 구분·전역 엑셀 업로드·업체명 매칭 (부록 M)."""

import io
from urllib.parse import unquote

import openpyxl

API = "/api/v1"


def _mk_client(client, headers, name):
    r = client.post(
        API + "/clients", headers=headers,
        json={"client_type": "TRANSPORT", "company_name": name},
    )
    assert r.status_code == 201, r.text
    return r.json()["client_id"]


def _mk_project(client, headers, name):
    r = client.post(API + "/projects", headers=headers, json={"project_name": name, "project_status": "기획"})
    assert r.status_code == 201, r.text
    return r.json()["project_id"]


def test_vehicle_status_codes_seeded(client, staff_headers):
    r = client.get(API + "/codes", headers=staff_headers, params={"category": "VEHICLE_STATUS"})
    assert r.status_code == 200, r.text
    assert {"운행", "폐차"} <= {c["code"] for c in r.json()}


def test_fleet_crud_and_region_derive(client, staff_headers):
    cid = _mk_client(client, staff_headers, "보유차량운수")
    r = client.post(
        f"{API}/clients/{cid}/vehicles", headers=staff_headers,
        json={"vehicle_no": "서울70사1234", "chassis_no": "VINCRUD1", "model_name": "BS106", "model_year": 2016},
    )
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["region"] == "서울"  # 차량번호 앞 2글자 파생
    assert body["status"] == "운행"  # 기본값
    # 중복 차대번호 → 409 (식별키는 차대번호)
    dup = client.post(
        f"{API}/clients/{cid}/vehicles", headers=staff_headers,
        json={"vehicle_no": "서울70사9999", "chassis_no": "VINCRUD1"},
    )
    assert dup.status_code == 409, dup.text
    # 차량번호만 같고 차대번호 다르면 허용(내연/전기 공존)
    ok = client.post(
        f"{API}/clients/{cid}/vehicles", headers=staff_headers,
        json={"vehicle_no": "서울70사1234", "chassis_no": "VINCRUD2"},
    )
    assert ok.status_code == 201, ok.text
    lr = client.get(f"{API}/clients/{cid}/vehicles", headers=staff_headers).json()
    assert lr["total"] == 2
    assert lr["participating_count"] == 0
    assert lr["unassigned_count"] == 2
    assert lr["items"][0]["participation"] is False


def test_fleet_participation_link(client, staff_headers):
    """fleet 차량과 같은 차량번호로 프로젝트 참여차량 등록 시 참여로 링크·표시."""
    cid = _mk_client(client, staff_headers, "참여링크운수")
    client.post(
        f"{API}/clients/{cid}/vehicles", headers=staff_headers,
        json={"vehicle_no": "부산71바5678", "client_id": cid},
    )
    pid = _mk_project(client, staff_headers, "참여링크사업")
    client.post(
        f"{API}/projects/{pid}/vehicles", headers=staff_headers,
        json={"vehicle_no": "부산71바5678", "client_id": cid, "introduction_type": "신규도입", "reduction_y1": 10},
    )
    lr = client.get(f"{API}/clients/{cid}/vehicles", headers=staff_headers).json()
    assert lr["participating_count"] == 1
    assert lr["unassigned_count"] == 0
    item = lr["items"][0]
    assert item["participation"] is True
    assert item["project_name"] == "참여링크사업"
    assert item["introduction_type"] == "신규도입"

    # participation 필터
    part = client.get(f"{API}/clients/{cid}/vehicles", headers=staff_headers, params={"participation": "participating"}).json()
    assert part["total"] == 1
    none = client.get(f"{API}/clients/{cid}/vehicles", headers=staff_headers, params={"participation": "unassigned"}).json()
    assert none["total"] == 0


_FLEET_HEADERS = [
    "차량번호", "업체명", "차대번호", "차명", "연식", "차량등록일", "차종",
    "길이(mm)", "너비(mm)", "높이(mm)", "총중량(kg)", "승차정원", "연료",
]


def _fleet_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BUS_LIST_ALL"
    ws.append(_FLEET_HEADERS)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_fleet_import_and_client_match(client, staff_headers):
    cid = _mk_client(client, staff_headers, "임포트매칭운수")
    rows = [
        # 업체명이 등록 운수사와 일치 → client 매칭
        ["대구72사0001", "임포트매칭운수", "VIN0001", "BS110", 2018, "2018-03-01", "대형 승합", 10600, 2490, 3200, 13455, 49, "경유"],
        # 미등록 업체 → client_id None
        ["광주73아0002", "미등록운수", "VIN0002", "NEW", 2020, "2020-05-01", "대형 승합", 10600, 2490, 3200, 13455, 45, "전기"],
    ]
    buf = _fleet_xlsx(rows)
    r = client.post(
        f"{API}/fleet/import", headers=staff_headers,
        files={"file": ("bus.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    res = r.json()
    assert res["created"] == 2
    assert res["client_matched"] == 1  # 임포트매칭운수만

    lr = client.get(f"{API}/clients/{cid}/vehicles", headers=staff_headers).json()
    assert lr["total"] == 1  # 매칭된 1대만 이 운수사 소속
    v = lr["items"][0]
    assert v["vehicle_no"] == "대구72사0001"
    assert v["model_year"] == 2018
    assert v["region"] == "대구"


def test_fleet_reimport_preserves_status(client, staff_headers):
    """월간 재업로드가 수기 폐차 status를 덮어쓰지 않는다(reviewer MEDIUM)."""
    cid = _mk_client(client, staff_headers, "재업로드보존운수")
    v = client.post(
        f"{API}/clients/{cid}/vehicles", headers=staff_headers,
        json={"vehicle_no": "대전75자9999", "chassis_no": "VINX", "status": "폐차"},
    ).json()
    assert v["status"] == "폐차"
    # 같은 차량번호로 재업로드(파일엔 status 컬럼 없음) → 폐차 보존, 스펙은 갱신
    buf = _fleet_xlsx([
        ["대전75자9999", "재업로드보존운수", "VINX", "BS110", 2019, "2019-01-01", "대형 승합", 1, 1, 1, 1, 40, "경유"],
    ])
    r = client.post(
        f"{API}/fleet/import", headers=staff_headers,
        files={"file": ("bus.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()
    assert r["updated"] == 1
    lr = client.get(f"{API}/clients/{cid}/vehicles", headers=staff_headers).json()
    item = lr["items"][0]
    assert item["status"] == "폐차"  # 수기 폐차 보존
    assert item["model_name"] == "BS110"  # 스펙은 파일로 갱신


def test_fleet_import_derives_introduction(client, staff_headers):
    """도입구분 자동 판별 — 참여차량 차량번호가 내연 fleet에 있으면 대체도입, 없으면 신규도입.

    introduction_type이 비어있는(None) 참여차량만 자동설정, 수기값은 보존.
    """
    pid = _mk_project(client, staff_headers, "도입판별사업")
    # 참여차량 3대: intro 미지정 2대 + 수기 '신규도입' 1대(내연 fleet에 있어도 불변)
    client.post(
        f"{API}/projects/{pid}/vehicles", headers=staff_headers,
        json={"vehicle_no": "서울70사0001", "reduction_y1": 5},  # intro 미지정 → 대체도입 기대
    )
    client.post(
        f"{API}/projects/{pid}/vehicles", headers=staff_headers,
        json={"vehicle_no": "부산99바0002", "reduction_y1": 5},  # intro 미지정, fleet 없음 → 신규도입
    )
    client.post(
        f"{API}/projects/{pid}/vehicles", headers=staff_headers,
        json={"vehicle_no": "인천88가0003", "introduction_type": "신규도입", "reduction_y1": 5},  # 수기 → 불변
    )
    # 내연 fleet 임포트(전부 경유) — "서울70사0001", "인천88가0003"
    buf = _fleet_xlsx([
        ["서울70사0001", "판별운수", "VINICE", "BS110", 2017, "2017-01-01", "대형 승합", 1, 1, 1, 1, 40, "경유"],
        ["인천88가0003", "판별운수", "VINICE2", "BS110", 2017, "2017-01-01", "대형 승합", 1, 1, 1, 1, 40, "경유"],
    ])
    r = client.post(
        f"{API}/fleet/import", headers=staff_headers,
        files={"file": ("bus.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()
    assert r["introduction_derived"] == 2  # intro None인 2대만 자동설정

    lr = client.get(f"{API}/projects/{pid}/vehicles", headers=staff_headers).json()
    intro = {v["vehicle_no"]: v["introduction_type"] for v in lr["items"]}
    assert intro["서울70사0001"] == "대체도입"  # 내연 fleet에 있음
    assert intro["부산99바0002"] == "신규도입"  # fleet에 없음
    assert intro["인천88가0003"] == "신규도입"  # 수기값 보존(자동판별 미적용)


def test_create_client_vehicle_backlinks_participation(client, staff_headers):
    """참여차량 먼저 등록 후 fleet 수기 생성 시 역링크되어 참여로 표시(reviewer MEDIUM)."""
    cid = _mk_client(client, staff_headers, "역링크운수")
    pid = _mk_project(client, staff_headers, "역링크사업")
    # fleet 없는 상태로 참여차량 등록 → client_vehicle_id None
    client.post(
        f"{API}/projects/{pid}/vehicles", headers=staff_headers,
        json={"vehicle_no": "인천76차4321", "client_id": cid, "reduction_y1": 3},
    )
    # 이후 fleet 수기 생성 → 역링크
    client.post(
        f"{API}/clients/{cid}/vehicles", headers=staff_headers,
        json={"vehicle_no": "인천76차4321"},
    )
    lr = client.get(f"{API}/clients/{cid}/vehicles", headers=staff_headers).json()
    assert lr["participating_count"] == 1
    assert lr["items"][0]["participation"] is True


# ── D. 전국 버스 명부 양식/미리보기(dry-run) ─────────────────────────────────
def test_fleet_template_download(client, staff_headers):
    """양식 다운로드 — 시트명·헤더 순서·필수 * ·예시행·파일명 헤더."""
    r = client.get(f"{API}/fleet/template", headers=staff_headers)
    assert r.status_code == 200, r.text
    assert "전국버스명부" in unquote(r.headers["content-disposition"])
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames[0] == "BUS_LIST_ALL"
    ws = wb["BUS_LIST_ALL"]
    header = [c.value for c in ws[1]]
    # 헤더 라벨(순서) == _FLEET_HEADERS, 차량번호에만 '*'
    assert [h.rstrip(" *") for h in header] == _FLEET_HEADERS
    assert "*" in header[0]  # 차량번호(필수)
    assert all("*" not in h for h in header[1:])
    # 예시행 존재 + 차량번호 예시 접두
    example = [c.value for c in ws[2]]
    assert example[0] and example[0].startswith("(예시) ")


def test_fleet_preview_template_example_skipped(client, staff_headers):
    """다운로드 양식을 그대로 preview에 올리면 예시행은 건너뜀(실반영 예측 안 됨)."""
    tpl = client.get(f"{API}/fleet/template", headers=staff_headers).content
    r = client.post(
        f"{API}/fleet/preview", headers=staff_headers,
        files={"file": ("tpl.xlsx", io.BytesIO(tpl), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    res = r.json()
    assert res["created"] == 0 and res["updated"] == 0
    assert res["skipped"] == 1
    assert res["rows"][0]["classification"] == "건너뜀"
    assert res["rows"][0]["reason"] == "예시행"


def test_fleet_preview_no_db_mutation(client, staff_headers):
    """preview는 DB 무변경 — preview 후에도 매칭 소속 차량 수 불변(생성 안 됨)."""
    cid = _mk_client(client, staff_headers, "미리보기무변경운수")
    buf = _fleet_xlsx([
        ["서울70사5001", "미리보기무변경운수", "PV0001", "일렉시티", 2021, "2021-01-01", "대형 승합", 1, 1, 1, 1, 40, "전기"],
    ])
    before = client.get(f"{API}/clients/{cid}/vehicles", headers=staff_headers).json()["total"]
    r1 = client.post(
        f"{API}/fleet/preview", headers=staff_headers,
        files={"file": ("bus.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()
    assert r1["created"] == 1
    after = client.get(f"{API}/clients/{cid}/vehicles", headers=staff_headers).json()["total"]
    assert before == after == 0  # preview로 생성 안 됨


def test_fleet_preview_import_parity_with_dup(client, staff_headers):
    """파리티 — 파일 내 차대번호 중복행 포함, preview 카운트 == import 결과 카운트."""
    _mk_client(client, staff_headers, "파리티운수")
    rows = [
        ["서울70사6001", "파리티운수", "DUP001", "일렉시티", 2021, "2021-01-01", "대형 승합", 1, 1, 1, 1, 40, "전기"],
        ["서울70사6002", "미등록", "DUP002", "일렉시티", 2021, "2021-01-01", "대형 승합", 1, 1, 1, 1, 40, "전기"],
        # 차대번호 DUP001 중복행 → 파일 내에서 갱신으로 잡혀야(신규는 1건)
        ["서울70사6001", "파리티운수", "DUP001", "일렉시티", 2022, "2022-01-01", "대형 승합", 1, 1, 1, 1, 40, "전기"],
        ["", "누락업체", "", "차명", 2020, "", "", 1, 1, 1, 1, 40, "전기"],  # 차량번호 없음 → 건너뜀
    ]
    prev = client.post(
        f"{API}/fleet/preview", headers=staff_headers,
        files={"file": ("bus.xlsx", _fleet_xlsx(rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()
    imp = client.post(
        f"{API}/fleet/import", headers=staff_headers,
        files={"file": ("bus.xlsx", _fleet_xlsx(rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()
    assert prev["created"] == imp["created"]
    assert prev["updated"] == imp["updated"]
    assert prev["skipped"] == imp["skipped"]
    assert prev["client_matched"] == imp["client_matched"]
    # 중복행이 갱신으로 잡혀 신규 2 / 갱신 1 / 건너뜀 1
    assert prev["created"] == 2 and prev["updated"] == 1 and prev["skipped"] == 1


def test_fleet_preview_missing_vehicle_no_column_422(client, staff_headers):
    """차량번호 헤더 없는 파일 → preview 422(import와 동일 메시지)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BUS_LIST_ALL"
    ws.append(["업체명", "차대번호"])  # 차량번호 없음
    ws.append(["어떤운수", "X"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post(
        f"{API}/fleet/preview", headers=staff_headers,
        files={"file": ("bad.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 422
    assert "차량번호" in r.json()["detail"]

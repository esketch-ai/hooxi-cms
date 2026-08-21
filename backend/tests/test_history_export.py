"""활동 이력 엑셀 내보내기 — 필터 필수(전체 금지)·행 상한·STAFF 허용·감사 로그."""

from io import BytesIO

from openpyxl import load_workbook

import models


def _cleanup(db):
    db.query(models.ActivityHistory).filter(
        models.ActivityHistory.title.like("TESTHX%")).delete(synchronize_session=False)
    db.query(models.AuditLog).filter(
        models.AuditLog.action == "DATA_EXPORT",
        models.AuditLog.target_type == "HISTORY").delete(synchronize_session=False)
    db.query(models.Client).filter(
        models.Client.company_name.like("TESTHX%")).delete(synchronize_session=False)
    db.commit()


def test_export_requires_filter_and_downloads(client, staff_headers):
    db = models.SessionLocal()
    try:
        _cleanup(db)
        c = models.Client(client_type="TRANSPORT", company_name="TESTHX운수", region="서울")
        db.add(c); db.commit()
        cid = c.client_id
    finally:
        db.close()
    # 이력 2건 등록
    for i in range(2):
        r = client.post("/api/v1/histories", headers=staff_headers, json={
            "client_id": cid, "activity_date": f"2026-08-1{i} 10:00:00",
            "activity_type": "CALL", "title": f"TESTHX 통화 {i}",
        })
        assert r.status_code == 201, r.text
    # 무필터 → 400 (전체 다운로드 금지)
    r0 = client.get("/api/v1/histories/export", headers=staff_headers)
    assert r0.status_code == 400
    assert "필터" in r0.json()["detail"]
    # 고객사 필터 → STAFF도 다운로드 가능
    r1 = client.get(f"/api/v1/histories/export?client_id={cid}", headers=staff_headers)
    assert r1.status_code == 200, r1.text
    assert "spreadsheetml" in r1.headers["content-type"]
    wb = load_workbook(BytesIO(r1.content))
    ws = wb.active
    texts = [str(c2.value) for row in ws.iter_rows() for c2 in row if c2.value]
    assert any("TESTHX 통화" in t for t in texts)
    assert any("활동일" in t for t in texts)  # 헤더
    # 감사 로그 — 건수·필터 요약만
    db = models.SessionLocal()
    try:
        log = (db.query(models.AuditLog)
               .filter_by(action="DATA_EXPORT", target_type="HISTORY").first())
        assert log is not None and "client_id" in (log.new_value or "")
        _cleanup(db)
    finally:
        db.close()

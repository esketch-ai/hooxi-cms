"""차량 월별 운행·충전 로그(D6, P1·P2) — 취합본 WIDE 업로드·정리·집계."""

import io

import openpyxl

import models

IMPORT = "/api/v1/vehicle-logs/import"
CONSOL = "/api/v1/vehicle-logs/consolidate"
AGG = "/api/v1/vehicle-logs/aggregate"
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _wide_xlsx():
    """취합본 WIDE: 2개월×(운행일수·운행거리·충전량). SPARSE 차량 포함."""
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["운수사명", "자동차등록번호",
               "2025년05월_운행일수", "2025년05월_운행거리", "2025년05월_충전량",
               "2025년06월_운행일수", "2025년06월_운행거리", "2025년06월_충전량"])
    # 운행+충전 완전
    ws.append(["춘천시민버스", "강원70자1088", 30, 6000, 7000, 31, 6200, 7100])
    # 충전 결여(운행만)
    ws.append(["춘천시민버스", "강원70자2000", 28, 5000, None, 29, 5100, None])
    wb2 = io.BytesIO(); wb.save(wb2); return wb2.getvalue()


def _clean():
    db = models.SessionLocal()
    try:
        db.query(models.VehicleMonthlyLog).delete(synchronize_session=False)
        db.query(models.VehicleCalcInput).delete(synchronize_session=False)
        db.query(models.ReductionRegistry).delete(synchronize_session=False)
        db.commit()
    finally:
        db.close()


def test_import_consolidate_aggregate(client, staff_headers):
    _clean()
    f = {"file": ("wide.xlsx", _wide_xlsx(), XLSX)}
    r = client.post(IMPORT, headers=staff_headers, files=f)
    assert r.status_code == 200, r.text
    body = r.json()
    # (차량×월) 1행씩: 차량1 2행 + 차량2 2행 = 4행(지표는 컬럼)
    assert body["vehicles"] == 2 and body["months"] == 2
    assert body["created"] == 4 and body["updated"] == 0

    # 재업로드 → upsert(갱신)
    r2 = client.post(IMPORT, headers=staff_headers, files={"file": ("wide.xlsx", _wide_xlsx(), XLSX)})
    assert r2.json()["updated"] == 4 and r2.json()["created"] == 0

    # 자동 정리 뷰 — 충전 결여 차량 감지
    con = client.get(CONSOL, headers=staff_headers).json()
    assert con["vehicle_count"] == 2
    assert con["missing_charge"] == 1 and con["missing_run"] == 0
    by = {v["vehicle_no"]: v for v in con["vehicles"]}
    assert by["강원70자2000"]["has_charge"] is False
    assert by["강원70자1088"]["months"]["2025-05"]["distance_km"] == 6000

    # 집계 → 연평균, 프로그램 차량(레지스트리)만. commit 없이 계산만.
    db = models.SessionLocal()
    try:
        db.add(models.ReductionRegistry(role="PROJECT", vehicle_no="강원70자1088", vin="EV1"))
        db.add(models.VehicleCalcInput(vehicle_no="강원70자1088", fuel="CNG"))
        db.commit()
    finally:
        db.close()
    agg = client.post(AGG + "?commit_project=true", headers=staff_headers).json()
    assert agg["aggregated"] == 1  # 레지스트리 차량만
    assert agg["updated"] == 1
    item = next(i for i in agg["items"] if i["vehicle_no"] == "강원70자1088")
    # (6000+6200)/(30+31)*365 = 12200/61*365 = 73000
    assert abs(item["project_distance"] - 73000.0) < 0.5
    assert abs(item["project_kwh"] - (14100 / 61 * 365)) < 0.5

    # VehicleCalcInput 사업측 갱신 확인
    db = models.SessionLocal()
    try:
        ci = db.query(models.VehicleCalcInput).filter_by(vehicle_no="강원70자1088").first()
        assert ci.project_distance is not None and float(ci.project_distance) > 70000
    finally:
        db.close()
    _clean()


def _bms_xlsx():
    """BMS취합 LONG: 운수사·차량번호·월·운행거리합계·운행횟수합계."""
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["운수사", "차량번호", "월", "운행형태", "운행횟수합계", "운행거리합계", "집계행수", "파일수"])
    ws.append(["경기버스", "경기72바1278", "2025-07", "일반형시내버스", 85.8, 7236.26, 30, 1])
    ws.append(["경기버스", "경기72바1278", "2025-08", "일반형시내버스", 88.8, 7730.67, 30, 1])
    b = io.BytesIO(); wb.save(b); return b.getvalue()


def test_import_raw_bms(client, staff_headers):
    _clean()
    # 다건 업로드(여기선 BMS .xlsx 1건) → import-raw 자동판별
    files = [("files", ("BMS취합_x.xlsx", _bms_xlsx(), XLSX))]
    r = client.post("/api/v1/vehicle-logs/import-raw", headers=staff_headers, files=files)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["parsed_files"] == 1 and body["files"] == 1
    assert body["created"] == 2 and body["vehicles"] == 1 and body["months"] == 2

    # 로그가 BMS 출처로 적재됐는지 — consolidate에 운행거리 반영
    con = client.get(CONSOL, headers=staff_headers).json()
    v = next(x for x in con["vehicles"] if x["vehicle_no"] == "경기72바1278")
    assert v["months"]["2025-07"]["distance_km"] == 7236.26
    assert v["has_charge"] is False  # BMS엔 충전량 없음

    # 재업로드 → (차량·월·BMS) 중복키 upsert
    r2 = client.post("/api/v1/vehicle-logs/import-raw", headers=staff_headers,
                     files=[("files", ("BMS취합_x.xlsx", _bms_xlsx(), XLSX))])
    assert r2.json()["updated"] == 2 and r2.json()["created"] == 0
    _clean()


def test_import_raw_rejects_empty(client, staff_headers):
    # 빈/무의미 파일 → 파싱 0 → 422
    wb = openpyxl.Workbook(); b = io.BytesIO(); wb.save(b)
    r = client.post("/api/v1/vehicle-logs/import-raw", headers=staff_headers,
                    files=[("files", ("빈.xlsx", b.getvalue(), XLSX))])
    assert r.status_code == 422


def test_scan_dropbox(client, staff_headers, monkeypatch):
    """Dropbox 폴더 스캔 → .xlsx/.xls 자동판별 파싱 → 미리보기·적재."""
    import services.dropbox_storage as ds

    _clean()
    monkeypatch.setattr(ds, "is_configured", lambda: True)
    monkeypatch.setattr(ds, "root", lambda: "")
    monkeypatch.setattr(ds, "list_folder", lambda p: [
        {"name": "BMS취합.xlsx", "path_display": "/etas/BMS취합.xlsx", "is_dir": False},
        {"name": "메모.txt", "path_display": "/etas/메모.txt", "is_dir": False},  # 비대상 → 스킵
    ])
    monkeypatch.setattr(ds, "download", lambda p: _bms_xlsx() if p.endswith(".xlsx") else b"x")

    # 미리보기 — DB 무변경
    pv = client.get("/api/v1/vehicle-logs/scan-preview", headers=staff_headers, params={"folder": "/etas"})
    assert pv.status_code == 200, pv.text
    assert pv.json()["parsed_files"] == 1 and pv.json()["vehicles"] == 1 and pv.json()["total"] == 2
    assert client.get(CONSOL, headers=staff_headers).json()["vehicle_count"] == 0  # 아직 미적재

    # 적재
    cm = client.post("/api/v1/vehicle-logs/scan-commit", headers=staff_headers, params={"folder": "/etas"})
    assert cm.status_code == 200, cm.text
    assert cm.json()["created"] == 2 and cm.json()["parsed_files"] == 1
    assert client.get(CONSOL, headers=staff_headers).json()["vehicle_count"] == 1
    _clean()


def test_scan_requires_folder(client, staff_headers, monkeypatch):
    import services.dropbox_storage as ds
    monkeypatch.setattr(ds, "is_configured", lambda: True)
    r = client.get("/api/v1/vehicle-logs/scan-preview", headers=staff_headers)
    assert r.status_code == 422  # 폴더 미지정·config 기본값 없음


def test_requires_auth(client):
    assert client.get(CONSOL).status_code == 401

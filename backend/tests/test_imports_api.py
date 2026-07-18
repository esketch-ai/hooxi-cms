"""엑셀 일괄 등록 API — /imports/{entity}/spec·template·preview·commit.

- preview는 DB 무변경, commit은 유효 행만 부분 반영 + 감사 로그(건수 요약만).
- 규격 단일 원천 증명: 라벨 1개를 import_spec에서만 바꾸면(monkeypatch)
  양식 헤더·파서 매칭·spec API가 동시에 추종한다.
"""

import dataclasses
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

import models
from services import import_spec

XLSX_MT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, path, content, headers):
    return client.post(
        path, files={"file": ("upload.xlsx", content, XLSX_MT)}, headers=headers
    )


def _count(model):
    db = models.SessionLocal()
    try:
        return db.query(model).count()
    finally:
        db.close()


# ---------------------------------------------------------------------------
# spec / template
# ---------------------------------------------------------------------------
def test_spec_endpoint(client, admin_headers):
    resp = client.get("/api/v1/imports/clients/spec", headers=admin_headers)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["entity"] == "clients"
    assert body["max_rows"] == import_spec.MAX_IMPORT_ROWS
    labels = [c["label"] for c in body["columns"]]
    assert labels == [c.label for c in import_spec.IMPORT_SPECS["clients"].columns]
    required = {c["label"] for c in body["columns"] if c["required"]}
    assert {"회사명", "구분"} <= required
    # 고정값 컬럼(자산 인증 방식) 허용 표기 안내
    resp = client.get("/api/v1/imports/assets/spec", headers=admin_headers)
    auth_col = next(c for c in resp.json()["columns"] if c["field"] == "auth_type")
    assert "아이디/비밀번호" in auth_col["allowed_values"]


def test_template_endpoint(client, admin_headers):
    resp = client.get("/api/v1/imports/clients/template", headers=admin_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(XLSX_MT)
    # 한글 파일명 RFC 5987 인코딩 (documents 관용구)
    assert "filename*=UTF-8''" in resp.headers["content-disposition"]
    wb = load_workbook(BytesIO(resp.content))
    headers_row = [c.value for c in wb.active[1]]
    assert "회사명 *" in headers_row and "구분 *" in headers_row


def test_unknown_entity_404(client, admin_headers):
    for path in (
        "/api/v1/imports/nope/spec",
        "/api/v1/imports/nope/template",
    ):
        assert client.get(path, headers=admin_headers).status_code == 404
    resp = _upload(client, "/api/v1/imports/nope/preview", _xlsx(["회사명"], []), admin_headers)
    assert resp.status_code == 404


def test_requires_auth_401(client):
    assert client.get("/api/v1/imports/clients/spec").status_code == 401
    assert client.get("/api/v1/imports/clients/template").status_code == 401
    # 무권한 403 케이스 없음 — master.write는 전 활성 역할(STAFF 이상) 허용(단건 등록과 동일)


# ---------------------------------------------------------------------------
# preview — DB 무변경
# ---------------------------------------------------------------------------
def test_preview_returns_rows_and_keeps_db(client, staff_headers):
    before = _count(models.Client)
    content = _xlsx(
        ["회사명", "구분", "사업자번호"],
        [
            ["미리보기상사", "운수사", "321-99-11110"],
            ["오류행상사", "우주선", None],
        ],
    )
    resp = _upload(client, "/api/v1/imports/clients/preview", content, staff_headers)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["total_rows"] == 2
    assert body["valid_rows"] == 1
    assert body["error_rows"] == 1
    error_row = next(r for r in body["rows"] if r["status"] == "ERROR")
    assert error_row["row"] == 3 and "우주선" in error_row["errors"][0]
    assert _count(models.Client) == before  # DB 무변경


def test_preview_empty_file_422(client, admin_headers):
    resp = client.post(
        "/api/v1/imports/clients/preview",
        files={"file": ("empty.xlsx", b"", XLSX_MT)},
        headers=admin_headers,
    )
    assert resp.status_code == 422


# ---------------------------------------------------------------------------
# commit — 부분 반영 + preview 일관 + 감사 로그(건수 요약만)
# ---------------------------------------------------------------------------
def test_commit_partial_apply_and_audit(client, admin_headers):
    content = _xlsx(
        ["회사명", "구분", "사업자번호", "담당 PM", "월간 보고서 수신"],
        [
            ["커밋일호상사", "운수사", "808-11-22233", "관리자", "예"],
            ["커밋이호상사", "TRANSPORT", None, None, None],
            ["커밋오류상사", "우주선", None, None, None],
        ],
    )
    preview = _upload(client, "/api/v1/imports/clients/preview", content, admin_headers).json()

    before = _count(models.Client)
    resp = _upload(client, "/api/v1/imports/clients/commit", content, admin_headers)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    # preview ↔ commit 일관 (같은 파일 → 같은 판정)
    assert body["created"] == preview["valid_rows"] == 2
    assert body["skipped"] == preview["error_rows"] == 1
    assert body["errors"][0]["row"] == 4
    assert _count(models.Client) == before + 2

    db = models.SessionLocal()
    try:
        created = (
            db.query(models.Client)
            .filter(models.Client.company_name == "커밋일호상사")
            .one()
        )
        # 단건 등록과 동일 효과 — 코드·PM 해석·Y/N 반영, 구독은 미생성
        assert created.client_type == "TRANSPORT"
        assert created.manager_id == "u-admin"
        assert created.report_yn == "Y"
        assert (
            db.query(models.ReportSubscription)
            .filter(models.ReportSubscription.client_id == created.client_id)
            .count()
            == 0
        )
        # 감사 로그 — 건수 요약만, 행 원문(회사명 등) 기록 금지 (R2-E6)
        log = (
            db.query(models.AuditLog)
            .filter(models.AuditLog.action == "EXCEL_IMPORT")
            .order_by(models.AuditLog.created_at.desc())
            .first()
        )
        assert log is not None and log.target_type == "CLIENT"
        assert "생성 2건" in log.new_value and "건너뜀 1건" in log.new_value
        assert "커밋일호상사" not in log.new_value
    finally:
        db.close()


def test_commit_assets_without_credentials(client, admin_headers):
    db = models.SessionLocal()
    try:
        db.add(models.Client(client_type="TRANSPORT", company_name="자산커밋운수"))
        db.commit()
    finally:
        db.close()

    content = _xlsx(
        ["고객사명", "대분류", "소분류", "수량", "인증 방식", "로그인 ID"],
        [["자산커밋운수", "모빌리티", "전기차", 5, "아이디/비밀번호", "fleet01"]],
    )
    resp = _upload(client, "/api/v1/imports/assets/commit", content, admin_headers)
    assert resp.status_code == 200, resp.text
    assert resp.json()["created"] == 1

    db = models.SessionLocal()
    try:
        asset = (
            db.query(models.Asset).filter(models.Asset.login_id == "fleet01").one()
        )
        assert asset.asset_group == "MOBILITY"
        assert asset.quantity == 5
        assert asset.auth_type == "ID_PW"
        # 인증 비밀값 미유통 — 엑셀 경로로는 절대 저장되지 않는다
        assert asset.login_password is None and asset.api_token is None
    finally:
        db.close()


def test_commit_duplicate_within_file_partial(client, admin_headers):
    content = _xlsx(
        ["회사명", "구분", "사업자번호"],
        [
            ["부분반영일호", "운수사", "444-55-66601"],
            ["부분반영이호", "운수사", "444-55-66601"],  # 파일 내 중복 → 스킵
        ],
    )
    resp = _upload(client, "/api/v1/imports/clients/commit", content, admin_headers)
    body = resp.json()
    assert body["created"] == 1 and body["skipped"] == 1
    assert "파일 내 중복" in body["errors"][0]["errors"][0]


# ---------------------------------------------------------------------------
# 규격 변경 용이성 — import_spec.py 1곳만 바꾸면 전부 추종 (설계 제1원칙)
# ---------------------------------------------------------------------------
def test_spec_label_change_single_source(client, admin_headers, monkeypatch):
    """컬럼 라벨 1개 변경(회사명→법인명)이 template·파서·spec API에 동시 반영."""
    spec = import_spec.IMPORT_SPECS["clients"]
    new_cols = tuple(
        dataclasses.replace(c, label="법인명") if c.field == "company_name" else c
        for c in spec.columns
    )
    monkeypatch.setitem(
        import_spec.IMPORT_SPECS, "clients", dataclasses.replace(spec, columns=new_cols)
    )

    # 1) spec API 추종
    body = client.get("/api/v1/imports/clients/spec", headers=admin_headers).json()
    labels = [c["label"] for c in body["columns"]]
    assert "법인명" in labels and "회사명" not in labels

    # 2) 양식 헤더 추종
    resp = client.get("/api/v1/imports/clients/template", headers=admin_headers)
    headers_row = [c.value for c in load_workbook(BytesIO(resp.content)).active[1]]
    assert "법인명 *" in headers_row and "회사명 *" not in headers_row

    # 3) 파서 매칭 추종 — 새 라벨은 통과, 옛 라벨은 필수 컬럼 누락(422)
    ok = _upload(
        client,
        "/api/v1/imports/clients/preview",
        _xlsx(["법인명", "구분"], [["새라벨상사", "운수사"]]),
        admin_headers,
    )
    assert ok.status_code == 200 and ok.json()["valid_rows"] == 1
    stale = _upload(
        client,
        "/api/v1/imports/clients/preview",
        _xlsx(["회사명", "구분"], [["옛라벨상사", "운수사"]]),
        admin_headers,
    )
    assert stale.status_code == 422 and "법인명" in stale.json()["detail"]

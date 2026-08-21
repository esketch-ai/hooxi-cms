"""자산관리 보고(P2) — 운수사×사업 정산 요약 매트릭스 + 엑셀 내보내기.

검증: 매트릭스 롤업 정확성(운수사 행 == 사업 드릴다운 합)·None 전파(전건 None→None, 0 아님)·
NULL client_id '(미지정)' 행으로 Σ행==총계 정합·finance-ledger와 예상지급액 총계 정합(동일집합)·
인가(OBSERVER 화이트리스트·외부역할 격리·export MANAGER 게이트)·export 균형 보안(행상한 400·
일일한도 429·DATA_EXPORT 감사 금액원문 미기록).
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

import models
from routers import asset_report as ar

API = "/api/v1"
PROJECTS = API + "/projects"
SUMMARY = API + "/asset-report/settlement-summary"
EXPORT = SUMMARY + "/export"
LEDGER = API + "/finance-ledger"

_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── 시드 헬퍼 ────────────────────────────────────────────────────────────────
def _mk_client(client, headers, tag):
    r = client.post(API + "/clients", headers=headers,
                    json={"client_type": "TRANSPORT", "company_name": "정산운수" + tag})
    assert r.status_code == 201, r.text
    return r.json()["client_id"]


def _capped_vehicle(per_year, cid):
    """부록 L 상한 없이 감축량이 잡히는 차량 payload — client_id 지정(None 허용)."""
    p = {"registered_at": "2016-01-01"}
    if cid is not None:
        p["client_id"] = cid
    for i in range(1, 9):
        p["reduction_y{0}".format(i)] = per_year
    return p


def _mk_project(client, headers, name):
    r = client.post(PROJECTS, headers=headers,
                    json={"project_name": name, "project_status": "기획"})
    assert r.status_code == 201, r.text
    return r.json()["project_id"]


def _add_vehicle(client, headers, pid, per_year, cid):
    r = client.post(f"{PROJECTS}/{pid}/vehicles", headers=headers,
                    json=_capped_vehicle(per_year, cid))
    assert r.status_code == 201, r.text


def _set_payout(client, headers, pid):
    r = client.put(f"{PROJECTS}/{pid}/payout-params", headers=headers,
                   json={"max_payment": 1200000, "approved_at": "2016-02-01"})
    assert r.status_code == 200, r.text


def _set_client_attr(client_id, **attrs):
    db = models.SessionLocal()
    try:
        c = db.get(models.Client, client_id)
        for k, v in attrs.items():
            setattr(c, k, v)
        db.commit()
    finally:
        db.close()


def _login_role(client, user_id, email, role, status="ACTIVE"):
    db = models.SessionLocal()
    try:
        u = db.get(models.User, user_id)
        if u is None:
            u = models.User(user_id=user_id, email=email, name=email.split("@")[0])
            db.add(u)
        u.role = role
        u.status = status
        db.commit()
    finally:
        db.close()
    # dev-login은 내부 전용(외부는 매직링크 전용)이므로 테스트 토큰은 직접 발급
    from auth import create_access_token

    db2 = models.SessionLocal()
    try:
        u2 = db2.query(models.User).filter(models.User.email == email).first()
        assert u2 is not None, email
        return {"Authorization": "Bearer {0}".format(create_access_token(u2))}
    finally:
        db2.close()


def _sum_opt_db(values):
    """None 안전 합(테스트 기대치) — 전건 None이면 None, 아니면 non-null 합."""
    parts = [float(v) for v in values if v is not None]
    return round(sum(parts), 2) if parts else None


def _approx_eq(a, b):
    """None-or-근사 동등 — 정합 게이트용."""
    if a is None or b is None:
        return a is None and b is None
    return abs(a - b) < 0.5


# ── 1) 매트릭스 롤업 정확성 — 운수사 행 == 사업 드릴다운 합 ────────────────────
def test_matrix_rollup_accuracy(client, manager_headers):
    ca = _mk_client(client, manager_headers, "A")
    cb = _mk_client(client, manager_headers, "B")
    p1 = _mk_project(client, manager_headers, "정산사업P1")
    p2 = _mk_project(client, manager_headers, "정산사업P2")
    # A: P1에 2대 + P2에 1대 = 2사업 3대 / B: P1에 1대 = 1사업 1대
    _add_vehicle(client, manager_headers, p1, 30, ca)
    _add_vehicle(client, manager_headers, p1, 30, ca)
    _add_vehicle(client, manager_headers, p1, 30, cb)
    _add_vehicle(client, manager_headers, p2, 20, ca)
    _set_payout(client, manager_headers, p1)
    _set_payout(client, manager_headers, p2)

    r = client.get(SUMMARY, headers=manager_headers, params={"client_id": ca})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["total"] == 1  # A 한 운수사만
    row = body["items"][0]
    assert row["client_id"] == ca
    assert row["participating_project_count"] == 2
    assert row["participating_vehicle_count"] == 3
    # 사업 정렬(project_name) + 사업별 차량수
    names = [p["project_name"] for p in row["projects"]]
    assert names == sorted(names)
    vcount = {p["project_name"]: p["vehicle_count"] for p in row["projects"]}
    assert vcount["정산사업P1"] == 2 and vcount["정산사업P2"] == 1

    # 운수사 행 == 사업 드릴다운 None-안전 합
    assert _approx_eq(row["expected_payout"],
                      _sum_opt_db(p["expected_payout"] for p in row["projects"]))
    assert _approx_eq(row["total_reduction"],
                      _sum_opt_db(p["total_reduction"] for p in row["projects"]))
    # 필터 총계 == 단일 운수사 행
    assert _approx_eq(body["totals"]["expected_payout"], row["expected_payout"])
    assert body["totals"]["participating_vehicle_count"] == 3
    assert body["totals"]["distinct_project_count"] == 2

    # DB 저장값 직접 합과 일치(재계산 아님 — 저장 파생값 합산)
    db = models.SessionLocal()
    try:
        vs = db.query(models.ProjectVehicle).filter(
            models.ProjectVehicle.client_id == ca).all()
        assert _approx_eq(row["expected_payout"],
                          _sum_opt_db(v.expected_payout for v in vs))
    finally:
        db.close()


# ── 2) 정합 게이트 — finance-ledger 예상지급액 총계와 동일(동일집합, None-safe) ──
def test_consistency_with_finance_ledger(client, manager_headers):
    s = client.get(SUMMARY, headers=manager_headers).json()
    fl = client.get(LEDGER, headers=manager_headers,
                    params={"page_size": 1}).json()
    assert _approx_eq(s["totals"]["expected_payout"],
                      fl["totals"]["expected_payment"])

    # 총감축량·잔여반영감축량 총계 == 전차량 저장값 None-안전 합
    db = models.SessionLocal()
    try:
        vs = db.query(models.ProjectVehicle).all()
        assert _approx_eq(s["totals"]["total_reduction"],
                          _sum_opt_db(v.total_reduction for v in vs))
        assert _approx_eq(s["totals"]["effective_reduction"],
                          _sum_opt_db(v.effective_reduction for v in vs))
    finally:
        db.close()


# ── 3) None 전파 — 예상지급액 전건 None 운수사/사업은 None(0 아님) ────────────
def test_none_propagation(client, manager_headers):
    cn = _mk_client(client, manager_headers, "N")
    p = _mk_project(client, manager_headers, "정산미승인N")
    _add_vehicle(client, manager_headers, p, 30, cn)
    _add_vehicle(client, manager_headers, p, 30, cn)
    # payout-params 미설정 → expected_payout·effective_reduction 전건 None(total_reduction은 값)

    r = client.get(SUMMARY, headers=manager_headers, params={"client_id": cn})
    assert r.status_code == 200
    row = r.json()["items"][0]
    assert row["expected_payout"] is None  # 0이 아니라 None
    assert row["effective_reduction"] is None
    assert row["total_reduction"] is not None and row["total_reduction"] > 0
    proj = row["projects"][0]
    assert proj["expected_payout"] is None and proj["total_reduction"] > 0


# ── 4) NULL client_id → '(미지정)' 행, Σ행 == 총계 정합 ──────────────────────
def test_unassigned_row_and_totals_reconcile(client, manager_headers):
    p = _mk_project(client, manager_headers, "정산미지정U")
    _add_vehicle(client, manager_headers, p, 30, None)  # client_id 미지정
    _set_payout(client, manager_headers, p)

    body = client.get(SUMMARY, headers=manager_headers).json()
    unassigned = [i for i in body["items"] if i["client_id"] is None]
    assert len(unassigned) == 1
    assert unassigned[0]["company_name"] == "(미지정)"

    # Σ행 == 전사 총계(차량 단순합·예상지급액 None-안전 합)
    assert (sum(i["participating_vehicle_count"] for i in body["items"])
            == body["totals"]["participating_vehicle_count"])
    assert _approx_eq(_sum_opt_db(i["expected_payout"] for i in body["items"]),
                      body["totals"]["expected_payout"])
    # distinct project == 고유 사업수(운수사 합산 아님)
    pids = {p["project_id"] for i in body["items"] for p in i["projects"]}
    assert body["totals"]["distinct_project_count"] == len(pids)


# ── 5) 필터(client_type·region) ──────────────────────────────────────────────
def test_filters(client, manager_headers):
    cf = _mk_client(client, manager_headers, "F")
    p = _mk_project(client, manager_headers, "정산필터F")
    _add_vehicle(client, manager_headers, p, 30, cf)
    _set_payout(client, manager_headers, p)
    _set_client_attr(cf, region="RGNTEST")

    # region 필터 → 그 운수사만
    r = client.get(SUMMARY, headers=manager_headers, params={"region": "RGNTEST"})
    assert r.status_code == 200
    ids = [i["client_id"] for i in r.json()["items"]]
    assert ids == [cf]

    # client_type 필터(전달값 그대로) — TRANSPORT 포함, 없는 구분은 빈 결과
    r = client.get(SUMMARY, headers=manager_headers, params={"client_type": "TRANSPORT"})
    assert cf in [i["client_id"] for i in r.json()["items"]]
    r = client.get(SUMMARY, headers=manager_headers, params={"client_type": "NOPE"})
    assert r.json()["items"] == []


# ── 6) 인가 — OBSERVER 화이트리스트·외부역할 격리·내부 3역할 통과 ─────────────
def test_authz_summary(client, manager_headers, admin_headers, staff_headers):
    # 내부 3역할 200
    for h in (admin_headers, manager_headers, staff_headers):
        assert client.get(SUMMARY, headers=h).status_code == 200
    # OBSERVER 200(정확매칭 화이트리스트)
    obs = _login_role(client, "u-ss-observer", "ss-observer@hooxipartners.com", "OBSERVER")
    assert client.get(SUMMARY, headers=obs).status_code == 200
    # 정확매칭 — 존재하나 화이트리스트 밖인 하위 경로(/export)는 get_current_user에서 403
    # (settlement-summary만 허용, /export는 미포함 → OBSERVER 자연 차단)
    assert client.get(EXPORT, headers=obs).status_code == 403
    # 외부역할 원천 403
    partner = _login_role(client, "u-ss-partner", "ss-partner@carrier.example", "PARTNER")
    investor = _login_role(client, "u-ss-investor", "ss-investor@fund.example", "INVESTOR")
    assert client.get(SUMMARY, headers=partner).status_code == 403
    assert client.get(SUMMARY, headers=investor).status_code == 403
    # 미인증 401
    assert client.get(SUMMARY).status_code == 401


# ── 7) export — MANAGER 200(xlsx)·평탄화·합계행 / STAFF·OBSERVER 403 ──────────
def test_export_ok_and_authz(client, manager_headers, staff_headers):
    cx = _mk_client(client, manager_headers, "X")
    p1 = _mk_project(client, manager_headers, "정산반출X1")
    p2 = _mk_project(client, manager_headers, "정산반출X2")
    _add_vehicle(client, manager_headers, p1, 30, cx)
    _add_vehicle(client, manager_headers, p2, 20, cx)
    _set_payout(client, manager_headers, p1)
    _set_payout(client, manager_headers, p2)

    r = client.get(EXPORT, headers=manager_headers, params={"client_id": cx})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(_XLSX)
    assert "filename*=UTF-8''" in r.headers["content-disposition"]

    ws = load_workbook(BytesIO(r.content)).active
    rows = [[c.value for c in row] for row in ws.iter_rows()
            if any(c.value is not None for c in row)]
    # rows[0]=워터마크, rows[1]=헤더, 데이터, rows[-1]=합계
    data = rows[2:-1]
    assert len(data) == 2  # 운수사×사업 평탄화(P1·P2)
    wm = ws.cell(row=1, column=1).value
    assert wm and "팀장" in wm  # 워터마크 내보낸 사람
    assert rows[-1][0] == "합계"

    # STAFF·OBSERVER 403(조회보다 좁은 MANAGER 게이트 — export는 화이트리스트 미포함)
    assert client.get(EXPORT, headers=staff_headers).status_code == 403
    obs = _login_role(client, "u-ss-obs-exp", "ss-obs-exp@hooxipartners.com", "OBSERVER")
    assert client.get(EXPORT, headers=obs).status_code == 403


# ── 8) export 균형 보안 — 행상한 400·일일한도 429·DATA_EXPORT 감사(금액 미기록) ──
def test_export_row_cap(client, manager_headers, monkeypatch):
    cc = _mk_client(client, manager_headers, "CAP")
    p = _mk_project(client, manager_headers, "정산상한CAP")
    _add_vehicle(client, manager_headers, p, 30, cc)
    _set_payout(client, manager_headers, p)
    monkeypatch.setattr(ar, "MAX_EXPORT_ROWS", 0)
    r = client.get(EXPORT, headers=manager_headers, params={"client_id": cc})
    assert r.status_code == 400, r.text


def test_export_daily_limit(client, manager_headers, monkeypatch):
    monkeypatch.setattr(ar, "DAILY_EXPORT_LIMIT", 0)
    r = client.get(EXPORT, headers=manager_headers, params={"client_id": "무관"})
    assert r.status_code == 429, r.text


def test_export_audit_no_secret(client, manager_headers):
    ca = _mk_client(client, manager_headers, "AUD")
    p = _mk_project(client, manager_headers, "정산감사AUD")
    _add_vehicle(client, manager_headers, p, 30, ca)
    _set_payout(client, manager_headers, p)

    r = client.get(EXPORT, headers=manager_headers, params={"client_id": ca})
    assert r.status_code == 200
    payout = client.get(SUMMARY, headers=manager_headers,
                        params={"client_id": ca}).json()["items"][0]["expected_payout"]

    db = models.SessionLocal()
    try:
        logs = (db.query(models.AuditLog)
                .filter(models.AuditLog.action == "DATA_EXPORT")
                .filter(models.AuditLog.new_value.like("%client={0}%".format(ca)))
                .all())
        assert len(logs) == 1
        log = logs[0]
        assert log.target_type == "ASSET_REPORT"
        assert log.new_value.startswith("rows=")
        # 금액 원문 미기록(R2-E6) — 예상지급액 값이 감사문자열에 없어야
        assert str(int(payout)) not in log.new_value
    finally:
        db.close()
